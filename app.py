"""
eBay Hub UK v1.0.0
Flask web application for eBay UK pallet reselling.
Dark cyberpunk theme, mobile-first design.
"""

import os
import csv
import io
import json
import time
import secrets
import threading
import requests as http_requests
from datetime import datetime, timedelta

from flask import (
    Flask, request, redirect, url_for, flash,
    render_template_string, jsonify, g, session
)

from modules.database import (
    get_db, close_db, init_db,
    get_config, set_config,
    query_db, execute_db
)
from modules.ebay_api import (
    get_ebay_client,
    get_shipping_options_grouped,
    validate_shipping_fit,
)
from modules.scraper import scrape_amazon_product, parse_specification, get_amazon_image_url

# ---------------------------------------------------------------------------
# Flask app setup
# ---------------------------------------------------------------------------

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', secrets.token_hex(32))
app.permanent_session_lifetime = timedelta(days=7)

# Initialise database on first request
with app.app_context():
    init_db()


@app.teardown_appcontext
def teardown_db(exception):
    close_db()


@app.before_request
def before_request():
    """Attach CSP nonce and check authentication."""
    request._csp_nonce = secrets.token_hex(16)

    # Simple PIN/password auth
    if request.path.startswith('/static') or request.path == '/login':
        return
    if not session.get('authenticated'):
        pin = get_config('app_pin', '')
        if pin:  # PIN is set — require login
            return redirect('/login')


@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        entered = request.form.get('pin', '')
        stored = get_config('app_pin', '')
        if entered == stored:
            session['authenticated'] = True
            session.permanent = True
            flash('Welcome back!', 'success')
            return redirect('/')
        else:
            flash('Wrong PIN.', 'error')
    return render_template_string(LOGIN_TEMPLATE)


@app.route('/logout')
def logout():
    session.clear()
    return redirect('/login')


LOGIN_TEMPLATE = """<!DOCTYPE html>
<html><head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>Login - eBay Hub UK</title>
<link href="https://fonts.googleapis.com/css2?family=Space+Grotesk:wght@400;600;700&display=swap" rel="stylesheet">
<style>
*{margin:0;padding:0;box-sizing:border-box}
body{background:#0a0a14;color:#e2e8f0;font-family:'Space Grotesk',sans-serif;display:flex;align-items:center;justify-content:center;height:100vh}
.login-box{background:rgba(15,15,30,0.8);border:1px solid rgba(143,245,255,0.12);padding:40px;max-width:360px;width:90%;text-align:center}
h1{font-size:1.4rem;color:#8ff5ff;margin-bottom:8px}
p{color:#64748b;font-size:0.85rem;margin-bottom:24px}
input{width:100%;padding:14px;background:#0a0a14;border:1px solid rgba(143,245,255,0.15);color:#e2e8f0;font-size:1.2rem;text-align:center;letter-spacing:8px;font-family:'Space Grotesk',sans-serif;margin-bottom:16px}
input:focus{outline:none;border-color:#8ff5ff}
button{width:100%;padding:14px;background:#8ff5ff;color:#0a0a14;border:none;font-weight:700;font-size:1rem;cursor:pointer;font-family:'Space Grotesk',sans-serif}
button:hover{background:#beee00}
.flash{padding:10px;margin-bottom:16px;font-size:0.85rem;background:rgba(239,68,68,0.15);border:1px solid rgba(239,68,68,0.3);color:#ef4444}
</style></head><body>
<div class="login-box">
<h1>eBay Hub UK</h1>
<p>Enter your PIN to continue</p>
{% with messages = get_flashed_messages() %}{% if messages %}{% for m in messages %}<div class="flash">{{ m }}</div>{% endfor %}{% endif %}{% endwith %}
<form method="POST">
<input type="password" name="pin" placeholder="PIN" autofocus required>
<button type="submit">Unlock</button>
</form>
</div>
</body></html>"""


@app.after_request
def add_security_headers(response):
    """Minimal security headers — no strict CSP for simplicity."""
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['X-Frame-Options'] = 'DENY'
    return response


# ---------------------------------------------------------------------------
# Template helpers
# ---------------------------------------------------------------------------

def nonce():
    return getattr(request, '_csp_nonce', '')


def amazon_image(asin):
    """Build Amazon product image URL from ASIN."""
    if asin:
        return f"https://images-na.ssl-images-amazon.com/images/I/{asin}._AC_SL1500_.jpg"
    return ""


def fmt_gbp(val):
    """Format a number as GBP."""
    try:
        return f"\u00a3{float(val):,.2f}"
    except (TypeError, ValueError):
        return "\u00a30.00"


def fmt_date(val):
    """Format an ISO date string nicely."""
    if not val:
        return "-"
    try:
        dt = datetime.fromisoformat(str(val).replace('Z', '+00:00'))
        return dt.strftime("%d %b %Y")
    except Exception:
        return str(val)[:10] if val else "-"


def fmt_datetime(val):
    """Format an ISO datetime string nicely."""
    if not val:
        return "-"
    try:
        dt = datetime.fromisoformat(str(val).replace('Z', '+00:00'))
        return dt.strftime("%d %b %Y %H:%M")
    except Exception:
        return str(val)[:16] if val else "-"


def condition_label(cond):
    """Human-friendly condition label."""
    labels = {
        'new': 'New',
        'like_new': 'Like New',
        'used': 'Used',
        'damaged': 'Damaged'
    }
    return labels.get(cond, cond or 'Unknown')


def status_color(status):
    """CSS class for status badges."""
    colors = {
        'active': 'badge-cyan',
        'warehouse': 'badge-purple',
        'listed': 'badge-cyan',
        'sold': 'badge-lime',
        'shipped': 'badge-lime',
        'delivered': 'badge-lime',
        'draft': 'badge-muted',
        'ended': 'badge-muted',
        'archived': 'badge-muted',
        'new': 'badge-pink',
    }
    return colors.get(status, 'badge-muted')


# Register template context
@app.context_processor
def inject_helpers():
    return dict(
        nonce=nonce,
        amazon_image=amazon_image,
        fmt_gbp=fmt_gbp,
        fmt_date=fmt_date,
        fmt_datetime=fmt_datetime,
        condition_label=condition_label,
        status_color=status_color,
        now=datetime.utcnow
    )


# ---------------------------------------------------------------------------
# Dashboard stats helpers
# ---------------------------------------------------------------------------

def get_dashboard_stats():
    """Compute dashboard statistics."""
    today = datetime.utcnow().strftime('%Y-%m-%d')
    week_ago = (datetime.utcnow() - timedelta(days=7)).strftime('%Y-%m-%d')
    month_ago = (datetime.utcnow() - timedelta(days=30)).strftime('%Y-%m-%d')

    # Sales totals
    sales_today = query_db(
        "SELECT COALESCE(SUM(price_gbp), 0) as total, COUNT(*) as cnt "
        "FROM sales WHERE DATE(sold_at) = ?", (today,), one=True
    )
    sales_week = query_db(
        "SELECT COALESCE(SUM(price_gbp), 0) as total, COUNT(*) as cnt "
        "FROM sales WHERE DATE(sold_at) >= ?", (week_ago,), one=True
    )
    sales_month = query_db(
        "SELECT COALESCE(SUM(price_gbp), 0) as total, COUNT(*) as cnt "
        "FROM sales WHERE DATE(sold_at) >= ?", (month_ago,), one=True
    )
    sales_all = query_db(
        "SELECT COALESCE(SUM(price_gbp), 0) as total, COUNT(*) as cnt FROM sales",
        one=True
    )

    # Counts
    active_listings = query_db(
        "SELECT COUNT(*) as cnt FROM ebay_listings WHERE status = 'active'", one=True
    )['cnt']
    to_ship = query_db(
        "SELECT COUNT(*) as cnt FROM sales WHERE status = 'new'", one=True
    )['cnt']
    total_products = query_db(
        "SELECT COUNT(*) as cnt FROM products", one=True
    )['cnt']
    total_pallets = query_db(
        "SELECT COUNT(*) as cnt FROM pallets", one=True
    )['cnt']
    warehouse_products = query_db(
        "SELECT COUNT(*) as cnt FROM products WHERE status = 'warehouse'", one=True
    )['cnt']

    # Frozen capital: cost of pallets with unsold products
    frozen = query_db(
        "SELECT COALESCE(SUM(p.purchase_price_gbp), 0) as total "
        "FROM pallets p WHERE p.status = 'active'",
        one=True
    )['total']

    # Total cost for profit calculation
    total_cost = query_db(
        "SELECT COALESCE(SUM(purchase_price_gbp), 0) as total FROM pallets",
        one=True
    )['total']

    total_revenue = sales_all['total']
    total_profit = total_revenue - total_cost

    # Monthly revenue data (last 6 months) for chart
    chart_data = []
    for i in range(5, -1, -1):
        d = datetime.utcnow() - timedelta(days=30 * i)
        month_start = d.replace(day=1).strftime('%Y-%m-%d')
        if i > 0:
            next_d = datetime.utcnow() - timedelta(days=30 * (i - 1))
            month_end = next_d.replace(day=1).strftime('%Y-%m-%d')
        else:
            month_end = '2099-12-31'
        row = query_db(
            "SELECT COALESCE(SUM(price_gbp), 0) as rev FROM sales "
            "WHERE DATE(sold_at) >= ? AND DATE(sold_at) < ?",
            (month_start, month_end), one=True
        )
        chart_data.append({
            'label': d.strftime('%b %Y'),
            'revenue': round(row['rev'], 2)
        })

    return {
        'sales_today': sales_today,
        'sales_week': sales_week,
        'sales_month': sales_month,
        'sales_all': sales_all,
        'active_listings': active_listings,
        'to_ship': to_ship,
        'total_products': total_products,
        'total_pallets': total_pallets,
        'warehouse_products': warehouse_products,
        'frozen_capital': frozen,
        'total_revenue': total_revenue,
        'total_cost': total_cost,
        'total_profit': total_profit,
        'chart_data': chart_data,
    }


# ===================================================================
# POST-only routes (no template rendering)
# ===================================================================

@app.route('/pallets/add', methods=['POST'])
def pallet_add():
    name = request.form.get('name', '').strip()
    supplier = request.form.get('supplier', '').strip()
    price = request.form.get('purchase_price_gbp', '0')
    date = request.form.get('purchase_date', '')
    notes = request.form.get('notes', '').strip()
    amazon_domain = request.form.get('amazon_domain', '').strip()
    if not name:
        flash('Pallet name is required.', 'error')
        return redirect(url_for('pallets_list'))

    # amazon_domain is an OPTIONAL override. Empty = scraper auto-detects
    # which locale the ASIN lives on. Only reject unknown non-empty values
    # so a typo doesn't silently hit a bogus host.
    from modules.scraper import AMAZON_DOMAINS
    valid_domains = {d for d, _ in AMAZON_DOMAINS}
    if amazon_domain and amazon_domain not in valid_domains:
        amazon_domain = ''  # fall through to auto-detect

    try:
        price = float(price)
    except ValueError:
        price = 0.0

    # Dedup: if an active pallet with the same name+supplier already exists,
    # reuse it instead of creating a duplicate. The uncle was re-uploading
    # the same pallet CSV multiple times (partial uploads, retries) and
    # ending up with 6x "#3 MIX / Jobalots / £300" rows cluttering the list.
    # Match is case-insensitive and trims whitespace so "#3 MIX" == "#3 mix ".
    existing = query_db(
        "SELECT id FROM pallets "
        "WHERE LOWER(TRIM(name)) = LOWER(TRIM(?)) "
        "AND LOWER(TRIM(COALESCE(supplier,''))) = LOWER(TRIM(COALESCE(?,''))) "
        "AND status = 'active' "
        "ORDER BY id DESC LIMIT 1",
        (name, supplier), one=True
    )
    merged_into_existing = False
    if existing:
        pallet_id = existing['id']
        merged_into_existing = True
        # Refresh the other fields in case they changed (notes, price, domain override)
        if amazon_domain or notes or price:
            execute_db(
                "UPDATE pallets SET "
                "purchase_price_gbp = CASE WHEN ? > 0 THEN ? ELSE purchase_price_gbp END, "
                "purchase_date = CASE WHEN ? <> '' THEN ? ELSE purchase_date END, "
                "notes = CASE WHEN ? <> '' THEN ? ELSE notes END, "
                "amazon_domain = CASE WHEN ? <> '' THEN ? ELSE amazon_domain END "
                "WHERE id = ?",
                (price, price, date, date, notes, notes, amazon_domain, amazon_domain, pallet_id)
            )
    else:
        pallet_id = execute_db(
            "INSERT INTO pallets (name, supplier, purchase_price_gbp, purchase_date, notes, amazon_domain) "
            "VALUES (?, ?, ?, ?, ?, ?)",
            (name, supplier, price, date, notes, amazon_domain)
        )

    # Import products from uploaded CSV/XLSX file
    imported = 0
    scraped_cnt = 0
    spec_file = request.files.get('spec_file')
    if spec_file and spec_file.filename:
        from modules.scraper import scrape_amazon_product, get_amazon_image_url
        fname = spec_file.filename.lower()
        try:
            rows = []
            if fname.endswith(('.xlsx', '.xls')):
                import openpyxl
                wb = openpyxl.load_workbook(spec_file, data_only=True)
                ws = wb.active
                headers = [str(c.value or '').strip().lower() for c in ws[1]]
                for row_cells in ws.iter_rows(min_row=2, values_only=True):
                    row = {}
                    for i, val in enumerate(row_cells):
                        if i < len(headers):
                            row[headers[i]] = str(val or '').strip()
                    rows.append(row)
            elif fname.endswith('.csv'):
                raw = spec_file.stream.read()
                for enc in ('utf-8-sig', 'utf-8', 'latin-1', 'cp1252'):
                    try:
                        text = raw.decode(enc)
                        break
                    except:
                        continue
                else:
                    text = raw.decode('utf-8', errors='replace')
                first_line = text.split('\n')[0]
                delimiter = ';' if first_line.count(';') > first_line.count(',') else ','
                reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
                rows = list(reader)

            def get_col(row, *names):
                # Exact match first, then substring
                for n in names:
                    for key in row:
                        if key and key.lower().strip() == n:
                            val = row[key].strip() if row[key] else ''
                            if val and val.lower() not in ('none', 'nan', 'null'):
                                return val
                for n in names:
                    for key in row:
                        if key and n in key.lower() and 'category' not in key.lower():
                            val = row[key].strip() if row[key] else ''
                            if val and val.lower() not in ('none', 'nan', 'null'):
                                return val
                return ''

            for row in rows:
                prod_name = get_col(row, 'name', 'title', 'product', 'nazwa')
                # Supplier description — ground truth about what's physically in the
                # pallet. Kept separate from the name so the name matches Amazon title
                # (for matching/dedup) while the description drives AI generation.
                supplier_desc = get_col(row, 'product description', 'description', 'opis')
                # If neither name nor title matched but we have a description, fall back
                # to it for the name so we still have something.
                if not prod_name:
                    prod_name = supplier_desc
                if not prod_name:
                    continue
                asin = get_col(row, 'asin').upper()
                ean = get_col(row, 'ean', 'barcode', 'upc', 'gtin')
                try:
                    qty = int(float(get_col(row, 'quantity', 'qty', 'ilosc', 'amount') or '1'))
                except:
                    qty = 1
                cond = get_col(row, 'condition', 'state', 'stan').lower()
                if cond not in ('new', 'like_new', 'used', 'damaged'):
                    cond = 'new'
                try:
                    ebay_price = float(get_col(row, 'price', 'ebay_price', 'rrp', 'cena') or '0')
                except:
                    ebay_price = 0.0

                image_url = get_amazon_image_url(asin) if asin else ''

                # Scrape from override locale if set, else auto-detect. First
                # successful auto-detect wins — we cache it onto the pallet
                # below so subsequent scrapes skip the cascade.
                if asin:
                    try:
                        data = scrape_amazon_product(asin, amazon_domain or None)
                        if data:
                            if data.get('title') and len(data['title']) > len(prod_name):
                                prod_name = data['title']
                            if data.get('image_url'):
                                image_url = data['image_url']
                            if data.get('price') and ebay_price == 0:
                                ebay_price = data['price']
                            # Cache detected domain on the pallet on first hit
                            if not amazon_domain and data.get('source_domain'):
                                amazon_domain = data['source_domain']
                                execute_db(
                                    "UPDATE pallets SET amazon_domain = ? WHERE id = ?",
                                    (amazon_domain, pallet_id)
                                )
                            scraped_cnt += 1
                    except:
                        pass

                execute_db(
                    "INSERT INTO products (pallet_id, name, asin, ean, quantity, "
                    "condition, ebay_price_gbp, image_url, supplier_description) "
                    "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
                    (pallet_id, prod_name, asin, ean, qty, cond, ebay_price, image_url,
                     supplier_desc)
                )
                imported += 1
        except Exception as e:
            flash(f'File import error: {e}', 'error')

    # Run auto-pipeline after import (scrape all images, AI titles/descriptions, create drafts)
    pipeline_msg = ''
    if imported > 0:
        try:
            processed, drafts = auto_process_products(pallet_id)
            pipeline_msg = f' Auto-pipeline: {processed} scraped, {drafts} drafts created.'
        except Exception as e:
            pipeline_msg = f' Auto-pipeline error: {e}'

    if merged_into_existing:
        msg = f'Pallet "{name}" already existed — merged into existing pallet (no duplicate created).'
    else:
        msg = f'Pallet "{name}" added successfully.'
    if imported > 0:
        msg += f' Imported {imported} products'
        if scraped_cnt > 0:
            where = amazon_domain or 'Amazon (auto-detected)'
            msg += f' (scraped {scraped_cnt} from {where})'
        msg += '.'
    msg += pipeline_msg
    flash(msg, 'success')
    return redirect(url_for('pallet_detail', pallet_id=pallet_id) if (imported > 0 or merged_into_existing) else url_for('pallets_list'))


@app.route('/pallet/<int:pallet_id>/delete', methods=['POST'])
def pallet_delete(pallet_id):
    pallet = query_db("SELECT * FROM pallets WHERE id = ?", (pallet_id,), one=True)
    if not pallet:
        flash('Pallet not found.', 'error')
        return redirect(url_for('pallets_list'))
    execute_db("DELETE FROM products WHERE pallet_id = ?", (pallet_id,))
    execute_db("DELETE FROM pallets WHERE id = ?", (pallet_id,))
    flash(f'Pallet "{pallet["name"]}" deleted.', 'success')
    return redirect(url_for('pallets_list'))


@app.route('/pallets/bulk-delete', methods=['POST'])
def pallets_bulk_delete():
    """Bulk-delete selected pallets + their products. JSON API, called from
    the 'DELETE SELECTED' bar on /pallets (mirrors Akces Hub's pattern)."""
    try:
        data = request.get_json(silent=True) or {}
        ids = data.get('ids', []) or []
        if not ids:
            return jsonify({'ok': False, 'error': 'No pallets selected'}), 400

        deleted = 0
        for pid in ids:
            try:
                pid_int = int(pid)
            except (TypeError, ValueError):
                continue
            execute_db("DELETE FROM products WHERE pallet_id = ?", (pid_int,))
            execute_db("DELETE FROM pallets WHERE id = ?", (pid_int,))
            deleted += 1
        return jsonify({'ok': True, 'deleted': deleted})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)[:200]}), 500


@app.route('/pallets/merge-duplicates', methods=['POST'])
def pallets_merge_duplicates():
    """Consolidate pallets that share the same (name, supplier) into a single
    pallet. Keeps the oldest one as canonical, moves all products into it,
    and deletes the duplicates. Matching is case/whitespace insensitive so
    "#3 MIX" and "#3 mix " collapse together."""
    rows = query_db("""
        SELECT id, LOWER(TRIM(name)) AS nkey,
               LOWER(TRIM(COALESCE(supplier,''))) AS skey
        FROM pallets
        WHERE status = 'active'
        ORDER BY id ASC
    """)
    # Group by (nkey, skey) → list of ids (ascending)
    groups = {}
    for r in rows:
        key = (r['nkey'], r['skey'])
        groups.setdefault(key, []).append(r['id'])

    merged_groups = 0
    merged_pallets = 0
    moved_products = 0
    for key, ids in groups.items():
        if len(ids) < 2:
            continue
        canonical = ids[0]
        dupes = ids[1:]
        # Move products and listings then drop the duplicate pallets
        for dup in dupes:
            # Count before moving for the flash message
            cnt = query_db("SELECT COUNT(*) AS c FROM products WHERE pallet_id = ?",
                           (dup,), one=True)
            moved_products += (cnt['c'] if cnt else 0)
            execute_db("UPDATE products SET pallet_id = ? WHERE pallet_id = ?",
                       (canonical, dup))
            execute_db("DELETE FROM pallets WHERE id = ?", (dup,))
            merged_pallets += 1
        merged_groups += 1

    if merged_groups == 0:
        flash('No duplicate pallets found. Everything is already unique.', 'info')
    else:
        flash(
            f'Merged {merged_pallets} duplicate pallet(s) into {merged_groups} canonical one(s). '
            f'{moved_products} products moved.',
            'success'
        )
    return redirect(url_for('pallets_list'))


@app.route('/pallet/<int:pallet_id>/edit', methods=['POST'])
def pallet_edit(pallet_id):
    """Edit the pallet's metadata (name, supplier, purchase price, date, notes).
    Before this existed the only way to fix a wrong price was to delete the
    pallet + products and re-upload the CSV, which was brutal."""
    pallet = query_db("SELECT id FROM pallets WHERE id = ?", (pallet_id,), one=True)
    if not pallet:
        flash('Pallet not found.', 'error')
        return redirect(url_for('pallets_list'))

    name = request.form.get('name', '').strip()
    supplier = request.form.get('supplier', '').strip()
    price_raw = request.form.get('purchase_price_gbp', '0')
    date = request.form.get('purchase_date', '').strip()
    notes = request.form.get('notes', '').strip()

    if not name:
        flash('Pallet name is required.', 'error')
        return redirect(url_for('pallet_detail', pallet_id=pallet_id))

    try:
        price = float(price_raw)
    except (ValueError, TypeError):
        price = 0.0

    execute_db(
        "UPDATE pallets SET name = ?, supplier = ?, purchase_price_gbp = ?, "
        "purchase_date = ?, notes = ? WHERE id = ?",
        (name, supplier, price, date, notes, pallet_id)
    )
    flash('Pallet updated.', 'success')
    return redirect(url_for('pallet_detail', pallet_id=pallet_id))


@app.route('/pallet/<int:pallet_id>/archive', methods=['POST'])
def pallet_archive(pallet_id):
    execute_db("UPDATE pallets SET status = 'archived' WHERE id = ?", (pallet_id,))
    flash('Pallet archived.', 'success')
    return redirect(url_for('pallet_detail', pallet_id=pallet_id))


@app.route('/pallet/<int:pallet_id>/set-domain', methods=['POST'])
def pallet_set_domain(pallet_id):
    """Update (or clear) the Amazon locale override for this pallet.
    Empty string = fall back to auto-detect on next scrape."""
    pallet = query_db("SELECT id FROM pallets WHERE id = ?", (pallet_id,), one=True)
    if not pallet:
        flash('Pallet not found.', 'error')
        return redirect(url_for('pallets_list'))

    from modules.scraper import AMAZON_DOMAINS
    valid_domains = {d for d, _ in AMAZON_DOMAINS}
    new_domain = request.form.get('amazon_domain', '').strip()

    # Empty means "clear the override → auto-detect on next scrape"
    if new_domain and new_domain not in valid_domains:
        flash('Unknown Amazon locale.', 'error')
        return redirect(url_for('pallet_detail', pallet_id=pallet_id))

    execute_db("UPDATE pallets SET amazon_domain = ? WHERE id = ?", (new_domain, pallet_id))
    if new_domain:
        flash(f'Locale override set to {new_domain}. Re-scrape to refresh.', 'success')
    else:
        flash('Locale override cleared — scraper will auto-detect on next run.', 'success')
    return redirect(url_for('pallet_detail', pallet_id=pallet_id))


# ---------------------------------------------------------------------------
# Background scrape jobs
#
# Why: scraping N products × up to 9 Amazon locales (auto-detect cascade) is
# 30-300s of blocking work. Cloudflare times out HTTP at 100s (error 524),
# so we cannot do this synchronously in a request handler. We also don't
# want to add a task queue dependency (Celery/RQ) on a Raspberry Pi.
#
# Solution: fire a daemon thread per scrape, track progress in an in-memory
# dict keyed by pallet_id, and expose /pallet/<id>/scrape-status for the UI
# to poll. Progress is best-effort and resets on app restart — that's fine
# because all writes go straight to SQLite, so even if the process dies
# mid-scrape the partial work is preserved.
# ---------------------------------------------------------------------------

_scrape_jobs = {}  # pallet_id -> {status, total, done, updated, where, error, started_at}
_scrape_lock = threading.Lock()


def _run_pallet_scrape(pallet_id, override):
    """Worker body that runs in a background thread.
    IMPORTANT: this runs OUTSIDE any Flask request context, so it must
    only use modules.database helpers (which are thread-safe via
    thread-local connections) and never touch request/session/g."""
    from modules.scraper import get_amazon_image_urls
    import requests as _req

    try:
        # Skip products that already have an image. The uncle asked for this
        # explicitly — once a product is scraped successfully there's no
        # point hitting Amazon again (wastes time and bumps into CAPTCHA
        # faster). A product counts as "already has image" if image_url is
        # set OR its images JSON gallery has at least one entry. If the user
        # wants to force a re-scrape they can clear image_url from the
        # product edit page first.
        products = query_db(
            "SELECT * FROM products "
            "WHERE pallet_id = ? AND asin != '' "
            "  AND (image_url IS NULL OR image_url = '') "
            "  AND (images IS NULL OR images = '' OR images = '[]')",
            (pallet_id,)
        )
        total_pending = len(products)

        # Include already-scraped count in the progress denominator so the
        # bar reads "47 / 50" instead of "3 / 3" for a mostly-done pallet.
        already = query_db(
            "SELECT COUNT(*) AS c FROM products "
            "WHERE pallet_id = ? AND asin != '' "
            "  AND ((image_url IS NOT NULL AND image_url <> '') "
            "       OR (images IS NOT NULL AND images <> '' AND images <> '[]'))",
            (pallet_id,), one=True
        )
        already_cnt = (already['c'] if already else 0)
        total_display = total_pending + already_cnt

        with _scrape_lock:
            _scrape_jobs[pallet_id]['total'] = total_display
            # Pre-credit the already-scraped products so the progress bar
            # starts at the right position and the "updated" count includes
            # what was there before.
            _scrape_jobs[pallet_id]['done'] = already_cnt

        if total_pending == 0:
            with _scrape_lock:
                _scrape_jobs[pallet_id].update(
                    status='done', where='', updated=already_cnt
                )
            return

        updated = 0
        detected_domain = None
        # effective_override: starts as override; once we auto-detect the
        # locale on product #1, reuse it for products 2..N so the cascade
        # only runs once per pallet instead of N times.
        effective_override = override

        for i, prod in enumerate(products):
            try:
                amz_data = scrape_amazon_product(prod['asin'], effective_override)
            except Exception:
                amz_data = None

            if amz_data and amz_data.get('image_url'):
                new_name = amz_data.get('title') or prod['name']
                new_image = amz_data['image_url']
                new_price = amz_data.get('price') or prod['ebay_price_gbp']
                images_json = json.dumps(amz_data.get('all_images', []))
                specs_json = json.dumps(amz_data.get('item_specifics', {}))
                execute_db(
                    "UPDATE products SET name = ?, image_url = ?, ebay_price_gbp = ?, images = ?, item_specifics = ? WHERE id = ?",
                    (new_name, new_image, new_price, images_json, specs_json, prod['id'])
                )
                updated += 1
                if not detected_domain:
                    detected_domain = amz_data.get('source_domain')
                    # Re-use the detected locale for remaining products
                    if not effective_override and detected_domain:
                        effective_override = detected_domain
            else:
                # Fallback: try static media.amazon.com image URLs (domain-agnostic)
                for url in get_amazon_image_urls(prod['asin']):
                    try:
                        r = _req.head(url, timeout=5, allow_redirects=True)
                        if r.status_code == 200 and 'image' in r.headers.get('content-type', ''):
                            execute_db("UPDATE products SET image_url = ? WHERE id = ?", (url, prod['id']))
                            updated += 1
                            break
                    except Exception:
                        continue

            with _scrape_lock:
                # done counts both pre-existing and newly-scraped so the
                # progress bar stays accurate when we skipped products.
                _scrape_jobs[pallet_id].update(
                    done=already_cnt + i + 1,
                    updated=already_cnt + updated,
                )

        # Cache first auto-detected locale on the pallet so future scrapes skip the cascade
        if not override and detected_domain:
            execute_db("UPDATE pallets SET amazon_domain = ? WHERE id = ?", (detected_domain, pallet_id))

        with _scrape_lock:
            _scrape_jobs[pallet_id].update(
                status='done',
                where=(override or detected_domain or 'Amazon'),
                updated=already_cnt + updated,
            )
    except Exception as e:
        with _scrape_lock:
            _scrape_jobs[pallet_id].update(status='error', error=str(e)[:300])


@app.route('/pallet/<int:pallet_id>/scrape', methods=['POST'])
def pallet_scrape(pallet_id):
    """Start a background scrape for all products with ASINs in this pallet.
    Returns immediately with a flash message; the actual work runs in a
    daemon thread. The UI can poll /pallet/<id>/scrape-status for progress.

    This avoids Cloudflare 524 timeouts — the origin request finishes in
    milliseconds, not minutes."""
    pallet = query_db("SELECT * FROM pallets WHERE id = ?", (pallet_id,), one=True)
    if not pallet:
        flash('Pallet not found.', 'error')
        return redirect(url_for('pallets_list'))

    # Empty -> auto-detect; set -> strict override
    override = (pallet.get('amazon_domain') or '').strip() or None

    # Refuse to queue a second scrape while one is already running
    with _scrape_lock:
        existing = _scrape_jobs.get(pallet_id)
        if existing and existing.get('status') == 'running':
            flash('A scrape is already running for this pallet — check back in a moment.', 'info')
            return redirect(url_for('pallet_detail', pallet_id=pallet_id))
        _scrape_jobs[pallet_id] = {
            'status': 'running',
            'total': 0,
            'done': 0,
            'updated': 0,
            'where': '',
            'error': '',
            'started_at': time.time(),
        }

    t = threading.Thread(
        target=_run_pallet_scrape,
        args=(pallet_id, override),
        daemon=True,
    )
    t.start()

    flash('Scrape started in the background. Progress will update on this page — no need to wait.', 'success')
    return redirect(url_for('pallet_detail', pallet_id=pallet_id))


@app.route('/pallet/<int:pallet_id>/scrape-status')
def pallet_scrape_status(pallet_id):
    """JSON status of the background scrape job for this pallet.
    Returns {status, total, done, updated, where, error}. Frontend polls this
    to show a progress bar without blocking the page."""
    with _scrape_lock:
        job = _scrape_jobs.get(pallet_id)
        if not job:
            return jsonify({'status': 'idle'})
        return jsonify({
            'status': job.get('status', 'idle'),
            'total': job.get('total', 0),
            'done': job.get('done', 0),
            'updated': job.get('updated', 0),
            'where': job.get('where', ''),
            'error': job.get('error', ''),
        })


@app.route('/pallet/<int:pallet_id>/mass-price', methods=['POST'])
def pallet_mass_price(pallet_id):
    """Save prices for all products in pallet at once."""
    pallet = query_db("SELECT * FROM pallets WHERE id = ?", (pallet_id,), one=True)
    if not pallet:
        flash('Pallet not found.', 'error')
        return redirect(url_for('pallets_list'))

    updated = 0
    for key, value in request.form.items():
        if key.startswith('price_'):
            try:
                pid = int(key.replace('price_', ''))
                price = float(value or 0)
                execute_db("UPDATE products SET ebay_price_gbp = ? WHERE id = ? AND pallet_id = ?",
                          (price, pid, pallet_id))
                updated += 1
            except (ValueError, TypeError):
                continue

    flash(f'Updated prices for {updated} products.', 'success')
    return redirect(url_for('pallet_detail', pallet_id=pallet_id))


@app.route('/pallet/<int:pallet_id>/create-drafts', methods=['POST'])
def pallet_create_drafts(pallet_id):
    """Create draft listings for all warehouse products (saved locally, NOT sent to eBay)."""
    pallet = query_db("SELECT * FROM pallets WHERE id = ?", (pallet_id,), one=True)
    if not pallet:
        flash('Pallet not found.', 'error')
        return redirect(url_for('pallets_list'))

    products = query_db(
        "SELECT * FROM products WHERE pallet_id = ? AND status = 'warehouse' AND name != ''",
        (pallet_id,)
    )

    if not products:
        flash('No eligible products (need warehouse status and a name).', 'info')
        return redirect(url_for('pallet_detail', pallet_id=pallet_id))

    created = 0
    skipped = 0
    for product in products:
        # Skip if draft already exists
        existing = query_db("SELECT id FROM ebay_listings WHERE product_id = ? AND status = 'draft'",
                           (product['id'],), one=True)
        if existing:
            skipped += 1
            continue

        title = product['name'][:80]
        price = product['ebay_price_gbp'] or 0.0
        description = (
            '<div style="font-family:Arial,sans-serif">'
            f'<h2>{title}</h2>'
            f'<p>{product["name"]}</p>'
            f'<p>Condition: {product["condition"].replace("_", " ").title()}</p>'
            '<p>Fast dispatch from UK warehouse.</p>'
            '</div>'
        )
        execute_db(
            "INSERT INTO ebay_listings (product_id, title, description, price_gbp, status) "
            "VALUES (?, ?, ?, ?, 'draft')",
            (product['id'], title, description, price)
        )
        created += 1

    msg = f'Created {created} drafts.'
    if skipped:
        msg += f' ({skipped} already had drafts)'
    msg += ' Go to Listings to review and publish.'
    flash(msg, 'success')
    return redirect(url_for('pallet_detail', pallet_id=pallet_id))


@app.route('/pallet/<int:pallet_id>/publish-all', methods=['POST'])
def pallet_publish_all(pallet_id):
    """Publish all draft listings to eBay."""
    pallet = query_db("SELECT * FROM pallets WHERE id = ?", (pallet_id,), one=True)
    if not pallet:
        flash('Pallet not found.', 'error')
        return redirect(url_for('pallets_list'))

    ebay = get_ebay_client(get_config)
    if not ebay.is_configured():
        flash('eBay API not configured. Go to Settings.', 'error')
        return redirect(url_for('pallet_detail', pallet_id=pallet_id))

    drafts = query_db(
        """SELECT l.*, p.image_url, p.images, p.custom_images, p.condition, p.quantity,
                  p.ean, p.id as prod_id,
                  p.shipping_method, p.shipping_cost_gbp, p.shipping_pricing_mode,
                  p.category, p.item_specifics as prod_specs,
                  p.weight_kg, p.length_cm, p.width_cm, p.height_cm
           FROM ebay_listings l
           JOIN products p ON p.id = l.product_id
           WHERE p.pallet_id = ? AND l.status = 'draft'""",
        (pallet_id,)
    )
    # Resolve the public base URL once per batch — used to turn custom image
    # paths into absolute URLs that eBay can fetch from the outside.
    _custom_img_base = (get_config('public_base_url', '') or request.host_url or '').rstrip('/')

    if not drafts:
        flash('No drafts to publish.', 'info')
        return redirect(url_for('pallet_detail', pallet_id=pallet_id))

    shipping_key = get_config('default_shipping', 'royal_mail_2nd')
    return_days = int(get_config('default_return_days', '30') or '30')
    default_pricing = get_config('default_shipping_pricing', 'flat') or 'flat'
    origin_postcode = get_config('seller_postcode', '')
    returns_policy = get_config('returns_policy', 'no') or 'no'

    published = 0
    failed = 0
    errors = []

    for draft in drafts:
        if (draft['price_gbp'] or 0) <= 0:
            failed += 1
            errors.append(f"{draft['title'][:40]}: no price")
            continue

        # Custom (uploaded) photos first, then Amazon. Caps at eBay's 12-picture limit.
        image_urls = []
        try:
            custom_rels = json.loads(draft.get('custom_images') or '[]')
            if isinstance(custom_rels, list) and _custom_img_base:
                for rel in custom_rels:
                    image_urls.append(f'{_custom_img_base}/static/{rel}')
        except (json.JSONDecodeError, TypeError):
            pass
        try:
            all_imgs = json.loads(draft.get('images') or '[]')
            for url in all_imgs:
                if url and url not in image_urls:
                    image_urls.append(url)
        except (json.JSONDecodeError, TypeError):
            pass
        if not image_urls and draft.get('image_url'):
            image_urls = [draft['image_url']]
        image_urls = image_urls[:12]

        # Parse item specifics from listing or product
        listing_specs = {}
        try:
            listing_specs = json.loads(draft.get('item_specifics') or draft.get('prod_specs') or '{}')
        except (json.JSONDecodeError, TypeError):
            pass

        try:
            # Use product-specific shipping or defaults
            prod_shipping = draft.get('shipping_method') or shipping_key
            prod_shipping_cost = draft.get('shipping_cost_gbp') or 0

            # Validate weight / dimensions against shipping method
            fit_ok, fit_err = validate_shipping_fit(
                prod_shipping,
                weight_kg=draft.get('weight_kg'),
                length_cm=draft.get('length_cm'),
                width_cm=draft.get('width_cm'),
                height_cm=draft.get('height_cm'),
            )
            if not fit_ok:
                failed += 1
                errors.append(f"{draft['title'][:40]}: {fit_err}")
                continue

            # Category from listing or product (format: "id:name" or just "id")
            _cat = (draft.get('category_id') or draft.get('category') or '175673').split(':')[0]
            prod_pricing = (draft.get('shipping_pricing_mode') or '').strip() or default_pricing
            result = ebay.create_listing({
                'title': draft['title'],
                'description': draft['description'],
                'price': draft['price_gbp'],
                'condition': draft['condition'],
                'quantity': draft['quantity'],
                'category_id': _cat,
                'image_urls': image_urls,
                'ean': draft.get('ean', ''),
                'dispatch_days': 3,
                'shipping_service': prod_shipping,
                'item_specifics': listing_specs,
                'shipping_cost': prod_shipping_cost,
                'shipping_pricing_mode': prod_pricing,
                'origin_postcode': origin_postcode,
                'weight_kg': draft.get('weight_kg'),
                'length_cm': draft.get('length_cm'),
                'width_cm': draft.get('width_cm'),
                'height_cm': draft.get('height_cm'),
                'return_days': return_days,
                'returns_policy': returns_policy,
            })

            if result and result.get('success'):
                execute_db(
                    "UPDATE ebay_listings SET ebay_item_id=?, status='active' WHERE id=?",
                    (result['ebay_item_id'], draft['id'])
                )
                execute_db("UPDATE products SET status='listed' WHERE id=?", (draft['prod_id'],))
                published += 1
            else:
                error_msg = result.get('error', 'Unknown error') if result else 'API error'
                failed += 1
                errors.append(f"{draft['title'][:40]}: {error_msg[:80]}")
        except Exception as e:
            failed += 1
            errors.append(f"{draft['title'][:40]}: {str(e)[:80]}")

    msg = f'Published {published} listings, {failed} failed.'
    if errors:
        msg += ' Errors: ' + '; '.join(errors[:5])
    flash(msg, 'success' if failed == 0 else 'warning')
    return redirect(url_for('pallet_detail', pallet_id=pallet_id))


@app.route('/pallet/<int:pallet_id>/add_product', methods=['POST'])
def add_product(pallet_id):
    name = request.form.get('name', '').strip()
    asin = request.form.get('asin', '').strip()
    ean = request.form.get('ean', '').strip()
    quantity = request.form.get('quantity', '1')
    condition = request.form.get('condition', 'new')
    ebay_price = request.form.get('ebay_price_gbp', '0')
    category = request.form.get('category', '').strip()

    if not name:
        flash('Product name is required.', 'error')
        return redirect(url_for('pallet_detail', pallet_id=pallet_id))

    try:
        quantity = int(quantity)
    except ValueError:
        quantity = 1
    try:
        ebay_price = float(ebay_price)
    except ValueError:
        ebay_price = 0.0

    pallet = query_db("SELECT * FROM pallets WHERE id = ?", (pallet_id,), one=True)
    product_count = query_db(
        "SELECT COUNT(*) as cnt FROM products WHERE pallet_id = ?",
        (pallet_id,), one=True
    )['cnt']
    cost_per_unit = (pallet['purchase_price_gbp'] or 0) / max(product_count + 1, 1)

    image_url = amazon_image(asin)

    execute_db(
        "INSERT INTO products (pallet_id, name, asin, ean, quantity, condition, "
        "ebay_price_gbp, cost_per_unit, category, image_url) "
        "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
        (pallet_id, name, asin, ean, quantity, condition,
         ebay_price, cost_per_unit, category, image_url)
    )
    flash(f'Product "{name}" added.', 'success')
    return redirect(url_for('pallet_detail', pallet_id=pallet_id))


@app.route('/product/<int:product_id>/update', methods=['POST'])
def product_update(product_id):
    name = request.form.get('name', '').strip()
    asin = request.form.get('asin', '').strip()
    ean = request.form.get('ean', '').strip()
    quantity = request.form.get('quantity', '1')
    condition = request.form.get('condition', 'new')
    ebay_price = request.form.get('ebay_price_gbp', '0')
    category = request.form.get('category', '').strip()
    status = request.form.get('status', 'warehouse')
    weight_kg = request.form.get('weight_kg', '0')
    length_cm = request.form.get('length_cm', '0')
    width_cm = request.form.get('width_cm', '0')
    height_cm = request.form.get('height_cm', '0')
    shipping_method = request.form.get('shipping_method', '')
    shipping_cost = request.form.get('shipping_cost_gbp', '0')
    shipping_pricing_mode = request.form.get('shipping_pricing_mode', '')
    if shipping_pricing_mode not in ('', 'flat', 'calculated'):
        shipping_pricing_mode = ''

    try:
        quantity = int(quantity)
    except ValueError:
        quantity = 1
    try:
        ebay_price = float(ebay_price)
    except ValueError:
        ebay_price = 0.0
    try:
        weight_kg = float(weight_kg or 0)
        length_cm = float(length_cm or 0)
        width_cm = float(width_cm or 0)
        height_cm = float(height_cm or 0)
        shipping_cost = float(shipping_cost or 0)
    except ValueError:
        weight_kg = length_cm = width_cm = height_cm = shipping_cost = 0.0

    # Keep existing image — only set new one if product has no image yet
    existing = query_db("SELECT image_url FROM products WHERE id = ?", (product_id,), one=True)
    image_url = (existing['image_url'] if existing and existing['image_url'] else amazon_image(asin))

    execute_db(
        "UPDATE products SET name=?, asin=?, ean=?, quantity=?, condition=?, "
        "ebay_price_gbp=?, category=?, image_url=?, status=?, "
        "weight_kg=?, length_cm=?, width_cm=?, height_cm=?, shipping_method=?, "
        "shipping_cost_gbp=?, shipping_pricing_mode=? WHERE id=?",
        (name, asin, ean, quantity, condition, ebay_price,
         category, image_url, status,
         weight_kg, length_cm, width_cm, height_cm, shipping_method,
         shipping_cost, shipping_pricing_mode, product_id)
    )
    flash('Product updated.', 'success')
    return redirect(url_for('product_detail', product_id=product_id))


@app.route('/product/<int:product_id>/delete', methods=['POST'])
def product_delete(product_id):
    product = query_db("SELECT * FROM products WHERE id = ?", (product_id,), one=True)
    if not product:
        flash('Product not found.', 'error')
        return redirect(url_for('pallets_list'))
    pallet_id = product['pallet_id']
    execute_db("DELETE FROM ebay_listings WHERE product_id = ?", (product_id,))
    execute_db("DELETE FROM sales WHERE product_id = ?", (product_id,))
    execute_db("DELETE FROM products WHERE id = ?", (product_id,))
    flash('Product deleted.', 'success')
    return redirect(url_for('pallet_detail', pallet_id=pallet_id))


# ---------------------------------------------------------------------------
# Custom images + colour-mismatch detection
# ---------------------------------------------------------------------------

# Minimal English colour vocabulary used to cross-check Amazon metadata against
# the supplier's ground-truth Product Description. Kept narrow on purpose —
# we're flagging pallet-seller confusion (Busybee: Amazon=Champagne/Warm white,
# supplier=Green), not doing general NLP. Multi-word terms first so
# "warm white" matches as one unit before being split on "white".
COLOUR_WORDS = (
    'warm white', 'cool white', 'cream white', 'off white', 'navy blue',
    'sky blue', 'rose gold',
    'red', 'orange', 'yellow', 'green', 'blue', 'purple', 'pink',
    'brown', 'black', 'white', 'grey', 'gray', 'silver', 'gold',
    'champagne', 'beige', 'cream', 'ivory', 'turquoise', 'cyan',
    'magenta', 'maroon', 'navy', 'olive', 'teal', 'lime', 'tan',
    'bronze', 'copper', 'pearl', 'charcoal',
)

# Colour words that are too generic to flag on their own — almost every
# package mentions "white box" or "black text". Only raise the warning if a
# STRONGER colour (non-neutral) mismatches.
NEUTRAL_COLOURS = {'white', 'black', 'grey', 'gray', 'cream', 'ivory', 'beige'}


def _extract_colours(text):
    """Return a set of colour words found in free-form text (lowercased)."""
    if not text:
        return set()
    lowered = ' ' + text.lower() + ' '
    found = set()
    # Replace matched terms with spaces so "warm white" doesn't also hit "white".
    for word in COLOUR_WORDS:
        needle = f' {word} '
        if needle in lowered:
            found.add(word)
            lowered = lowered.replace(needle, ' ' * (len(needle)))
    return found


def detect_colour_mismatch(product):
    """
    Compare colour words in the supplier description against Amazon-derived
    metadata (item_specifics + Amazon title stored in name). Returns None when
    no meaningful mismatch, or a dict {supplier: [...], amazon: [...]} when
    the two sources disagree on colour.
    """
    supplier_desc = (product.get('supplier_description') or '').strip()
    if not supplier_desc:
        return None

    # Amazon side: name (Amazon title replaces supplier title when longer) + specs
    amazon_text_parts = [product.get('name') or '']
    specs_raw = product.get('item_specifics') or ''
    try:
        specs = json.loads(specs_raw) if specs_raw else {}
        if isinstance(specs, dict):
            for k, v in specs.items():
                kl = k.lower()
                if 'colour' in kl or 'color' in kl:
                    amazon_text_parts.append(str(v))
    except (json.JSONDecodeError, TypeError):
        pass
    amazon_text = ' '.join(amazon_text_parts)

    supplier_colours = _extract_colours(supplier_desc)
    amazon_colours = _extract_colours(amazon_text)

    if not supplier_colours or not amazon_colours:
        return None

    # Overlap = at least one shared colour → no mismatch to flag.
    if supplier_colours & amazon_colours:
        return None

    # Require at least one NON-neutral colour on each side. Without that we're
    # just flagging "white box" vs "black ink" kind of noise.
    supplier_strong = supplier_colours - NEUTRAL_COLOURS
    amazon_strong = amazon_colours - NEUTRAL_COLOURS
    if not supplier_strong and not amazon_strong:
        return None

    return {
        'supplier': sorted(supplier_colours),
        'amazon': sorted(amazon_colours),
    }


UPLOAD_ROOT = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                           'static', 'uploads', 'products')
ALLOWED_IMAGE_EXT = {'.jpg', '.jpeg', '.png', '.webp', '.gif'}
MAX_IMAGE_BYTES = 8 * 1024 * 1024  # 8 MB per file — eBay's own cap is ~12 MB


def _product_image_dir(product_id):
    d = os.path.join(UPLOAD_ROOT, str(int(product_id)))
    os.makedirs(d, exist_ok=True)
    return d


def _load_custom_images(product):
    try:
        data = json.loads(product.get('custom_images') or '[]')
        return data if isinstance(data, list) else []
    except (json.JSONDecodeError, TypeError):
        return []


@app.route('/product/<int:product_id>/upload-images', methods=['POST'])
def product_upload_images(product_id):
    """Upload user-supplied product photos. Files are stored in
    static/uploads/products/<id>/ and their relative paths appended to the
    product's custom_images JSON array."""
    product = query_db("SELECT * FROM products WHERE id = ?", (product_id,), one=True)
    if not product:
        flash('Product not found.', 'error')
        return redirect(url_for('pallets_list'))

    files = request.files.getlist('images')
    if not files:
        flash('No files selected.', 'error')
        return redirect(url_for('product_detail', product_id=product_id))

    saved = _load_custom_images(product)
    dest_dir = _product_image_dir(product_id)
    added = 0
    rejected = 0

    for f in files:
        if not f or not f.filename:
            continue
        ext = os.path.splitext(f.filename)[1].lower()
        if ext not in ALLOWED_IMAGE_EXT:
            rejected += 1
            continue
        # Enforce size by seeking — Flask doesn't cap uploads by default here.
        f.stream.seek(0, io.SEEK_END)
        size = f.stream.tell()
        f.stream.seek(0)
        if size > MAX_IMAGE_BYTES or size == 0:
            rejected += 1
            continue
        # UUID-ish filename, preserving extension. Avoids collisions and strips
        # any path tricks from the original filename.
        token = secrets.token_hex(8)
        fname = f'{token}{ext}'
        path = os.path.join(dest_dir, fname)
        try:
            f.save(path)
        except Exception as e:
            print(f'[UPLOAD] save failed for {fname}: {e}')
            rejected += 1
            continue
        rel = f'uploads/products/{int(product_id)}/{fname}'
        saved.append(rel)
        added += 1

    execute_db(
        "UPDATE products SET custom_images = ? WHERE id = ?",
        (json.dumps(saved), product_id)
    )
    if added:
        flash(f'Uploaded {added} photo{"s" if added != 1 else ""}.' +
              (f' {rejected} rejected (bad type or size).' if rejected else ''),
              'success')
    else:
        flash(f'No photos uploaded. {rejected} rejected (bad type or size).', 'error')
    return redirect(url_for('product_detail', product_id=product_id))


@app.route('/product/<int:product_id>/delete-image', methods=['POST'])
def product_delete_image(product_id):
    """Remove one user-uploaded image. Expects form field 'path' with the
    relative path stored in custom_images."""
    product = query_db("SELECT * FROM products WHERE id = ?", (product_id,), one=True)
    if not product:
        flash('Product not found.', 'error')
        return redirect(url_for('pallets_list'))

    target = (request.form.get('path') or '').strip()
    if not target:
        return redirect(url_for('product_detail', product_id=product_id))

    # Defence-in-depth — only delete paths that belong to this product's folder.
    expected_prefix = f'uploads/products/{int(product_id)}/'
    if not target.startswith(expected_prefix) or '..' in target:
        flash('Invalid image path.', 'error')
        return redirect(url_for('product_detail', product_id=product_id))

    saved = [p for p in _load_custom_images(product) if p != target]
    execute_db(
        "UPDATE products SET custom_images = ? WHERE id = ?",
        (json.dumps(saved), product_id)
    )

    abs_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'static',
                            target.replace('/', os.sep))
    try:
        if os.path.isfile(abs_path):
            os.remove(abs_path)
    except Exception as e:
        print(f'[UPLOAD] delete failed for {target}: {e}')

    flash('Photo removed.', 'success')
    return redirect(url_for('product_detail', product_id=product_id))


@app.route('/product/<int:product_id>/list_ebay', methods=['POST'])
def list_on_ebay(product_id):
    product = query_db("SELECT * FROM products WHERE id = ?", (product_id,), one=True)
    if not product:
        flash('Product not found.', 'error')
        return redirect(url_for('pallets_list'))

    title = request.form.get('title', product['name'])[:80]
    description = request.form.get('description', '')
    price = request.form.get('price', str(product['ebay_price_gbp']))

    try:
        price = float(price)
    except ValueError:
        price = product['ebay_price_gbp']

    # Generate HTML description if user left it empty
    if not description.strip():
        description = (
            '<div style="font-family:Arial,sans-serif">'
            f'<h2>{title}</h2>'
            f'<p>{product["name"]}</p>'
            f'<p>Condition: {product["condition"].replace("_", " ").title()}</p>'
            '<p>Fast dispatch from UK warehouse.</p>'
            '</div>'
        )

    action = request.form.get('action', 'draft')

    # Check if draft already exists — update instead of creating new
    existing_draft = query_db(
        "SELECT id FROM ebay_listings WHERE product_id = ? AND status = 'draft'",
        (product_id,), one=True
    )
    if existing_draft:
        execute_db(
            "UPDATE ebay_listings SET title=?, description=?, price_gbp=? WHERE id=?",
            (title, description, price, existing_draft['id'])
        )
        listing_id = existing_draft['id']
    else:
        listing_id = execute_db(
            "INSERT INTO ebay_listings (product_id, title, description, price_gbp, status) "
            "VALUES (?, ?, ?, ?, 'draft')",
            (product_id, title, description, price)
        )

    if action == 'draft':
        flash('Draft saved! You can publish it later.', 'success')
        return redirect(url_for('product_detail', product_id=product_id))

    # Publish to eBay
    ebay = get_ebay_client(get_config)
    if not ebay.is_configured():
        flash('Draft saved. Configure eBay API in Settings to publish.', 'info')
        return redirect(url_for('product_detail', product_id=product_id))

    shipping_key = product.get('shipping_method') or get_config('default_shipping', 'royal_mail_2nd')
    shipping_cost = product.get('shipping_cost_gbp') or 0
    # Per-product pricing mode; '' means fall back to Settings default.
    pricing_mode = (product.get('shipping_pricing_mode') or '').strip() \
        or get_config('default_shipping_pricing', 'flat') or 'flat'
    origin_postcode = get_config('seller_postcode', '')
    return_days = int(get_config('default_return_days', '30') or '30')
    returns_policy = get_config('returns_policy', 'no') or 'no'
    cat = (product.get('category') or '175673').split(':')[0]

    # Validate product fits the shipping method (weight + dimensions).
    # Missing data = pass — only block when we KNOW it won't fit.
    fit_ok, fit_err = validate_shipping_fit(
        shipping_key,
        weight_kg=product.get('weight_kg'),
        length_cm=product.get('length_cm'),
        width_cm=product.get('width_cm'),
        height_cm=product.get('height_cm'),
    )
    if not fit_ok:
        flash(f'Cannot publish — {fit_err}', 'error')
        return redirect(url_for('product_detail', product_id=product_id))

    # Build the eBay image list. Custom (user-uploaded) photos come first —
    # they're the actual pallet contents, not Amazon's variant stock shots.
    # Then Amazon images fill any remaining slots up to eBay's 12-picture cap.
    image_urls = []
    # Custom uploads — convert relative paths to absolute URLs eBay can fetch.
    # Prefer the configured public URL (set when the app is behind ngrok on the
    # Pi), fall back to request.host_url so dev/LAN setups still work.
    custom_rels = _load_custom_images(product)
    if custom_rels:
        base = (get_config('public_base_url', '') or request.host_url or '').rstrip('/')
        for rel in custom_rels:
            if base:
                image_urls.append(f'{base}/static/{rel}')
    try:
        all_imgs = json.loads(product.get('images') or '[]')
        for url in all_imgs:
            if url and url not in image_urls:
                image_urls.append(url)
    except (json.JSONDecodeError, TypeError):
        pass
    if not image_urls and product.get('image_url'):
        image_urls = [product['image_url']]
    image_urls = image_urls[:12]

    # Parse item specifics
    prod_specs = {}
    try:
        prod_specs = json.loads(product.get('item_specifics') or '{}')
    except (json.JSONDecodeError, TypeError):
        pass

    result = ebay.create_listing({
        'title': title,
        'description': description,
        'price': price,
        'condition': product['condition'],
        'quantity': product['quantity'],
        'category_id': cat,
        'ean': product['ean'],
        'image_urls': image_urls,
        'dispatch_days': 3,
        'shipping_service': shipping_key,
        'shipping_cost': shipping_cost,
        'shipping_pricing_mode': pricing_mode,
        'origin_postcode': origin_postcode,
        'weight_kg': product.get('weight_kg'),
        'length_cm': product.get('length_cm'),
        'width_cm': product.get('width_cm'),
        'height_cm': product.get('height_cm'),
        'return_days': return_days,
        'returns_policy': returns_policy,
        'item_specifics': prod_specs,
    })
    if result and result.get('success'):
        execute_db("UPDATE ebay_listings SET ebay_item_id=?, status='active' WHERE id=?",
                   (result['ebay_item_id'], listing_id))
        execute_db("UPDATE products SET status='listed' WHERE id=?", (product_id,))
        fees_msg = f' (fees: GBP {result["fees"]:.2f})' if result.get('fees') else ''
        flash(f'Published on eBay!{fees_msg} Item ID: {result["ebay_item_id"]}', 'success')
    else:
        error = result.get('error', 'Unknown error') if result else 'API error'
        flash(f'Draft saved. eBay error: {error}', 'warning')

    return redirect(url_for('product_detail', product_id=product_id))


@app.route('/order/<int:order_id>/ship', methods=['POST'])
def order_mark_shipped(order_id):
    execute_db(
        "UPDATE sales SET status='shipped', shipped_at=CURRENT_TIMESTAMP WHERE id=?",
        (order_id,)
    )
    order = query_db("SELECT * FROM sales WHERE id = ?", (order_id,), one=True)
    if order and order['product_id']:
        execute_db(
            "UPDATE products SET status='shipped' WHERE id=?",
            (order['product_id'],)
        )
    flash('Order marked as shipped.', 'success')
    return redirect(url_for('orders_list'))


@app.route('/api/add_sale', methods=['POST'])
def api_add_sale():
    product_id = request.form.get('product_id')
    price = request.form.get('price_gbp', '0')
    buyer = request.form.get('buyer', '')
    address = request.form.get('shipping_address', '')

    try:
        price = float(price)
    except ValueError:
        price = 0.0

    execute_db(
        "INSERT INTO sales (product_id, price_gbp, buyer, shipping_address) "
        "VALUES (?, ?, ?, ?)",
        (product_id, price, buyer, address)
    )

    if product_id:
        execute_db("UPDATE products SET status='sold' WHERE id=?", (product_id,))

    flash('Sale recorded.', 'success')
    return redirect(url_for('orders_list'))


@app.route('/product/<int:product_id>/sell-private', methods=['POST'])
def product_sell_private(product_id):
    """Record a private sale — the uncle sold this item in person, to a friend,
    at a local market, etc. Still counts toward revenue/profit on dashboard.
    No shipping_address (private sales are typically collected in person)."""
    product = query_db("SELECT id, status FROM products WHERE id = ?",
                       (product_id,), one=True)
    if not product:
        flash('Product not found.', 'error')
        return redirect(url_for('pallets_list'))

    try:
        price = float(request.form.get('price_gbp', '0') or '0')
    except ValueError:
        price = 0.0
    if price <= 0:
        flash('Enter a sale price greater than 0.', 'error')
        return redirect(url_for('product_detail', product_id=product_id))

    buyer = (request.form.get('buyer') or '').strip() or 'Private buyer'
    notes = (request.form.get('notes') or '').strip()
    mark_shipped = request.form.get('mark_shipped') == '1'

    # Private sales are typically completed on the spot — default to 'shipped'
    # if the uncle ticked the "already handed over" box, otherwise 'new'.
    status = 'shipped' if mark_shipped else 'new'
    shipped_at = 'CURRENT_TIMESTAMP' if mark_shipped else 'NULL'

    execute_db(
        f"INSERT INTO sales (product_id, price_gbp, buyer, status, source, notes, shipped_at) "
        f"VALUES (?, ?, ?, ?, 'private', ?, {shipped_at})",
        (product_id, price, buyer, status, notes)
    )

    # Mark product sold/shipped so it stops appearing as "warehouse stock".
    product_new_status = 'shipped' if mark_shipped else 'sold'
    execute_db("UPDATE products SET status=? WHERE id=?",
               (product_new_status, product_id))

    flash(f'Private sale recorded: £{price:.2f}.', 'success')
    return redirect(url_for('product_detail', product_id=product_id))


# ===================================================================
# CSS THEME
# ===================================================================

CSS_THEME = """
/* eBay Hub UK - Cyberpunk Theme */
@import url('https://fonts.googleapis.com/css2?family=Space+Grotesk:wght@300;400;500;600;700&display=swap');

:root {
    --cyan: #8ff5ff;
    --lime: #beee00;
    --pink: #ff6b9b;
    --purple: #a855f7;
    --bg-dark: #0a0a0f;
    --bg-card: #12121a;
    --bg-card-hover: #1a1a26;
    --bg-input: #1a1a26;
    --border: #2a2a3a;
    --text: #e0e0e8;
    --text-muted: #6a6a80;
    --danger: #ff4444;
    --warning: #ffaa00;
    --radius: 12px;
    --radius-sm: 8px;
}

* { margin: 0; padding: 0; box-sizing: border-box; }

body {
    font-family: 'Space Grotesk', sans-serif;
    background: var(--bg-dark);
    color: var(--text);
    min-height: 100vh;
    line-height: 1.5;
}

/* Scrollbar */
::-webkit-scrollbar { width: 8px; }
::-webkit-scrollbar-track { background: var(--bg-dark); }
::-webkit-scrollbar-thumb { background: var(--border); border-radius: 4px; }
::-webkit-scrollbar-thumb:hover { background: var(--purple); }

/* Layout */
.app-container {
    max-width: 1280px;
    margin: 0 auto;
    padding: 0 16px 80px 16px;
}

/* Navigation */
.navbar {
    background: var(--bg-card);
    border-bottom: 1px solid var(--border);
    padding: 12px 0;
    position: sticky;
    top: 0;
    z-index: 100;
    backdrop-filter: blur(10px);
}
.navbar-inner {
    max-width: 1280px;
    margin: 0 auto;
    padding: 0 16px;
    display: flex;
    align-items: center;
    justify-content: space-between;
    flex-wrap: wrap;
    gap: 8px;
}
.navbar-brand {
    display: flex;
    align-items: center;
    gap: 8px;
    text-decoration: none;
    font-weight: 700;
    font-size: 1.2rem;
    color: var(--cyan);
}
.navbar-brand .material-symbols-outlined { font-size: 28px; }
.nav-links {
    display: flex;
    gap: 4px;
    flex-wrap: wrap;
}
.nav-links a {
    display: flex;
    align-items: center;
    gap: 4px;
    padding: 8px 12px;
    color: var(--text-muted);
    text-decoration: none;
    border-radius: var(--radius-sm);
    font-size: 0.85rem;
    font-weight: 500;
    transition: all 0.2s;
}
.nav-links a:hover, .nav-links a.active {
    color: var(--cyan);
    background: rgba(143, 245, 255, 0.08);
}
.nav-links a .material-symbols-outlined { font-size: 20px; }

/* Page header */
.page-header {
    display: flex;
    align-items: center;
    justify-content: space-between;
    flex-wrap: wrap;
    gap: 12px;
    margin: 24px 0 20px 0;
}
.page-header h1 {
    font-size: 1.5rem;
    font-weight: 700;
    color: var(--text);
}
.page-header h1 span { color: var(--cyan); }

/* Cards */
.card {
    background: var(--bg-card);
    border: 1px solid var(--border);
    border-radius: var(--radius);
    padding: 20px;
    transition: border-color 0.2s;
}
.card:hover { border-color: rgba(143, 245, 255, 0.2); }
.card-title {
    font-size: 0.8rem;
    text-transform: uppercase;
    letter-spacing: 0.05em;
    color: var(--text-muted);
    margin-bottom: 8px;
}
.card-value {
    font-size: 1.8rem;
    font-weight: 700;
    color: var(--cyan);
}
.card-value.lime { color: var(--lime); }
.card-value.pink { color: var(--pink); }
.card-value.purple { color: var(--purple); }
.card-subtitle {
    font-size: 0.8rem;
    color: var(--text-muted);
    margin-top: 4px;
}

/* Stats grid */
.stats-grid {
    display: grid;
    grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
    gap: 16px;
    margin-bottom: 24px;
}

/* Tables */
.table-wrap {
    overflow-x: auto;
    border-radius: var(--radius);
    border: 1px solid var(--border);
    background: var(--bg-card);
}
table {
    width: 100%;
    border-collapse: collapse;
}
th {
    text-align: left;
    padding: 12px 16px;
    font-size: 0.75rem;
    text-transform: uppercase;
    letter-spacing: 0.05em;
    color: var(--text-muted);
    border-bottom: 1px solid var(--border);
    white-space: nowrap;
}
td {
    padding: 12px 16px;
    border-bottom: 1px solid var(--border);
    vertical-align: middle;
}
tr:last-child td { border-bottom: none; }
tr:hover td { background: var(--bg-card-hover); }
.table-link {
    color: var(--cyan);
    text-decoration: none;
    font-weight: 500;
}
.table-link:hover { text-decoration: underline; }

/* Badges */
.badge {
    display: inline-flex;
    align-items: center;
    padding: 2px 10px;
    border-radius: 20px;
    font-size: 0.75rem;
    font-weight: 600;
    text-transform: uppercase;
    letter-spacing: 0.03em;
}
.badge-cyan { background: rgba(143, 245, 255, 0.12); color: var(--cyan); }
.badge-lime { background: rgba(190, 238, 0, 0.12); color: var(--lime); }
.badge-pink { background: rgba(255, 107, 155, 0.12); color: var(--pink); }
.badge-purple { background: rgba(168, 85, 247, 0.12); color: var(--purple); }
.badge-muted { background: rgba(106, 106, 128, 0.12); color: var(--text-muted); }
.badge-danger { background: rgba(255, 68, 68, 0.12); color: var(--danger); }

/* Buttons */
.btn {
    display: inline-flex;
    align-items: center;
    gap: 6px;
    padding: 10px 20px;
    border: none;
    border-radius: var(--radius-sm);
    font-family: 'Space Grotesk', sans-serif;
    font-size: 0.9rem;
    font-weight: 600;
    cursor: pointer;
    transition: all 0.2s;
    text-decoration: none;
    white-space: nowrap;
}
.btn .material-symbols-outlined { font-size: 20px; }
.btn-cyan {
    background: var(--cyan);
    color: #0a0a0f;
}
.btn-cyan:hover { background: #adf8ff; box-shadow: 0 0 20px rgba(143, 245, 255, 0.3); }
.btn-lime {
    background: var(--lime);
    color: #0a0a0f;
}
.btn-lime:hover { background: #d4ff2a; box-shadow: 0 0 20px rgba(190, 238, 0, 0.3); }
.btn-pink {
    background: var(--pink);
    color: #0a0a0f;
}
.btn-pink:hover { background: #ff8db5; box-shadow: 0 0 20px rgba(255, 107, 155, 0.3); }
.btn-purple {
    background: var(--purple);
    color: #fff;
}
.btn-purple:hover { background: #b86ef9; box-shadow: 0 0 20px rgba(168, 85, 247, 0.3); }
.btn-outline {
    background: transparent;
    border: 1px solid var(--border);
    color: var(--text);
}
.btn-outline:hover { border-color: var(--cyan); color: var(--cyan); }
.btn-danger {
    background: rgba(255, 68, 68, 0.15);
    color: var(--danger);
    border: 1px solid rgba(255, 68, 68, 0.3);
}
.btn-danger:hover { background: rgba(255, 68, 68, 0.25); }
.btn-sm { padding: 6px 14px; font-size: 0.8rem; }
.btn-block { width: 100%; justify-content: center; }

/* Filter tabs */
.filter-tabs {
    display: flex;
    gap: 4px;
    margin-bottom: 20px;
    flex-wrap: wrap;
}
.filter-tab {
    padding: 8px 16px;
    border-radius: var(--radius-sm);
    background: var(--bg-card);
    border: 1px solid var(--border);
    color: var(--text-muted);
    text-decoration: none;
    font-size: 0.85rem;
    font-weight: 500;
    transition: all 0.2s;
}
.filter-tab:hover { border-color: var(--cyan); color: var(--cyan); }
.filter-tab.active {
    background: rgba(143, 245, 255, 0.1);
    border-color: var(--cyan);
    color: var(--cyan);
}
.filter-count {
    display: inline-flex;
    align-items: center;
    justify-content: center;
    min-width: 20px;
    height: 20px;
    padding: 0 6px;
    border-radius: 10px;
    background: rgba(143, 245, 255, 0.15);
    font-size: 0.7rem;
    margin-left: 6px;
}

/* Forms */
.form-group {
    margin-bottom: 16px;
}
.form-label {
    display: block;
    margin-bottom: 6px;
    font-size: 0.85rem;
    font-weight: 500;
    color: var(--text-muted);
}
.form-control {
    width: 100%;
    padding: 10px 14px;
    background: var(--bg-input);
    border: 1px solid var(--border);
    border-radius: var(--radius-sm);
    color: var(--text);
    font-family: 'Space Grotesk', sans-serif;
    font-size: 0.9rem;
    transition: border-color 0.2s;
}
.form-control:focus {
    outline: none;
    border-color: var(--cyan);
    box-shadow: 0 0 0 3px rgba(143, 245, 255, 0.1);
}
select.form-control {
    appearance: none;
    background-image: url("data:image/svg+xml,%3Csvg width='12' height='8' viewBox='0 0 12 8' xmlns='http://www.w3.org/2000/svg'%3E%3Cpath d='M1 1l5 5 5-5' stroke='%236a6a80' stroke-width='2' fill='none'/%3E%3C/svg%3E");
    background-repeat: no-repeat;
    background-position: right 14px center;
    padding-right: 36px;
}
textarea.form-control { resize: vertical; min-height: 80px; }
.form-row {
    display: grid;
    grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
    gap: 16px;
}
.form-hint {
    font-size: 0.75rem;
    color: var(--text-muted);
    margin-top: 4px;
}

/* Modal */
.modal-overlay {
    display: none;
    position: fixed;
    top: 0; left: 0; right: 0; bottom: 0;
    background: rgba(0, 0, 0, 0.7);
    backdrop-filter: blur(4px);
    z-index: 200;
    align-items: center;
    justify-content: center;
    padding: 16px;
}
.modal-overlay.active { display: flex; }
.modal {
    background: var(--bg-card);
    border: 1px solid var(--border);
    border-radius: var(--radius);
    padding: 24px;
    width: 100%;
    max-width: 520px;
    max-height: 90vh;
    overflow-y: auto;
}
.modal-title {
    font-size: 1.2rem;
    font-weight: 700;
    margin-bottom: 20px;
    color: var(--cyan);
}

/* Flash messages */
.flash-container { margin: 16px 0; }
.flash {
    padding: 12px 16px;
    border-radius: var(--radius-sm);
    margin-bottom: 8px;
    font-size: 0.9rem;
    display: flex;
    align-items: center;
    gap: 8px;
}
.flash-success { background: rgba(190, 238, 0, 0.1); border: 1px solid rgba(190, 238, 0, 0.3); color: var(--lime); }
.flash-error { background: rgba(255, 68, 68, 0.1); border: 1px solid rgba(255, 68, 68, 0.3); color: var(--danger); }
.flash-warning { background: rgba(255, 170, 0, 0.1); border: 1px solid rgba(255, 170, 0, 0.3); color: var(--warning); }
.flash-info { background: rgba(143, 245, 255, 0.1); border: 1px solid rgba(143, 245, 255, 0.3); color: var(--cyan); }

/* Product grid */
.product-grid {
    display: grid;
    grid-template-columns: repeat(auto-fill, minmax(280px, 1fr));
    gap: 16px;
    margin-bottom: 24px;
}
.product-card {
    background: var(--bg-card);
    border: 1px solid var(--border);
    border-radius: var(--radius);
    overflow: hidden;
    transition: all 0.2s;
    text-decoration: none;
    color: var(--text);
    display: block;
}
.product-card:hover {
    border-color: var(--cyan);
    transform: translateY(-2px);
    box-shadow: 0 8px 24px rgba(0, 0, 0, 0.3);
}
.product-card-img {
    width: 100%;
    height: 160px;
    object-fit: contain;
    background: #1a1a26;
    padding: 12px;
}
.product-card-body { padding: 14px; }
.product-card-name {
    font-weight: 600;
    font-size: 0.9rem;
    margin-bottom: 6px;
    display: -webkit-box;
    -webkit-line-clamp: 2;
    -webkit-box-orient: vertical;
    overflow: hidden;
}
.product-card-meta {
    display: flex;
    align-items: center;
    justify-content: space-between;
    margin-top: 8px;
}
.product-card-price {
    font-weight: 700;
    color: var(--lime);
    font-size: 1.1rem;
}

/* Chart bar */
.chart-container { margin-bottom: 24px; }
.chart-bars {
    display: flex;
    align-items: flex-end;
    gap: 8px;
    height: 160px;
    padding: 0 4px;
}
.chart-bar-wrap {
    flex: 1;
    display: flex;
    flex-direction: column;
    align-items: center;
    height: 100%;
    justify-content: flex-end;
}
.chart-bar {
    width: 100%;
    max-width: 60px;
    background: linear-gradient(to top, var(--cyan), var(--purple));
    border-radius: 4px 4px 0 0;
    min-height: 2px;
    transition: height 0.3s;
}
.chart-bar-label {
    font-size: 0.65rem;
    color: var(--text-muted);
    margin-top: 6px;
    text-align: center;
}
.chart-bar-value {
    font-size: 0.7rem;
    color: var(--cyan);
    margin-bottom: 4px;
    font-weight: 600;
}

/* Detail page layout */
.detail-header {
    display: flex;
    flex-direction: column;
    gap: 24px;
    margin-bottom: 24px;
}
.detail-image {
    width: 200px;
    height: 200px;
    object-fit: contain;
    background: var(--bg-card);
    border: 1px solid var(--border);
    border-radius: var(--radius);
    padding: 12px;
    flex-shrink: 0;
}
.detail-info { flex: 1; min-width: 250px; }
.detail-info h2 {
    font-size: 1.3rem;
    margin-bottom: 12px;
}
.detail-meta {
    display: grid;
    grid-template-columns: auto 1fr;
    gap: 6px 16px;
    font-size: 0.9rem;
}
.detail-meta dt { color: var(--text-muted); }
.detail-meta dd { color: var(--text); }

/* Section headings */
.section-title {
    font-size: 1.1rem;
    font-weight: 600;
    margin: 24px 0 16px 0;
    padding-bottom: 8px;
    border-bottom: 1px solid var(--border);
    color: var(--cyan);
}

/* Empty state */
.empty-state {
    text-align: center;
    padding: 48px 20px;
    color: var(--text-muted);
}
.empty-state .material-symbols-outlined {
    font-size: 48px;
    margin-bottom: 12px;
    opacity: 0.3;
}
.empty-state p { font-size: 1rem; margin-bottom: 16px; }

/* Utility */
.text-cyan { color: var(--cyan); }
.text-lime { color: var(--lime); }
.text-pink { color: var(--pink); }
.text-purple { color: var(--purple); }
.text-muted { color: var(--text-muted); }
.text-danger { color: var(--danger); }
.text-right { text-align: right; }
.mt-16 { margin-top: 16px; }
.mb-16 { margin-bottom: 16px; }
.flex-between {
    display: flex;
    align-items: center;
    justify-content: space-between;
    gap: 12px;
    flex-wrap: wrap;
}
.gap-8 { gap: 8px; }
.d-flex { display: flex; }
.align-center { align-items: center; }
.inline-form { display: inline; }

/* Responsive */
@media (max-width: 768px) {
    .navbar-inner { flex-direction: column; align-items: flex-start; }
    .nav-links { width: 100%; overflow-x: auto; }
    .page-header { flex-direction: column; align-items: flex-start; }
    .page-header h1 { font-size: 1.2rem; }
    .stats-grid { grid-template-columns: repeat(2, 1fr); }
    .card-value { font-size: 1.4rem; }
    .detail-header { flex-direction: column; }
    .detail-image { width: 100%; height: 200px; }
    .form-row { grid-template-columns: 1fr; }
    .product-grid { grid-template-columns: 1fr; }
    .modal { padding: 16px; }
}

@media (max-width: 480px) {
    .stats-grid { grid-template-columns: 1fr; }
    .nav-links a span.nav-text { display: none; }
}
"""

# ===================================================================
# TEMPLATES
# ===================================================================

# -------------------------------------------------------------------
# Base template (layout wrapper)
# -------------------------------------------------------------------
TEMPLATE_BASE = """<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{{ page_title | default('eBay Hub UK') }}</title>
    <meta name="theme-color" content="#8ff5ff">
    <meta name="apple-mobile-web-app-capable" content="yes">
    <meta name="apple-mobile-web-app-status-bar-style" content="black-translucent">
    <link rel="manifest" href="/static/manifest.json">
    <link rel="apple-touch-icon" href="/static/icon-192.png">
    <link rel="preconnect" href="https://fonts.googleapis.com">
    <link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
    <link nonce="{{ nonce() }}" href="https://fonts.googleapis.com/css2?family=Space+Grotesk:wght@300;400;500;600;700&display=swap" rel="stylesheet">
    <link nonce="{{ nonce() }}" href="https://fonts.googleapis.com/css2?family=Material+Symbols+Outlined:opsz,wght,FILL,GRAD@20..48,100..700,0..1,-50..200&display=swap" rel="stylesheet">
    <style nonce="{{ nonce() }}">""" + CSS_THEME + """</style>
</head>
<body>

<nav class="navbar">
    <div class="navbar-inner">
        <a href="/" class="navbar-brand">
            <span class="material-symbols-outlined">hub</span>
            eBay Hub UK
        </a>
        <div class="nav-links">
            <a href="/" class="{{ 'active' if active_page == 'dashboard' else '' }}">
                <span class="material-symbols-outlined">dashboard</span>
                <span class="nav-text">Dashboard</span>
            </a>
            <a href="/pallets" class="{{ 'active' if active_page == 'pallets' else '' }}">
                <span class="material-symbols-outlined">inventory_2</span>
                <span class="nav-text">Pallets</span>
            </a>
            <a href="/listings" class="{{ 'active' if active_page == 'listings' else '' }}">
                <span class="material-symbols-outlined">sell</span>
                <span class="nav-text">Listings</span>
            </a>
            <a href="/orders" class="{{ 'active' if active_page == 'orders' else '' }}">
                <span class="material-symbols-outlined">local_shipping</span>
                <span class="nav-text">Orders</span>
            </a>
            <a href="/settings" class="{{ 'active' if active_page == 'settings' else '' }}">
                <span class="material-symbols-outlined">settings</span>
                <span class="nav-text">Settings</span>
            </a>
            <a href="/help" class="{{ 'active' if active_page == 'help' else '' }}">
                <span class="material-symbols-outlined">help</span>
            </a>
            <a href="/logout" style="color:#ef4444">
                <span class="material-symbols-outlined">logout</span>
            </a>
        </div>
    </div>
</nav>

<div class="app-container">
    {% with messages = get_flashed_messages(with_categories=true) %}
    {% if messages %}
    <div class="flash-container">
        {% for category, message in messages %}
        <div class="flash flash-{{ category }}">
            <span class="material-symbols-outlined">
                {% if category == 'success' %}check_circle{% elif category == 'error' %}error{% elif category == 'warning' %}warning{% else %}info{% endif %}
            </span>
            {{ message }}
        </div>
        {% endfor %}
    </div>
    {% endif %}
    {% endwith %}

    {{ content }}
</div>

<!-- Loading Spinner Overlay -->
<div id="loadingOverlay" style="display:none;position:fixed;inset:0;background:rgba(0,0,0,0.92);z-index:99999;align-items:center;justify-content:center;flex-direction:column;gap:16px">
    <div style="width:56px;height:56px;border:3px solid rgba(143,245,255,0.15);border-top:3px solid #8ff5ff;border-radius:50%;animation:spin 0.8s linear infinite"></div>
    <div id="loadingText" style="color:#8ff5ff;font-family:'Space Grotesk',sans-serif;font-size:1.1rem;font-weight:700">Processing...</div>
    <div id="loadingSubtext" style="color:rgba(255,255,255,0.5);font-size:0.8rem">This may take a few minutes for large files</div>
    <div id="loadingTimer" style="color:rgba(143,245,255,0.4);font-size:0.75rem;font-family:monospace">00:00</div>
    <div style="margin-top:12px;padding:12px 24px;background:rgba(143,245,255,0.06);border:1px solid rgba(143,245,255,0.12);max-width:320px;text-align:center">
        <div style="font-size:0.7rem;color:rgba(255,255,255,0.3)">Do NOT close this page</div>
    </div>
</div>
<style>@keyframes spin{to{transform:rotate(360deg)}}</style>
<script>
if('serviceWorker' in navigator){navigator.serviceWorker.register('/static/sw.js')}
// Show spinner on form submit (for imports/scraping)
var _loadingStart = 0;
document.querySelectorAll('form').forEach(function(f){
    f.addEventListener('submit', function(){
        var overlay = document.getElementById('loadingOverlay');
        var btn = f.querySelector('button[type="submit"]');
        var hasFile = f.querySelector('input[type="file"]');
        var action = f.action || '';
        var isSlow = hasFile || action.includes('/scrape') || action.includes('/import') || action.includes('/add') || action.includes('/list-all') || action.includes('/list_ebay') || action.includes('/publish-all') || action.includes('/create-drafts') || action.includes('/auto-categories') || action.includes('/auto-pipeline') || action.includes('/mass-price');
        if(isSlow){
            overlay.style.display='flex';
            if(btn) btn.disabled=true;
            _loadingStart = Date.now();
            // Context-aware messages
            var msg = 'Processing...';
            var sub = 'Please wait';
            if(action.includes('/import') || (hasFile && action.includes('/add'))) { msg='Importing products...'; sub='Scraping Amazon UK for images & prices'; }
            else if(action.includes('/scrape')) { msg='Scraping images...'; sub='Fetching from Amazon UK (~3s per product)'; }
            else if(action.includes('/publish') || action.includes('/list')) { msg='Publishing to eBay...'; sub='Sending listings to eBay UK'; }
            else if(action.includes('/create-drafts')) { msg='Creating drafts...'; sub='Saving listings locally'; }
            document.getElementById('loadingText').textContent = msg;
            document.getElementById('loadingSubtext').textContent = sub;
            // Timer
            setInterval(function(){
                var elapsed = Math.floor((Date.now() - _loadingStart) / 1000);
                var min = Math.floor(elapsed/60);
                var sec = elapsed%60;
                document.getElementById('loadingTimer').textContent = (min<10?'0':'')+min+':'+(sec<10?'0':'')+sec;
            }, 1000);
        }
    });
});
</script>
</body>
</html>"""

# -------------------------------------------------------------------
# Helper: wrap content in base
# -------------------------------------------------------------------
def render_page(content_template, page_title='eBay Hub UK', active_page='', **kwargs):
    """Render a content template inside the base layout."""
    from markupsafe import Markup
    content_html = render_template_string(content_template, **kwargs)
    return render_template_string(
        TEMPLATE_BASE,
        content=Markup(content_html),
        page_title=page_title,
        active_page=active_page,
        **kwargs
    )


# -------------------------------------------------------------------
# Dashboard Template
# -------------------------------------------------------------------
TEMPLATE_DASHBOARD_CONTENT = """
<div class="page-header">
    <h1><span>Dashboard</span></h1>
</div>

<div class="stats-grid">
    <div class="card">
        <div class="card-title">Today's Sales</div>
        <div class="card-value">{{ fmt_gbp(stats.sales_today.total) }}</div>
        <div class="card-subtitle">{{ stats.sales_today.cnt }} order{{ 's' if stats.sales_today.cnt != 1 }}</div>
    </div>
    <div class="card">
        <div class="card-title">This Week</div>
        <div class="card-value lime">{{ fmt_gbp(stats.sales_week.total) }}</div>
        <div class="card-subtitle">{{ stats.sales_week.cnt }} order{{ 's' if stats.sales_week.cnt != 1 }}</div>
    </div>
    <div class="card">
        <div class="card-title">This Month</div>
        <div class="card-value purple">{{ fmt_gbp(stats.sales_month.total) }}</div>
        <div class="card-subtitle">{{ stats.sales_month.cnt }} order{{ 's' if stats.sales_month.cnt != 1 }}</div>
    </div>
    <div class="card">
        <div class="card-title">Total Profit</div>
        <div class="card-value {{ 'lime' if stats.total_profit >= 0 else 'text-danger' }}">
            {{ fmt_gbp(stats.total_profit) }}
        </div>
        <div class="card-subtitle">Revenue: {{ fmt_gbp(stats.total_revenue) }}</div>
    </div>
</div>

<div class="stats-grid">
    <div class="card">
        <div class="card-title">Active Listings</div>
        <div class="card-value cyan">{{ stats.active_listings }}</div>
    </div>
    <div class="card">
        <div class="card-title">To Ship</div>
        <div class="card-value pink">{{ stats.to_ship }}</div>
        {% if stats.to_ship > 0 %}
        <a href="/orders?status=new" class="btn btn-pink btn-sm mt-16">
            <span class="material-symbols-outlined">local_shipping</span> View Orders
        </a>
        {% endif %}
    </div>
    <div class="card">
        <div class="card-title">Products</div>
        <div class="card-value purple">{{ stats.total_products }}</div>
        <div class="card-subtitle">{{ stats.warehouse_products }} in warehouse</div>
    </div>
    <div class="card">
        <div class="card-title">Frozen Capital</div>
        <div class="card-value">{{ fmt_gbp(stats.frozen_capital) }}</div>
        <div class="card-subtitle">{{ stats.total_pallets }} pallet{{ 's' if stats.total_pallets != 1 }}</div>
    </div>
</div>

<!-- Revenue Chart -->
<div class="card chart-container">
    <div class="card-title">Monthly Revenue</div>
    {% set max_rev = stats.chart_data | map(attribute='revenue') | max %}
    <div class="chart-bars">
        {% for bar in stats.chart_data %}
        <div class="chart-bar-wrap">
            <div class="chart-bar-value">{{ fmt_gbp(bar.revenue) }}</div>
            <div class="chart-bar" style="height: {{ (bar.revenue / max_rev * 120) if max_rev > 0 else 2 }}px"></div>
            <div class="chart-bar-label">{{ bar.label }}</div>
        </div>
        {% endfor %}
    </div>
</div>

<!-- Quick Actions -->
<div class="card">
    <div class="card-title">Quick Actions</div>
    <div class="d-flex gap-8" style="flex-wrap: wrap; margin-top: 12px;">
        <a href="/pallets" class="btn btn-cyan">
            <span class="material-symbols-outlined">add</span> New Pallet
        </a>
        <a href="/listings" class="btn btn-purple">
            <span class="material-symbols-outlined">sell</span> View Listings
        </a>
        <a href="/orders?status=new" class="btn btn-pink">
            <span class="material-symbols-outlined">local_shipping</span> Pending Orders
        </a>
        <a href="/settings" class="btn btn-outline">
            <span class="material-symbols-outlined">settings</span> Settings
        </a>
    </div>
</div>
"""

# -------------------------------------------------------------------
# Route: Dashboard
# -------------------------------------------------------------------
@app.route('/')
def dashboard():
    stats = get_dashboard_stats()
    return render_page(
        TEMPLATE_DASHBOARD_CONTENT,
        page_title='Dashboard - eBay Hub UK',
        active_page='dashboard',
        stats=stats
    )


# -------------------------------------------------------------------
# Pallets Template
# -------------------------------------------------------------------
TEMPLATE_PALLETS_CONTENT = """
<div class="page-header">
    <h1><span>Pallets</span></h1>
    <div class="d-flex gap-8">
        <button type="button" id="bulk-select-btn" class="btn btn-outline btn-sm" onclick="toggleSelectAllPallets()">
            <span class="material-symbols-outlined">check_box</span> Select All
        </button>
        <form method="POST" action="/pallets/merge-duplicates" class="inline-form"
              onsubmit="return confirm('Find pallets with the same name+supplier and merge them into one? Products will be moved to the oldest pallet, and the duplicates will be deleted.')">
            <button type="submit" class="btn btn-outline btn-sm" style="border-color:rgba(168,85,247,0.3);color:#a855f7">
                <span class="material-symbols-outlined">merge</span> Merge Duplicates
            </button>
        </form>
        <button class="btn btn-cyan" onclick="document.getElementById('addPalletModal').classList.add('active')">
            <span class="material-symbols-outlined">add</span> Add Pallet
        </button>
    </div>
</div>

<!-- Sticky bulk-delete bar (shows when any pallet row is checked) -->
<div id="bulk-delete-bar" style="display:none;position:sticky;top:0;z-index:50;background:linear-gradient(135deg,#ef4444,#dc2626);border-radius:10px;padding:12px 16px;margin-bottom:12px;align-items:center;justify-content:space-between;box-shadow:0 4px 15px rgba(239,68,68,0.4)">
    <span style="color:#fff;font-weight:600;font-size:0.9rem" id="bulk-delete-count">0 selected</span>
    <button type="button" onclick="bulkDeletePallets()" style="background:#fff;color:#dc2626;border:none;padding:8px 20px;border-radius:8px;font-weight:700;cursor:pointer;font-size:0.85rem">
        <span class="material-symbols-outlined">delete</span> DELETE SELECTED
    </button>
</div>

{% if pallets %}
<div class="table-wrap">
    <table>
        <thead>
            <tr>
                <th style="width:36px"><input type="checkbox" id="header-select-all" onchange="toggleSelectAllPallets()"></th>
                <th>Name</th>
                <th>Supplier</th>
                <th>Cost</th>
                <th>Products</th>
                <th>Sold</th>
                <th>Revenue</th>
                <th>ROI</th>
                <th>Status</th>
            </tr>
        </thead>
        <tbody>
            {% for p in pallets %}
            <tr id="pallet-row-{{ p.id }}">
                <td><input type="checkbox" class="pallet-cb" value="{{ p.id }}" onchange="updateBulkDelete()"></td>
                <td>
                    <a href="/pallet/{{ p.id }}" class="table-link">{{ p.name }}</a>
                </td>
                <td class="text-muted">{{ p.supplier or '-' }}</td>
                <td>{{ fmt_gbp(p.purchase_price_gbp) }}</td>
                <td>{{ p.product_count }}</td>
                <td>{{ p.sold_count }}</td>
                <td class="text-lime">{{ fmt_gbp(p.revenue) }}</td>
                <td>
                    {% if p.purchase_price_gbp > 0 %}
                        {% set roi = ((p.revenue - p.purchase_price_gbp) / p.purchase_price_gbp * 100) %}
                        <span class="{{ 'text-lime' if roi >= 0 else 'text-danger' }}">
                            {{ "%.0f"|format(roi) }}%
                        </span>
                    {% else %}-{% endif %}
                </td>
                <td><span class="badge {{ status_color(p.status) }}">{{ p.status }}</span></td>
            </tr>
            {% endfor %}
        </tbody>
    </table>
</div>
{% else %}
<div class="empty-state">
    <span class="material-symbols-outlined">inventory_2</span>
    <p>No pallets yet. Add your first pallet to get started!</p>
    <button class="btn btn-cyan" onclick="document.getElementById('addPalletModal').classList.add('active')">
        <span class="material-symbols-outlined">add</span> Add Pallet
    </button>
</div>
{% endif %}

<script>
function updateBulkDelete() {
    var checked = document.querySelectorAll('.pallet-cb:checked');
    var bar = document.getElementById('bulk-delete-bar');
    var count = document.getElementById('bulk-delete-count');
    if (checked.length > 0) {
        bar.style.display = 'flex';
        count.textContent = checked.length + ' selected';
    } else {
        bar.style.display = 'none';
    }
}

function toggleSelectAllPallets() {
    var cbs = document.querySelectorAll('.pallet-cb');
    var allChecked = [].every.call(cbs, function(cb){ return cb.checked; });
    [].forEach.call(cbs, function(cb){ cb.checked = !allChecked; });
    var hdr = document.getElementById('header-select-all');
    if (hdr) hdr.checked = !allChecked;
    updateBulkDelete();
}

function bulkDeletePallets() {
    var checked = document.querySelectorAll('.pallet-cb:checked');
    var ids = [].map.call(checked, function(cb){ return cb.value; });
    if (ids.length === 0) return;
    if (!confirm('Delete ' + ids.length + ' pallets with all their products? This cannot be undone!')) return;

    fetch('/pallets/bulk-delete', {
        method: 'POST',
        headers: {'Content-Type': 'application/json'},
        body: JSON.stringify({ids: ids})
    })
    .then(function(r){ return r.json(); })
    .then(function(data){
        if (data.ok) {
            ids.forEach(function(id){
                var row = document.getElementById('pallet-row-' + id);
                if (row) row.remove();
            });
            updateBulkDelete();
            location.reload();
        } else {
            alert('Error: ' + (data.error || 'Unknown'));
        }
    })
    .catch(function(e){ alert('Error: ' + e); });
}
</script>

<!-- Add Pallet Modal -->
<div id="addPalletModal" class="modal-overlay" onclick="if(event.target===this)this.classList.remove('active')">
    <div class="modal">
        <div class="modal-title">Add New Pallet</div>
        <form method="POST" action="/pallets/add" enctype="multipart/form-data">
            <div class="form-group">
                <label class="form-label">Pallet Name *</label>
                <input type="text" name="name" class="form-control" placeholder="e.g. Amazon Returns Batch #12" required>
            </div>
            <div class="form-row">
                <div class="form-group">
                    <label class="form-label">Supplier</label>
                    <input type="text" name="supplier" class="form-control" placeholder="e.g. Wholesale Co">
                </div>
                <div class="form-group">
                    <label class="form-label">Purchase Price (GBP)</label>
                    <input type="number" step="0.01" name="purchase_price_gbp" class="form-control" placeholder="0.00">
                </div>
            </div>
            <div class="form-row">
                <div class="form-group">
                    <label class="form-label">Purchase Date</label>
                    <input type="date" name="purchase_date" class="form-control">
                </div>
                <div class="form-group">
                    <label class="form-label">Amazon Locale Override (optional)</label>
                    <select name="amazon_domain" class="form-control">
                        <option value="" selected>Auto-detect (scraper finds it)</option>
                        <option value="amazon.co.uk">Amazon.co.uk (UK)</option>
                        <option value="amazon.de">Amazon.de (Germany)</option>
                        <option value="amazon.com">Amazon.com (US)</option>
                        <option value="amazon.pl">Amazon.pl (Poland)</option>
                        <option value="amazon.fr">Amazon.fr (France)</option>
                        <option value="amazon.it">Amazon.it (Italy)</option>
                        <option value="amazon.es">Amazon.es (Spain)</option>
                        <option value="amazon.nl">Amazon.nl (Netherlands)</option>
                        <option value="amazon.se">Amazon.se (Sweden)</option>
                    </select>
                    <div style="font-size:0.7rem;color:var(--text-muted);margin-top:4px">
                        Leave on Auto-detect — only set if auto picks the wrong variant for your pallet.
                    </div>
                </div>
            </div>
            <div class="form-group">
                <label class="form-label">Notes</label>
                <textarea name="notes" class="form-control" rows="2" placeholder="Optional notes..."></textarea>
            </div>
            <div style="border-top:1px solid rgba(255,255,255,0.08);padding-top:14px;margin-top:14px">
                <label class="form-label"><span class="material-symbols-outlined" style="font-size:0.9rem;vertical-align:middle">upload_file</span> Import Specification (CSV / XLSX)</label>
                <input type="file" name="spec_file" accept=".csv,.xlsx,.xls" class="form-control" style="padding:8px">
                <div style="font-size:0.7rem;color:var(--text-muted);margin-top:4px">Upload supplier file with products. ASINs are auto-scraped from whichever Amazon locale they live on (always fetched in English).</div>
            </div>
            <div class="d-flex gap-8" style="justify-content: flex-end; margin-top: 20px;">
                <button type="button" class="btn btn-outline" onclick="document.getElementById('addPalletModal').classList.remove('active')">Cancel</button>
                <button type="submit" class="btn btn-cyan">
                    <span class="material-symbols-outlined">add</span> Add Pallet
                </button>
            </div>
        </form>
    </div>
</div>
"""

@app.route('/pallets')
def pallets_list():
    pallets = query_db("""
        SELECT p.*,
            COUNT(pr.id) as product_count,
            SUM(CASE WHEN pr.status = 'sold' THEN 1 ELSE 0 END) as sold_count,
            COALESCE(SUM(CASE WHEN pr.status = 'sold' THEN pr.ebay_price_gbp ELSE 0 END), 0) as revenue
        FROM pallets p
        LEFT JOIN products pr ON pr.pallet_id = p.id
        GROUP BY p.id
        ORDER BY p.created_at DESC
    """)
    return render_page(
        TEMPLATE_PALLETS_CONTENT,
        page_title='Pallets - eBay Hub UK',
        active_page='pallets',
        pallets=pallets
    )


# -------------------------------------------------------------------
# Pallet Detail Template
# -------------------------------------------------------------------
TEMPLATE_PALLET_DETAIL_CONTENT = """
<div class="page-header">
    <h1>
        <a href="/pallets" class="text-muted" style="text-decoration: none;">Pallets</a>
        <span class="text-muted">/</span>
        <span>{{ pallet.name }}</span>
    </h1>
    <div class="d-flex gap-8">
        <a href="/pallet/{{ pallet.id }}/import" class="btn btn-purple btn-sm">
            <span class="material-symbols-outlined">upload_file</span> Import CSV
        </a>
        <form method="POST" action="/pallet/{{ pallet.id }}/create-drafts" class="inline-form">
            <button type="submit" class="btn btn-outline btn-sm" style="border-color:rgba(245,158,11,0.3);color:#f59e0b">
                <span class="material-symbols-outlined">edit_note</span> Create Drafts
            </button>
        </form>
        <form method="POST" action="/pallet/{{ pallet.id }}/publish-all" class="inline-form"
              onsubmit="return confirm('Publish all drafts to eBay? Listings go LIVE immediately.') && (document.getElementById('loadingOverlay').style.display='flex',document.getElementById('loadingText').textContent='Publishing to eBay...',true)">
            <button type="submit" class="btn btn-lime btn-sm">
                <span class="material-symbols-outlined">sell</span> Publish All
            </button>
        </form>
        <form method="POST" action="/pallet/{{ pallet.id }}/auto-categories" class="inline-form"
              onsubmit="document.getElementById('loadingOverlay').style.display='flex';document.getElementById('loadingText').textContent='Matching categories...'">
            <button type="submit" class="btn btn-outline btn-sm" style="border-color:rgba(168,85,247,0.3);color:#a855f7">
                <span class="material-symbols-outlined">category</span> Auto Categories
            </button>
        </form>
        <form method="POST" action="/pallet/{{ pallet.id }}/auto-pipeline" class="inline-form"
              onsubmit="document.getElementById('loadingOverlay').style.display='flex';document.getElementById('loadingText').textContent='Running auto-pipeline (scrape, AI, drafts)...'">
            <button type="submit" class="btn btn-sm" style="background:linear-gradient(135deg,rgba(0,255,136,0.15),rgba(139,92,246,0.15));border:1px solid rgba(0,255,136,0.4);color:#00ff88;">
                <span class="material-symbols-outlined">auto_fix_high</span> Auto Pipeline
            </button>
        </form>
        <form method="POST" action="/pallet/{{ pallet.id }}/scrape" class="inline-form">
            <button type="submit" class="btn btn-cyan btn-sm">
                <span class="material-symbols-outlined">photo_camera</span> Scrape Images
            </button>
        </form>
        <button type="button" class="btn btn-outline btn-sm" onclick="document.getElementById('editPalletModal').classList.add('active')">
            <span class="material-symbols-outlined">edit</span> Edit
        </button>
        <form method="POST" action="/pallet/{{ pallet.id }}/archive" class="inline-form">
            <button type="submit" class="btn btn-outline btn-sm">
                <span class="material-symbols-outlined">archive</span> Archive
            </button>
        </form>
        <form method="POST" action="/pallet/{{ pallet.id }}/delete" class="inline-form" onsubmit="return confirm('Delete this pallet and all its products? This cannot be undone.')">
            <button type="submit" class="btn btn-sm" style="background:rgba(239,68,68,0.15);border:1px solid rgba(239,68,68,0.3);color:#ef4444">
                <span class="material-symbols-outlined">delete</span> Delete
            </button>
        </form>
    </div>
</div>

<!-- Edit Pallet Modal — name / supplier / purchase price / date / notes -->
<div id="editPalletModal" class="modal-overlay" onclick="if(event.target===this)this.classList.remove('active')">
    <div class="modal-content">
        <div class="modal-header">
            <h2>Edit Pallet</h2>
            <button class="modal-close" onclick="document.getElementById('editPalletModal').classList.remove('active')">&times;</button>
        </div>
        <form method="POST" action="/pallet/{{ pallet.id }}/edit">
            <div class="form-group">
                <label>Pallet Name *</label>
                <input type="text" name="name" class="form-control" required value="{{ pallet.name or '' }}">
            </div>
            <div class="form-group">
                <label>Supplier</label>
                <input type="text" name="supplier" class="form-control" value="{{ pallet.supplier or '' }}">
            </div>
            <div class="form-group">
                <label>Purchase Price (&pound;)</label>
                <input type="number" step="0.01" min="0" name="purchase_price_gbp" class="form-control" value="{{ '%.2f'|format(pallet.purchase_price_gbp or 0) }}">
            </div>
            <div class="form-group">
                <label>Purchase Date</label>
                <input type="date" name="purchase_date" class="form-control" value="{{ pallet.purchase_date or '' }}">
            </div>
            <div class="form-group">
                <label>Notes</label>
                <textarea name="notes" class="form-control" rows="3">{{ pallet.notes or '' }}</textarea>
            </div>
            <div class="d-flex gap-8" style="justify-content:flex-end">
                <button type="button" class="btn btn-outline" onclick="document.getElementById('editPalletModal').classList.remove('active')">Cancel</button>
                <button type="submit" class="btn btn-cyan">
                    <span class="material-symbols-outlined">save</span> Save
                </button>
            </div>
        </form>
    </div>
</div>

<!-- Background-scrape progress bar (hidden unless a job is running for this pallet).
     Polls /pallet/<id>/scrape-status every 2s; reloads when the job finishes so the
     newly scraped images/titles appear without a manual refresh. -->
<div id="scrape-progress" style="display:none;background:linear-gradient(135deg,rgba(0,255,136,0.12),rgba(6,182,212,0.12));border:1px solid rgba(0,255,136,0.35);border-radius:10px;padding:12px 16px;margin-bottom:12px">
    <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:8px">
        <span style="color:#00ff88;font-weight:600;font-size:0.9rem">
            <span class="material-symbols-outlined" style="vertical-align:middle;font-size:1.1rem">sync</span>
            Scraping Amazon <span id="scrape-progress-label">...</span>
        </span>
        <span id="scrape-progress-count" style="color:rgba(255,255,255,0.7);font-size:0.8rem">0 / 0</span>
    </div>
    <div style="background:rgba(0,0,0,0.3);border-radius:6px;height:8px;overflow:hidden">
        <div id="scrape-progress-bar" style="background:linear-gradient(90deg,#00ff88,#06b6d4);height:100%;width:0%;transition:width 0.3s"></div>
    </div>
</div>
<script>
(function(){
    var palletId = {{ pallet.id }};
    var bar = document.getElementById('scrape-progress');
    var label = document.getElementById('scrape-progress-label');
    var count = document.getElementById('scrape-progress-count');
    var fill = document.getElementById('scrape-progress-bar');
    var wasRunning = false;

    function poll() {
        fetch('/pallet/' + palletId + '/scrape-status')
            .then(function(r){ return r.json(); })
            .then(function(j){
                if (j.status === 'running') {
                    wasRunning = true;
                    bar.style.display = 'block';
                    count.textContent = (j.done || 0) + ' / ' + (j.total || '?');
                    label.textContent = '(' + (j.updated || 0) + ' updated)';
                    var pct = j.total ? Math.round((j.done / j.total) * 100) : 5;
                    fill.style.width = pct + '%';
                } else if (j.status === 'done') {
                    if (wasRunning) { location.reload(); }
                    else { bar.style.display = 'none'; }
                } else if (j.status === 'error') {
                    bar.style.display = 'block';
                    label.textContent = 'ERROR: ' + (j.error || 'Unknown');
                    fill.style.background = '#ef4444';
                } else {
                    bar.style.display = 'none';
                }
            })
            .catch(function(){});
    }
    poll();
    setInterval(poll, 2000);
})();
</script>

<!-- Pallet Stats -->
<div class="stats-grid">
    <div class="card">
        <div class="card-title">Pallet Cost</div>
        <div class="card-value">{{ fmt_gbp(pallet.purchase_price_gbp) }}</div>
        <div class="card-subtitle">{{ pallet.supplier or 'No supplier' }} &middot; {{ fmt_date(pallet.purchase_date) }}</div>
    </div>
    <div class="card">
        <div class="card-title">Products</div>
        <div class="card-value purple">{{ stats.total }}</div>
        <div class="card-subtitle">{{ stats.warehouse }} warehouse / {{ stats.listed }} listed / {{ stats.sold }} sold</div>
    </div>
    <div class="card">
        <div class="card-title">Actual Revenue</div>
        <div class="card-value lime">{{ fmt_gbp(stats.revenue) }}</div>
    </div>
    <div class="card">
        <div class="card-title">Actual Profit</div>
        <div class="card-value {{ 'lime' if profit >= 0 else 'text-danger' }}">{{ fmt_gbp(profit) }}</div>
        {% if pallet.purchase_price_gbp > 0 %}
        <div class="card-subtitle">ROI: {{ "%.0f"|format((profit / pallet.purchase_price_gbp) * 100) }}%</div>
        {% endif %}
    </div>
</div>

{% if estimated_revenue > 0 %}
<div class="stats-grid" style="margin-top:8px">
    <div class="card" style="border-left:3px solid #f59e0b">
        <div class="card-title" style="color:#f59e0b">Est. Revenue (if all sold)</div>
        <div class="card-value" style="color:#f59e0b">{{ fmt_gbp(estimated_revenue) }}</div>
        <div class="card-subtitle">Sum of all set prices</div>
    </div>
    <div class="card" style="border-left:3px solid #f59e0b">
        <div class="card-title" style="color:#f59e0b">Est. Profit (after fees)</div>
        <div class="card-value {{ 'lime' if estimated_profit >= 0 else 'text-danger' }}">{{ fmt_gbp(estimated_profit) }}</div>
        <div class="card-subtitle">After eBay ~12.8% fee + cost</div>
    </div>
    <div class="card" style="border-left:3px solid #f59e0b">
        <div class="card-title" style="color:#f59e0b">Est. ROI</div>
        <div class="card-value {{ 'lime' if estimated_roi >= 0 else 'text-danger' }}">{{ "%.0f"|format(estimated_roi) }}%</div>
        <div class="card-subtitle">{{ products_with_price }}/{{ stats.total }} products priced</div>
    </div>
</div>
{% endif %}

<!-- Amazon locale — auto-detected on first scrape and cached. Override dropdown
     is for cases where auto-detection pulls the wrong variant (same ASIN on
     multiple locales pointing at different products). -->
<div class="card mb-16" style="border-left:3px solid #8ff5ff">
    <form method="POST" action="/pallet/{{ pallet.id }}/set-domain" class="d-flex" style="gap:12px;align-items:center;flex-wrap:wrap">
        <div style="flex:1;min-width:200px">
            <div style="font-size:0.75rem;color:var(--text-muted);text-transform:uppercase;letter-spacing:0.5px">Amazon Locale</div>
            <div style="font-size:1rem;font-weight:600;color:#8ff5ff">
                {% if pallet.amazon_domain %}{{ pallet.amazon_domain }} <span style="font-size:0.7rem;color:var(--text-muted);font-weight:normal">(cached/override)</span>{% else %}Auto-detect <span style="font-size:0.7rem;color:var(--text-muted);font-weight:normal">(scraper finds it on first run)</span>{% endif %}
            </div>
            <div style="font-size:0.7rem;color:var(--text-muted);margin-top:2px">
                Scraper cascades through Amazon locales to find where the ASIN is listed. Pages are always fetched in English.
            </div>
        </div>
        <select name="amazon_domain" class="form-control" style="width:auto;min-width:180px">
            <option value="" {% if not pallet.amazon_domain %}selected{% endif %}>Auto-detect</option>
            <option value="amazon.co.uk" {% if pallet.amazon_domain == 'amazon.co.uk' %}selected{% endif %}>Amazon.co.uk (UK)</option>
            <option value="amazon.de" {% if pallet.amazon_domain == 'amazon.de' %}selected{% endif %}>Amazon.de (Germany)</option>
            <option value="amazon.com" {% if pallet.amazon_domain == 'amazon.com' %}selected{% endif %}>Amazon.com (US)</option>
            <option value="amazon.pl" {% if pallet.amazon_domain == 'amazon.pl' %}selected{% endif %}>Amazon.pl (Poland)</option>
            <option value="amazon.fr" {% if pallet.amazon_domain == 'amazon.fr' %}selected{% endif %}>Amazon.fr (France)</option>
            <option value="amazon.it" {% if pallet.amazon_domain == 'amazon.it' %}selected{% endif %}>Amazon.it (Italy)</option>
            <option value="amazon.es" {% if pallet.amazon_domain == 'amazon.es' %}selected{% endif %}>Amazon.es (Spain)</option>
            <option value="amazon.nl" {% if pallet.amazon_domain == 'amazon.nl' %}selected{% endif %}>Amazon.nl (Netherlands)</option>
            <option value="amazon.se" {% if pallet.amazon_domain == 'amazon.se' %}selected{% endif %}>Amazon.se (Sweden)</option>
        </select>
        <button type="submit" class="btn btn-outline btn-sm">
            <span class="material-symbols-outlined">save</span> Save
        </button>
    </form>
</div>

{% if pallet.notes %}
<div class="card mb-16">
    <div class="card-title">Notes</div>
    <p>{{ pallet.notes }}</p>
</div>
{% endif %}

<!-- Mass Price Editor -->
<div class="card mb-16">
    <div class="flex-between" style="margin-bottom:12px">
        <div class="card-title" style="margin:0"><span class="material-symbols-outlined" style="font-size:18px;vertical-align:middle">paid</span> Set Prices</div>
        <div class="d-flex gap-8">
            <button onclick="applyMultiplier()" class="btn btn-outline btn-sm" style="font-size:0.7rem">Apply Multiplier</button>
            <button onclick="document.getElementById('massPriceForm').submit()" class="btn btn-lime btn-sm">
                <span class="material-symbols-outlined">save</span> Save All Prices
            </button>
        </div>
    </div>
    <div style="display:flex;gap:8px;align-items:center;margin-bottom:12px;font-size:0.8rem;color:var(--text-muted)">
        <label>Multiplier: cost &times;</label>
        <input type="number" id="priceMultiplier" value="2.5" step="0.1" min="1" style="width:70px;padding:6px;background:var(--bg);border:1px solid var(--border);color:var(--text);text-align:center;font-size:0.85rem">
        <span style="color:var(--text-muted);font-size:0.7rem">(e.g. 2.5 = 150% markup)</span>
    </div>
    <form method="POST" action="/pallet/{{ pallet.id }}/mass-price" id="massPriceForm">
        <div style="display:grid;grid-template-columns:1fr 100px;gap:6px;font-size:0.75rem;color:var(--text-muted);padding:0 4px;margin-bottom:4px">
            <div>Product</div>
            <div style="text-align:center">Price (GBP)</div>
        </div>
        {% for p in products %}
        <div style="display:grid;grid-template-columns:1fr 100px;gap:6px;align-items:center;padding:6px 4px;border-bottom:1px solid rgba(255,255,255,0.04)">
            <div style="font-size:0.8rem;white-space:nowrap;overflow:hidden;text-overflow:ellipsis" title="{{ p.name }}">{{ p.name[:45] }}</div>
            <input type="number" name="price_{{ p.id }}" value="{{ '%.2f'|format(p.ebay_price_gbp or 0) }}" step="0.01" min="0" class="mass-price-input" data-pid="{{ p.id }}"
                   style="padding:6px;background:var(--bg);border:1px solid var(--border);color:#8ff5ff;text-align:center;font-size:0.85rem;font-weight:700;font-family:'Space Grotesk',sans-serif">
        </div>
        {% endfor %}
    </form>
</div>

<script>
function applyMultiplier() {
    var mult = parseFloat(document.getElementById('priceMultiplier').value) || 2.5;
    var palletCost = {{ pallet.purchase_price_gbp or 0 }};
    var productCount = {{ products|length or 1 }};
    var costPerUnit = palletCost / Math.max(productCount, 1);
    document.querySelectorAll('.mass-price-input').forEach(function(inp) {
        inp.value = (costPerUnit * mult).toFixed(2);
        inp.style.borderColor = '#beee00';
        setTimeout(function(){ inp.style.borderColor = ''; }, 1000);
    });
}
</script>

<!-- Add Product -->
<div class="flex-between mb-16">
    <h3 class="section-title" style="margin: 0; border: none; padding: 0;">Products</h3>
    <button class="btn btn-cyan btn-sm" onclick="document.getElementById('addProductModal').classList.add('active')">
        <span class="material-symbols-outlined">add</span> Add Product
    </button>
</div>

{% if products %}
<div class="product-grid">
    {% for p in products %}
    <a href="/product/{{ p.id }}" class="product-card">
        {% if p.image_url %}
        <img src="{{ p.image_url }}" alt="{{ p.name }}" class="product-card-img"
             onerror="this.style.display='none'">
        {% else %}
        <div class="product-card-img" style="display:flex;align-items:center;justify-content:center;">
            <span class="material-symbols-outlined" style="font-size:48px;opacity:0.2;">image</span>
        </div>
        {% endif %}
        <div class="product-card-body">
            <div class="product-card-name">{{ p.name }}</div>
            <div style="font-size:0.75rem;color:var(--text-muted);">
                {% if p.asin %}ASIN: {{ p.asin }}{% endif %}
                {% if p.ean %} &middot; EAN: {{ p.ean }}{% endif %}
            </div>
            <div class="product-card-meta">
                <div class="product-card-price">{{ fmt_gbp(p.ebay_price_gbp) }}</div>
                <span class="badge {{ status_color(p.status) }}">{{ p.status }}</span>
            </div>
            <div style="font-size:0.75rem;color:var(--text-muted);margin-top:4px;">
                Qty: {{ p.quantity }} &middot; {{ condition_label(p.condition) }}
            </div>
        </div>
    </a>
    {% endfor %}
</div>
{% else %}
<div class="empty-state">
    <span class="material-symbols-outlined">category</span>
    <p>No products in this pallet yet.</p>
    <div class="d-flex gap-8" style="justify-content: center;">
        <button class="btn btn-cyan" onclick="document.getElementById('addProductModal').classList.add('active')">
            <span class="material-symbols-outlined">add</span> Add Product
        </button>
        <a href="/pallet/{{ pallet.id }}/import" class="btn btn-purple">
            <span class="material-symbols-outlined">upload_file</span> Import CSV
        </a>
    </div>
</div>
{% endif %}

<!-- Add Product Modal -->
<div id="addProductModal" class="modal-overlay" onclick="if(event.target===this)this.classList.remove('active')">
    <div class="modal">
        <div class="modal-title">Add Product</div>
        <form method="POST" action="/pallet/{{ pallet.id }}/add_product">
            <div class="form-group">
                <label class="form-label">Product Name *</label>
                <input type="text" name="name" class="form-control" placeholder="e.g. Sony WH-1000XM5" required>
            </div>
            <div class="form-row">
                <div class="form-group">
                    <label class="form-label">ASIN</label>
                    <input type="text" name="asin" class="form-control" placeholder="B0BS1N8GK7">
                    <div class="form-hint">Amazon product ID (for image)</div>
                </div>
                <div class="form-group">
                    <label class="form-label">EAN / Barcode</label>
                    <input type="text" name="ean" class="form-control" placeholder="5027242923485">
                </div>
            </div>
            <div class="form-row">
                <div class="form-group">
                    <label class="form-label">Quantity</label>
                    <input type="number" name="quantity" class="form-control" value="1" min="1">
                </div>
                <div class="form-group">
                    <label class="form-label">Condition</label>
                    <select name="condition" class="form-control">
                        <option value="new">New</option>
                        <option value="like_new">Like New</option>
                        <option value="used">Used</option>
                        <option value="damaged">Damaged</option>
                    </select>
                </div>
            </div>
            <div class="form-row">
                <div class="form-group">
                    <label class="form-label">eBay Price (GBP)</label>
                    <input type="number" step="0.01" name="ebay_price_gbp" class="form-control" placeholder="0.00">
                </div>
                <div class="form-group">
                    <label class="form-label">Category</label>
                    <input type="text" name="category" class="form-control" placeholder="e.g. Electronics">
                </div>
            </div>
            <div class="d-flex gap-8" style="justify-content: flex-end; margin-top: 20px;">
                <button type="button" class="btn btn-outline" onclick="document.getElementById('addProductModal').classList.remove('active')">Cancel</button>
                <button type="submit" class="btn btn-cyan">
                    <span class="material-symbols-outlined">add</span> Add Product
                </button>
            </div>
        </form>
    </div>
</div>
"""

@app.route('/pallet/<int:pallet_id>')
def pallet_detail(pallet_id):
    pallet = query_db("SELECT * FROM pallets WHERE id = ?", (pallet_id,), one=True)
    if not pallet:
        flash('Pallet not found.', 'error')
        return redirect(url_for('pallets_list'))

    products = query_db(
        "SELECT * FROM products WHERE pallet_id = ? ORDER BY created_at DESC",
        (pallet_id,)
    )

    stats = query_db("""
        SELECT
            COUNT(*) as total,
            SUM(CASE WHEN status = 'listed' THEN 1 ELSE 0 END) as listed,
            SUM(CASE WHEN status = 'sold' THEN 1 ELSE 0 END) as sold,
            SUM(CASE WHEN status = 'warehouse' THEN 1 ELSE 0 END) as warehouse,
            COALESCE(SUM(CASE WHEN status = 'sold' THEN ebay_price_gbp ELSE 0 END), 0) as revenue,
            COALESCE(SUM(ebay_price_gbp * quantity), 0) as potential_revenue
        FROM products WHERE pallet_id = ?
    """, (pallet_id,), one=True)

    profit = (stats['revenue'] or 0) - (pallet['purchase_price_gbp'] or 0)

    # Estimated revenue/profit from set prices
    est = query_db("""
        SELECT COALESCE(SUM(ebay_price_gbp * quantity), 0) as revenue,
               COUNT(CASE WHEN ebay_price_gbp > 0 THEN 1 END) as priced
        FROM products WHERE pallet_id = ? AND status IN ('warehouse', 'listed')
    """, (pallet_id,), one=True)
    estimated_revenue = est['revenue'] or 0
    ebay_fee = estimated_revenue * 0.128  # ~12.8% eBay final value fee
    estimated_profit = estimated_revenue - ebay_fee - (pallet['purchase_price_gbp'] or 0)
    estimated_roi = (estimated_profit / (pallet['purchase_price_gbp'] or 1)) * 100 if pallet['purchase_price_gbp'] else 0

    return render_page(
        TEMPLATE_PALLET_DETAIL_CONTENT,
        page_title=f'{pallet["name"]} - eBay Hub UK',
        active_page='pallets',
        pallet=pallet, products=products, stats=stats, profit=profit,
        estimated_revenue=estimated_revenue, estimated_profit=estimated_profit,
        estimated_roi=estimated_roi, products_with_price=est['priced'] or 0
    )


# -------------------------------------------------------------------
# CSV Import Template
# -------------------------------------------------------------------
TEMPLATE_CSV_IMPORT_CONTENT = """
<div class="page-header">
    <h1>
        <a href="/pallets" class="text-muted" style="text-decoration: none;">Pallets</a>
        <span class="text-muted">/</span>
        <a href="/pallet/{{ pallet.id }}" class="text-muted" style="text-decoration: none;">{{ pallet.name }}</a>
        <span class="text-muted">/</span>
        <span>Import</span>
    </h1>
</div>

<div class="card">
    <div class="card-title">Upload Specification (CSV or Excel)</div>
    <p style="margin: 12px 0; color: var(--text-muted); font-size: 0.9rem;">
        Upload the CSV or XLSX file from your joblot supplier. The app will auto-detect columns and scrape product data from Amazon UK.
    </p>

    <div class="table-wrap mb-16">
        <table>
            <thead>
                <tr>
                    <th>Column</th>
                    <th>Required</th>
                    <th>Description</th>
                </tr>
            </thead>
            <tbody>
                <tr><td>name / title / product</td><td><span class="text-lime">Yes</span></td><td>Product name</td></tr>
                <tr><td>asin</td><td><span class="text-muted">No</span></td><td>Amazon ASIN (for auto-scrape)</td></tr>
                <tr><td>ean / barcode</td><td><span class="text-muted">No</span></td><td>EAN / Barcode</td></tr>
                <tr><td>quantity / qty</td><td><span class="text-muted">No</span></td><td>Quantity (default: 1)</td></tr>
                <tr><td>condition / state</td><td><span class="text-muted">No</span></td><td>new / like_new / used / damaged</td></tr>
                <tr><td>price / ebay_price / rrp</td><td><span class="text-muted">No</span></td><td>Price in GBP</td></tr>
            </tbody>
        </table>
    </div>

    <form method="POST" enctype="multipart/form-data" action="/pallet/{{ pallet.id }}/import">
        <div class="form-group">
            <label class="form-label">Select file (CSV or XLSX)</label>
            <input type="file" name="csv_file" accept=".csv,.xlsx,.xls" class="form-control" required>
        </div>
        <div class="form-group" style="margin-top: 12px;">
            <label style="display: flex; align-items: center; gap: 8px; cursor: pointer; font-size: 0.85rem; color: var(--text-muted);">
                <input type="checkbox" name="auto_scrape" value="1" checked style="accent-color: #8ff5ff;">
                Auto-scrape Amazon{% if pallet.amazon_domain %} ({{ pallet.amazon_domain }}){% else %} (auto-detect locale){% endif %} for products with ASIN (images, titles, prices)
            </label>
        </div>
        <div class="d-flex gap-8" style="margin-top: 20px;">
            <a href="/pallet/{{ pallet.id }}" class="btn btn-outline">Cancel</a>
            <button type="submit" class="btn btn-purple">
                <span class="material-symbols-outlined">upload_file</span> Import Products
            </button>
        </div>
    </form>
</div>
"""

@app.route('/pallet/<int:pallet_id>/import', methods=['GET', 'POST'])
def csv_import(pallet_id):
    pallet = query_db("SELECT * FROM pallets WHERE id = ?", (pallet_id,), one=True)
    if not pallet:
        flash('Pallet not found.', 'error')
        return redirect(url_for('pallets_list'))

    if request.method == 'POST':
        file = request.files.get('csv_file')
        if not file or not file.filename:
            flash('Please upload a file.', 'error')
            return redirect(url_for('csv_import', pallet_id=pallet_id))

        fname = file.filename.lower()
        auto_scrape = request.form.get('auto_scrape') == '1'
        # Empty = auto-detect per ASIN; set = strict override.
        pallet_domain = (pallet.get('amazon_domain') or '').strip()

        try:
            rows = []

            # Parse XLSX
            if fname.endswith(('.xlsx', '.xls')):
                try:
                    import openpyxl
                except ImportError:
                    flash('openpyxl not installed. Run: pip install openpyxl', 'error')
                    return redirect(url_for('csv_import', pallet_id=pallet_id))
                wb = openpyxl.load_workbook(file, data_only=True)
                ws = wb.active
                # Find header row (first row with text)
                headers = []
                for cell in ws[1]:
                    headers.append(str(cell.value or '').strip().lower())
                for row_cells in ws.iter_rows(min_row=2, values_only=True):
                    row = {}
                    for i, val in enumerate(row_cells):
                        if i < len(headers):
                            row[headers[i]] = str(val or '').strip()
                    rows.append(row)

            # Parse CSV
            elif fname.endswith('.csv'):
                raw = file.stream.read()
                # Try UTF-8 first, fallback to latin-1
                for enc in ('utf-8-sig', 'utf-8', 'latin-1', 'cp1252'):
                    try:
                        text = raw.decode(enc)
                        break
                    except:
                        continue
                else:
                    text = raw.decode('utf-8', errors='replace')
                # Detect delimiter
                first_line = text.split('\n')[0]
                delimiter = ';' if first_line.count(';') > first_line.count(',') else ','
                reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
                rows = list(reader)
            else:
                flash('Unsupported file format. Use CSV or XLSX.', 'error')
                return redirect(url_for('csv_import', pallet_id=pallet_id))

            # Column name mapping (flexible)
            def get_col(row, *names):
                # Exact match first, then substring
                for n in names:
                    for key in row:
                        if key and key.lower().strip() == n:
                            val = row[key].strip() if row[key] else ''
                            if val and val.lower() not in ('none', 'nan', 'null'):
                                return val
                for n in names:
                    for key in row:
                        if key and n in key.lower() and 'category' not in key.lower():
                            val = row[key].strip() if row[key] else ''
                            if val and val.lower() not in ('none', 'nan', 'null'):
                                return val
                return ''

            count = 0
            scraped = 0
            from modules.scraper import scrape_amazon_product, get_amazon_image_url

            for row in rows:
                name = get_col(row, 'name', 'title', 'product', 'nazwa')
                # Supplier ground-truth description (actual pallet contents) — kept
                # separate from name so AI can weight it above Amazon variant metadata.
                supplier_desc = get_col(row, 'product description', 'description', 'opis')
                if not name:
                    name = supplier_desc
                if not name:
                    continue

                asin = get_col(row, 'asin').upper()
                ean = get_col(row, 'ean', 'barcode', 'upc', 'gtin')
                try:
                    qty = int(float(get_col(row, 'quantity', 'qty', 'ilosc', 'amount') or '1'))
                except:
                    qty = 1
                cond = get_col(row, 'condition', 'state', 'stan').lower()
                if cond not in ('new', 'like_new', 'used', 'damaged'):
                    cond = 'new'
                try:
                    price = float(get_col(row, 'price', 'ebay_price', 'rrp', 'cena') or '0')
                except:
                    price = 0.0

                image_url = get_amazon_image_url(asin) if asin else ''

                # Auto-scrape: use pallet's override if set, otherwise the scraper
                # cascades through AMAZON_DOMAINS and returns the first locale that
                # actually lists this ASIN.
                if auto_scrape and asin:
                    try:
                        data = scrape_amazon_product(asin, pallet_domain or None)
                        if data:
                            if data.get('title') and len(data['title']) > len(name):
                                name = data['title']
                            if data.get('image_url'):
                                image_url = data['image_url']
                            if data.get('price') and price == 0:
                                price = data['price']
                            # First successful auto-detect wins — cache on pallet
                            if not pallet_domain and data.get('source_domain'):
                                pallet_domain = data['source_domain']
                                execute_db(
                                    "UPDATE pallets SET amazon_domain = ? WHERE id = ?",
                                    (pallet_domain, pallet_id)
                                )
                            scraped += 1
                    except Exception as e:
                        print(f"[WARN] Scrape failed for {asin}: {e}")

                execute_db(
                    "INSERT INTO products (pallet_id, name, asin, ean, quantity, "
                    "condition, ebay_price_gbp, image_url, supplier_description) "
                    "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
                    (pallet_id, name, asin, ean, qty, cond, price, image_url,
                     supplier_desc)
                )
                count += 1

            msg = f'Imported {count} products'
            if scraped > 0:
                where = pallet_domain or 'Amazon (auto-detected)'
                msg += f' (scraped {scraped} from {where})'

            # Run auto-pipeline (all images, AI titles/descriptions, drafts)
            if count > 0:
                try:
                    processed, drafts = auto_process_products(pallet_id)
                    msg += f'. Auto-pipeline: {processed} scraped, {drafts} drafts created'
                except Exception as e:
                    msg += f'. Auto-pipeline error: {e}'

            flash(msg + '.', 'success')
            return redirect(url_for('pallet_detail', pallet_id=pallet_id))

        except Exception as e:
            flash(f'Error importing file: {e}', 'error')
            return redirect(url_for('csv_import', pallet_id=pallet_id))

    return render_page(
        TEMPLATE_CSV_IMPORT_CONTENT,
        page_title=f'Import CSV - {pallet["name"]} - eBay Hub UK',
        active_page='pallets',
        pallet=pallet
    )


# -------------------------------------------------------------------
# Product Detail Template
# -------------------------------------------------------------------
TEMPLATE_PRODUCT_DETAIL_CONTENT = """
<div class="page-header">
    <h1>
        <a href="/pallets" class="text-muted" style="text-decoration: none;">Pallets</a>
        <span class="text-muted">/</span>
        {% if pallet %}
        <a href="/pallet/{{ pallet.id }}" class="text-muted" style="text-decoration: none;">{{ pallet.name }}</a>
        <span class="text-muted">/</span>
        {% endif %}
        <span>{{ product.name[:40] }}{% if product.name|length > 40 %}...{% endif %}</span>
    </h1>
</div>

{% if colour_mismatch %}
<!-- Colour-mismatch warning: Amazon variant metadata doesn't line up with what the
     supplier says is physically in the pallet. Classic case: Amazon ASIN is the
     "Champagne / Warm white" variant, supplier shipped the "Green" variant. -->
<div class="card" style="border:1px solid rgba(245,158,11,0.4);background:rgba(245,158,11,0.08);margin-bottom:16px">
    <div style="display:flex;gap:12px;align-items:flex-start">
        <span class="material-symbols-outlined" style="color:#f59e0b;font-size:28px;flex-shrink:0">warning</span>
        <div style="flex:1">
            <div style="font-weight:700;color:#f59e0b;margin-bottom:4px">Colour mismatch — check before listing</div>
            <div style="font-size:0.85rem;color:var(--text);line-height:1.5">
                Supplier says: <strong>{{ colour_mismatch.supplier|join(', ') }}</strong>.
                Amazon says: <strong>{{ colour_mismatch.amazon|join(', ') }}</strong>.<br>
                The ASIN on Amazon is probably a different variant than what's in the pallet.
                Upload your own photos below and the AI will lean on the supplier description
                instead of Amazon's variant metadata.
            </div>
        </div>
    </div>
</div>
{% endif %}

<div class="detail-header">
    <div style="flex-shrink:0;">
        {% set amazon_images_raw = product.images|default('', true) %}
        {% set has_any_image = (amazon_images_raw and amazon_images_raw not in ('[]', '')) or custom_images or product.image_url %}
        {% if has_any_image %}
        <!-- 3D Image Carousel -->
        <style>
        .carousel-3d{position:relative;height:520px;display:flex;align-items:center;justify-content:center;perspective:1200px;overflow:visible;margin:0 auto;max-width:900px}
        .carousel-3d .c3d-slide{position:absolute;transition:all 0.5s ease;border-radius:14px;overflow:hidden;background:rgba(255,255,255,0.97);display:flex;align-items:center;justify-content:center;padding:12px}
        .carousel-3d .c3d-slide img{max-width:100%;max-height:100%;object-fit:contain}
        .carousel-3d .c3d-center{width:420px;height:480px;z-index:10;transform:translateX(0) scale(1);box-shadow:0 25px 70px rgba(0,0,0,0.5);cursor:zoom-in}
        .carousel-3d .c3d-left{width:250px;height:320px;z-index:5;transform:translateX(-280px) rotateY(30deg) scale(0.8);opacity:0.6;filter:brightness(0.6);cursor:pointer}
        .carousel-3d .c3d-right{width:250px;height:320px;z-index:5;transform:translateX(280px) rotateY(-30deg) scale(0.8);opacity:0.6;filter:brightness(0.6);cursor:pointer}
        .carousel-3d .c3d-hidden{width:180px;height:220px;z-index:1;opacity:0;transform:scale(0.3)}
        .c3d-arrow{position:absolute;top:50%;transform:translateY(-50%);z-index:20;width:44px;height:44px;border-radius:50%;background:rgba(143,245,255,0.15);border:1px solid rgba(143,245,255,0.3);color:#8ff5ff;display:flex;align-items:center;justify-content:center;cursor:pointer;font-size:1.4rem;transition:all 0.2s;backdrop-filter:blur(8px)}
        .c3d-arrow:hover{background:rgba(143,245,255,0.3);box-shadow:0 0 20px rgba(143,245,255,0.3)}
        .c3d-arrow.left{left:8px}
        .c3d-arrow.right{right:8px}
        .c3d-dots{display:flex;gap:6px;justify-content:center;margin-top:12px}
        .c3d-dot{width:8px;height:8px;border-radius:50%;background:rgba(255,255,255,0.15);cursor:pointer;transition:all 0.3s}
        .c3d-dot.active{background:#8ff5ff;box-shadow:0 0 8px rgba(143,245,255,0.5)}
        @media(max-width:600px){.carousel-3d{height:320px}.carousel-3d .c3d-center{width:240px;height:280px}.carousel-3d .c3d-left,.carousel-3d .c3d-right{width:140px;height:180px;transform:translateX(-160px) rotateY(25deg) scale(0.8)}.carousel-3d .c3d-right{transform:translateX(160px) rotateY(-25deg) scale(0.8)}}
        </style>
        <div class="carousel-3d" id="carousel3d">
            <div class="c3d-arrow left" onclick="c3dNav(-1)"><span class="material-symbols-outlined">chevron_left</span></div>
            <div class="c3d-arrow right" onclick="c3dNav(1)"><span class="material-symbols-outlined">chevron_right</span></div>
        </div>
        <div class="c3d-dots" id="c3dDots"></div>
        <script>
        (function(){
            // Custom (user-uploaded) photos take priority — they're the actual
            // pallet contents, not Amazon's variant stock shots.
            var customImgs = {{ custom_images|tojson }};
            var customUrls = (customImgs || []).map(function(p){ return '/static/' + p; });
            var amazonImgs = [];
            try { amazonImgs = JSON.parse({{ amazon_images_raw|tojson }}); } catch(e){}
            if (!amazonImgs || !amazonImgs.length) {
                amazonImgs = ['{{ product.image_url or "" }}'];
            }
            var imgs = customUrls.concat(amazonImgs);
            imgs = imgs.filter(function(u){ return u && u.length > 5; });
            if (imgs.length === 0) return;

            var idx = 0;
            var container = document.getElementById('carousel3d');
            var dots = document.getElementById('c3dDots');

            function render() {
                // Clear slides
                var old = container.querySelectorAll('.c3d-slide');
                old.forEach(function(el){ el.remove(); });

                var positions = ['c3d-hidden','c3d-left','c3d-center','c3d-right','c3d-hidden'];
                var offsets = [-2,-1,0,1,2];

                offsets.forEach(function(off, pi){
                    var i = (idx + off + imgs.length) % imgs.length;
                    var slide = document.createElement('div');
                    slide.className = 'c3d-slide ' + positions[pi];
                    var img = document.createElement('img');
                    img.src = imgs[i];
                    img.onerror = function(){ this.style.display='none'; };
                    if (positions[pi] === 'c3d-center') {
                        img.onclick = function(){ window.open(imgs[idx], '_blank'); };
                    } else if (positions[pi] === 'c3d-left') {
                        slide.style.cursor = 'pointer';
                        slide.onclick = function(){ c3dNav(-1); };
                    } else if (positions[pi] === 'c3d-right') {
                        slide.style.cursor = 'pointer';
                        slide.onclick = function(){ c3dNav(1); };
                    }
                    slide.appendChild(img);
                    container.appendChild(slide);
                });

                // Update dots
                dots.innerHTML = '';
                imgs.forEach(function(_, i){
                    var d = document.createElement('div');
                    d.className = 'c3d-dot' + (i === idx ? ' active' : '');
                    d.onclick = function(){ idx = i; render(); };
                    dots.appendChild(d);
                });
            }

            window.c3dNav = function(dir) {
                idx = (idx + dir + imgs.length) % imgs.length;
                render();
            };

            // Swipe support
            var startX = 0;
            container.addEventListener('touchstart', function(e){ startX = e.touches[0].clientX; });
            container.addEventListener('touchend', function(e){
                var diff = e.changedTouches[0].clientX - startX;
                if (Math.abs(diff) > 50) c3dNav(diff > 0 ? -1 : 1);
            });

            // Keyboard
            document.addEventListener('keydown', function(e){
                if (e.key === 'ArrowLeft') c3dNav(-1);
                if (e.key === 'ArrowRight') c3dNav(1);
            });

            render();
        })();
        </script>
        {% elif product.image_url %}
        <div style="text-align:center">
            <img src="{{ product.image_url }}" alt="{{ product.name }}"
                 style="width:100%;max-height:500px;object-fit:contain;border-radius:12px;background:rgba(255,255,255,0.95);cursor:zoom-in"
                 onclick="window.open(this.src,'_blank')"
                 onerror="this.style.display='none'">
        </div>
        {% endif %}
    </div>
    <div class="detail-info">
        <h2>{{ product.name }}</h2>
        <dl class="detail-meta">
            <dt>ASIN</dt><dd>{{ product.asin or '-' }}</dd>
            <dt>EAN</dt><dd>{{ product.ean or '-' }}</dd>
            <dt>Condition</dt><dd>{{ condition_label(product.condition) }}</dd>
            <dt>Quantity</dt><dd>{{ product.quantity }}</dd>
            <dt>eBay Price</dt><dd class="text-lime">{{ fmt_gbp(product.ebay_price_gbp) }}</dd>
            <dt>Status</dt><dd><span class="badge {{ status_color(product.status) }}">{{ product.status }}</span></dd>
            <dt>Category</dt><dd>{{ product.category or '-' }}</dd>
            <dt>Added</dt><dd>{{ fmt_datetime(product.created_at) }}</dd>
        </dl>
    </div>
</div>

<!-- Your Own Photos — override Amazon variant shots with actual pallet contents.
     Custom uploads appear FIRST in the carousel and when publishing to eBay. -->
<div class="section-title">Your Photos</div>
<div class="card">
    <div style="font-size:0.85rem;color:var(--text-muted);margin-bottom:12px">
        Upload pictures of what's physically in the pallet. These appear before Amazon photos
        in the listing — useful when the Amazon ASIN shows a different colour/variant.
        Max 8 MB per file. JPG/PNG/WebP/GIF.
    </div>
    <form method="POST" action="/product/{{ product.id }}/upload-images" enctype="multipart/form-data">
        <div class="form-group">
            <input type="file" name="images" accept="image/jpeg,image/png,image/webp,image/gif"
                   multiple class="form-control"
                   style="padding:10px;cursor:pointer">
        </div>
        <button type="submit" class="btn btn-purple btn-sm">
            <span class="material-symbols-outlined">upload</span> Upload Photos
        </button>
    </form>
    {% if custom_images %}
    <div style="display:grid;grid-template-columns:repeat(auto-fill,minmax(140px,1fr));gap:10px;margin-top:16px">
        {% for rel in custom_images %}
        <div style="position:relative;border-radius:8px;overflow:hidden;background:rgba(255,255,255,0.95);aspect-ratio:1;display:flex;align-items:center;justify-content:center">
            <img src="/static/{{ rel }}" style="max-width:100%;max-height:100%;object-fit:contain;cursor:zoom-in"
                 onclick="window.open('/static/{{ rel }}','_blank')">
            <form method="POST" action="/product/{{ product.id }}/delete-image"
                  style="position:absolute;top:4px;right:4px;margin:0"
                  onsubmit="return confirm('Delete this photo?')">
                <input type="hidden" name="path" value="{{ rel }}">
                <button type="submit"
                        style="width:28px;height:28px;border-radius:50%;border:none;background:rgba(255,68,68,0.9);color:white;cursor:pointer;display:flex;align-items:center;justify-content:center;padding:0"
                        title="Delete photo">
                    <span class="material-symbols-outlined" style="font-size:18px">close</span>
                </button>
            </form>
        </div>
        {% endfor %}
    </div>
    {% endif %}
</div>

<!-- Supplier Description — the ground truth from the joblot spec. Shown so the
     uncle can see what the AI will anchor the listing on, vs the Amazon blurb. -->
{% if product.supplier_description %}
<div class="section-title">Supplier Description</div>
<div class="card">
    <div style="font-size:0.8rem;color:var(--text-muted);margin-bottom:6px">
        From the supplier's CSV/ODS. Used as the AI's source of truth for what's in the pallet.
    </div>
    <div style="font-size:0.9rem;line-height:1.5;white-space:pre-wrap">{{ product.supplier_description }}</div>
</div>
{% endif %}

<!-- Item Specifics -->
{% set specs_raw = product.item_specifics|default('', true) %}
{% if specs_raw and specs_raw != '{}' and specs_raw != '' %}
<div class="section-title">Item Specifics</div>
<div class="card">
    <div class="table-wrap">
        <table>
            <thead><tr><th>Property</th><th>Value</th></tr></thead>
            <tbody id="specsTable"></tbody>
        </table>
    </div>
</div>
<script>
(function(){
    try {
        var specs = JSON.parse({{ specs_raw|tojson }});
        var tbody = document.getElementById('specsTable');
        for (var key in specs) {
            var tr = document.createElement('tr');
            var td1 = document.createElement('td');
            td1.textContent = key;
            td1.style.cssText = 'font-weight:600;color:#8ff5ff;width:40%;';
            var td2 = document.createElement('td');
            td2.textContent = specs[key];
            tr.appendChild(td1);
            tr.appendChild(td2);
            tbody.appendChild(tr);
        }
    } catch(e){}
})();
</script>
{% endif %}

<!-- Edit Product -->
<div class="section-title">Edit Product</div>
<div class="card">
    <form method="POST" action="/product/{{ product.id }}/update">
        <div class="form-row">
            <div class="form-group">
                <label class="form-label">Name</label>
                <input type="text" name="name" class="form-control" value="{{ product.name }}" required>
            </div>
            <div class="form-group">
                <label class="form-label">Status</label>
                <select name="status" class="form-control">
                    <option value="warehouse" {{ 'selected' if product.status == 'warehouse' }}>Warehouse</option>
                    <option value="listed" {{ 'selected' if product.status == 'listed' }}>Listed</option>
                    <option value="sold" {{ 'selected' if product.status == 'sold' }}>Sold</option>
                    <option value="shipped" {{ 'selected' if product.status == 'shipped' }}>Shipped</option>
                </select>
            </div>
        </div>
        <div class="form-row">
            <div class="form-group">
                <label class="form-label">ASIN</label>
                <input type="text" name="asin" class="form-control" value="{{ product.asin }}">
            </div>
            <div class="form-group">
                <label class="form-label">EAN</label>
                <input type="text" name="ean" class="form-control" value="{{ product.ean }}">
            </div>
        </div>
        <div class="form-row">
            <div class="form-group">
                <label class="form-label">Quantity</label>
                <input type="number" name="quantity" class="form-control" value="{{ product.quantity }}" min="1">
            </div>
            <div class="form-group">
                <label class="form-label">Condition</label>
                <select name="condition" class="form-control">
                    <option value="new" {{ 'selected' if product.condition == 'new' }}>New</option>
                    <option value="like_new" {{ 'selected' if product.condition == 'like_new' }}>Like New</option>
                    <option value="used" {{ 'selected' if product.condition == 'used' }}>Used</option>
                    <option value="damaged" {{ 'selected' if product.condition == 'damaged' }}>Damaged</option>
                </select>
            </div>
        </div>
        <div class="form-row">
            <div class="form-group">
                <label class="form-label">eBay Price (GBP)</label>
                <input type="number" step="0.01" name="ebay_price_gbp" class="form-control" value="{{ product.ebay_price_gbp }}">
            </div>
            <div class="form-group">
                <label class="form-label">Category</label>
                <input type="text" name="category" class="form-control" value="{{ product.category }}">
            </div>
        </div>

        <!-- Shipping & Dimensions -->
        <div style="border-top:1px solid rgba(255,255,255,0.06);margin-top:16px;padding-top:16px">
            <div style="font-size:0.8rem;font-weight:700;color:#f59e0b;margin-bottom:12px;display:flex;align-items:center;gap:6px">
                <span class="material-symbols-outlined" style="font-size:1rem">local_shipping</span> Shipping & Dimensions
            </div>
            <div class="form-row">
                <div class="form-group">
                    <label class="form-label">Weight (kg)</label>
                    <input type="number" step="0.01" name="weight_kg" class="form-control" value="{{ product.weight_kg or '' }}" placeholder="0.00">
                </div>
                <div class="form-group">
                    <label class="form-label">Length (cm)</label>
                    <input type="number" step="1" name="length_cm" class="form-control" value="{{ product.length_cm|int if product.length_cm else '' }}" placeholder="0">
                </div>
                <div class="form-group">
                    <label class="form-label">Width (cm)</label>
                    <input type="number" step="1" name="width_cm" class="form-control" value="{{ product.width_cm|int if product.width_cm else '' }}" placeholder="0">
                </div>
                <div class="form-group">
                    <label class="form-label">Height (cm)</label>
                    <input type="number" step="1" name="height_cm" class="form-control" value="{{ product.height_cm|int if product.height_cm else '' }}" placeholder="0">
                </div>
            </div>
            <div class="form-row">
                <div class="form-group">
                    <label class="form-label">Shipping Method</label>
                    <select name="shipping_method" class="form-control">
                        <option value="" {{ 'selected' if not product.shipping_method }}>Use default (Settings)</option>
                        {% for group_label, items in shipping_groups %}
                        <optgroup label="{{ group_label }}">
                            {% for key, label in items %}
                            <option value="{{ key }}" {{ 'selected' if product.shipping_method == key }}>{{ label }}</option>
                            {% endfor %}
                        </optgroup>
                        {% endfor %}
                    </select>
                    <div class="form-hint">Limits (max weight / dimensions) shown next to each option.</div>
                </div>
                <div class="form-group">
                    <label class="form-label">Shipping Cost (GBP)</label>
                    <input type="number" step="0.01" name="shipping_cost_gbp" class="form-control" value="{{ product.shipping_cost_gbp or '' }}" placeholder="np. 5.99 — koszt kuriera">
                    <div class="form-hint">
                        <strong>0</strong> = "Free delivery" na eBay (wliczone w cenę, lepszy ranking — <em>wujek płaci kurierowi</em>).<br>
                        <strong>&gt; 0</strong> = kupujący dopłaca tę kwotę za wysyłkę (wujek nie dokłada).<br>
                        W trybie <strong>Calculated</strong> to pole jest ignorowane — eBay liczy sam.
                    </div>
                </div>
            </div>
            <div class="form-row">
                <div class="form-group">
                    <label class="form-label">Shipping Pricing Mode (for this product)</label>
                    <select name="shipping_pricing_mode" class="form-control">
                        <option value="" {{ 'selected' if not product.shipping_pricing_mode }}>Use default (Settings)</option>
                        <option value="flat" {{ 'selected' if product.shipping_pricing_mode == 'flat' }}>Flat rate (ręczna cena powyżej)</option>
                        <option value="calculated" {{ 'selected' if product.shipping_pricing_mode == 'calculated' }}>Calculated (eBay liczy z wagi/wymiarów)</option>
                    </select>
                    <div class="form-hint">Calculated wymaga wagi + wymiarów + Royal Mail/Parcelforce jako metody. Inaczej apka użyje Flat rate.</div>
                </div>
            </div>
        </div>

        <div class="d-flex gap-8" style="margin-top: 16px;">
            <button type="submit" class="btn btn-cyan">
                <span class="material-symbols-outlined">save</span> Save Changes
            </button>
        </div>
    </form>
</div>

<!-- List on eBay -->
<div class="section-title">eBay Listing</div>
<div class="card">
    <form method="POST" action="/product/{{ product.id }}/list_ebay">
        <div class="form-group">
            <label class="form-label">Listing Title</label>
            <input type="text" name="title" class="form-control" value="{{ draft.title if draft else product.name }}" maxlength="80">
            <div class="form-hint">Max 80 characters. Make it descriptive for eBay search.</div>
        </div>
        <div class="form-group">
            <label class="form-label">Description</label>
            <textarea name="description" id="descInput" class="form-control" rows="6" placeholder="Product description for eBay listing..." oninput="updatePreview()">{{ draft.description if draft else '' }}</textarea>
            <div style="margin-top:8px">
                <button type="button" onclick="document.getElementById('descPreview').style.display=document.getElementById('descPreview').style.display==='none'?'block':'none'" class="btn btn-outline btn-sm" style="font-size:0.7rem">
                    <span class="material-symbols-outlined" style="font-size:0.85rem">visibility</span> Toggle Preview
                </button>
            </div>
            <div id="descPreview" style="display:none;margin-top:8px;padding:16px;background:#fff;color:#333;font-family:Arial,sans-serif;font-size:14px;line-height:1.6;border:1px solid #ddd;max-height:400px;overflow-y:auto">
                <div id="descPreviewContent" style="word-wrap:break-word">Click "Generate Description" or type HTML above to see preview</div>
            </div>
        </div>
        <div class="form-group">
            <label class="form-label">Price (GBP)</label>
            <input type="number" step="0.01" name="price" class="form-control" value="{{ draft.price_gbp if draft else product.ebay_price_gbp }}">
        </div>
        <div class="d-flex gap-8" style="margin-top: 16px;flex-wrap:wrap">
            <button type="submit" name="action" value="draft" class="btn btn-outline btn-sm" style="border-color:#f59e0b;color:#f59e0b">
                <span class="material-symbols-outlined">save</span> Save Draft
            </button>
            <button type="submit" name="action" value="publish" class="btn btn-lime"
                    onclick="return confirm('Publish to eBay? This listing will go LIVE immediately.')">
                <span class="material-symbols-outlined">sell</span> Publish to eBay
            </button>
            <button type="button" class="btn btn-outline btn-sm" onclick="generateAI('title')" id="genTitleBtn">
                <span class="material-symbols-outlined">auto_awesome</span> Generate Title
            </button>
            <button type="button" class="btn btn-outline btn-sm" onclick="generateAI('description')" id="genDescBtn">
                <span class="material-symbols-outlined">auto_awesome</span> Generate Description
            </button>
        </div>
    </form>
</div>
<script>
function updatePreview() {
    var html = document.getElementById('descInput').value;
    var preview = document.getElementById('descPreviewContent');
    if (html.trim()) {
        preview.innerHTML = html;
        document.getElementById('descPreview').style.display = 'block';
    }
}
// Auto-show preview if draft has description
document.addEventListener('DOMContentLoaded', function() { updatePreview(); });
function generateAI(type) {
    var btn = document.getElementById(type === 'title' ? 'genTitleBtn' : 'genDescBtn');
    var oldText = btn.innerHTML;
    btn.innerHTML = '<span class="material-symbols-outlined">hourglass_empty</span> Generating...';
    btn.disabled = true;
    var name = document.querySelector('[name="title"]').value || '{{ product.name }}';
    fetch('/api/generate-' + type, {
        method: 'POST',
        headers: {'Content-Type': 'application/json'},
        body: JSON.stringify({product_id: {{ product.id }}, product_name: name, condition: '{{ product.condition }}'})
    }).then(r => r.json()).then(d => {
        if (d.ok) {
            if (type === 'title') document.querySelector('[name="title"]').value = d.text;
            else { document.querySelector('[name="description"]').value = d.text; updatePreview(); }
            btn.innerHTML = '<span class="material-symbols-outlined">check</span> Done!';
            setTimeout(() => { btn.innerHTML = oldText; btn.disabled = false; }, 2000);
        } else {
            alert('Error: ' + (d.error || 'AI generation failed'));
            btn.innerHTML = oldText; btn.disabled = false;
        }
    }).catch(e => { alert('Error: ' + e); btn.innerHTML = oldText; btn.disabled = false; });
}
</script>
<!-- Listings History -->
{% if listings %}
<div class="section-title">Listing History</div>
<div class="table-wrap">
    <table>
        <thead>
            <tr>
                <th>Title</th>
                <th>Price</th>
                <th>Status</th>
                <th>eBay ID</th>
                <th>Views</th>
                <th>Watchers</th>
                <th>Created</th>
            </tr>
        </thead>
        <tbody>
            {% for l in listings %}
            <tr>
                <td>{{ l.title[:50] }}</td>
                <td>{{ fmt_gbp(l.price_gbp) }}</td>
                <td><span class="badge {{ status_color(l.status) }}">{{ l.status }}</span></td>
                <td class="text-muted">{{ l.ebay_item_id or '-' }}</td>
                <td>{{ l.views }}</td>
                <td>{{ l.watchers }}</td>
                <td class="text-muted">{{ fmt_datetime(l.created_at) }}</td>
            </tr>
            {% endfor %}
        </tbody>
    </table>
</div>
{% endif %}

<!-- Sell Privately — when the uncle sells in person, to a friend, at a local
     market, or anywhere outside eBay. Still counts toward revenue/profit.
     Shown whenever no sale has been recorded yet — even if the product was
     manually set to 'sold' (so uncle can backfill a private sale after the fact). -->
{% if not sales %}
<div class="section-title">Sprzedaż prywatna</div>
<div class="card">
    <div style="font-size:0.85rem;color:var(--text-muted);margin-bottom:12px">
        Use this when the item is sold outside eBay (in person, cash, to a friend).
        It will be added to revenue on the dashboard and marked as sold.
        {% if product.status in ('sold', 'shipped') %}
        <br><span style="color:var(--warning)">Status is already "{{ product.status }}" but no sale record exists —
        fill this in to backfill the revenue.</span>
        {% endif %}
    </div>
    <form method="POST" action="/product/{{ product.id }}/sell-private"
          onsubmit="return confirm('Record private sale for £' + this.price_gbp.value + '?')">
        <div class="form-row">
            <div class="form-group">
                <label class="form-label">Sale price (GBP) *</label>
                <input type="number" name="price_gbp" step="0.01" min="0.01"
                       class="form-control" required
                       value="{{ '%.2f'|format(product.ebay_price_gbp or 0) }}">
            </div>
            <div class="form-group">
                <label class="form-label">Buyer (optional)</label>
                <input type="text" name="buyer" class="form-control"
                       placeholder="e.g. John (friend) / local market">
            </div>
        </div>
        <div class="form-group">
            <label class="form-label">Notes (optional)</label>
            <input type="text" name="notes" class="form-control"
                   placeholder="e.g. paid cash, collected in person">
        </div>
        <div style="display:flex;gap:12px;align-items:center;margin-top:8px">
            <label style="display:flex;gap:6px;align-items:center;cursor:pointer;font-size:0.9rem">
                <input type="checkbox" name="mark_shipped" value="1" checked>
                Already handed over (mark as shipped)
            </label>
            <button type="submit" class="btn btn-lime">
                <span class="material-symbols-outlined">payments</span>
                Record private sale
            </button>
        </div>
    </form>
</div>
{% endif %}

<!-- Sales History -->
{% if sales %}
<div class="section-title">Sales History</div>
<div class="table-wrap">
    <table>
        <thead>
            <tr>
                <th>Buyer</th>
                <th>Price</th>
                <th>Source</th>
                <th>Status</th>
                <th>Sold</th>
                <th>Shipped</th>
                <th>Notes</th>
            </tr>
        </thead>
        <tbody>
            {% for s in sales %}
            <tr>
                <td>{{ s.buyer or 'Unknown' }}</td>
                <td class="text-lime">{{ fmt_gbp(s.price_gbp) }}</td>
                <td>
                    {% if (s.source or 'ebay') == 'private' %}
                    <span class="badge" style="background:rgba(190,238,0,0.15);color:var(--lime);border:1px solid rgba(190,238,0,0.3)">PRIVATE</span>
                    {% else %}
                    <span class="badge" style="background:rgba(143,245,255,0.12);color:var(--cyan);border:1px solid rgba(143,245,255,0.25)">eBay</span>
                    {% endif %}
                </td>
                <td><span class="badge {{ status_color(s.status) }}">{{ s.status }}</span></td>
                <td class="text-muted">{{ fmt_datetime(s.sold_at) }}</td>
                <td class="text-muted">{{ fmt_datetime(s.shipped_at) }}</td>
                <td class="text-muted" style="max-width:200px;overflow:hidden;text-overflow:ellipsis;">
                    {{ s.notes or '-' }}
                </td>
            </tr>
            {% endfor %}
        </tbody>
    </table>
</div>
{% endif %}

<!-- Delete Product -->
<div style="margin-top: 40px; padding-top: 20px; border-top: 1px solid var(--border);">
    <form method="POST" action="/product/{{ product.id }}/delete"
          onsubmit="return confirm('Delete this product? This cannot be undone.')">
        <button type="submit" class="btn btn-danger btn-sm">
            <span class="material-symbols-outlined">delete</span> Delete Product
        </button>
    </form>
</div>
"""

@app.route('/product/<int:product_id>')
def product_detail(product_id):
    product = query_db("SELECT * FROM products WHERE id = ?", (product_id,), one=True)
    if not product:
        flash('Product not found.', 'error')
        return redirect(url_for('pallets_list'))

    pallet = None
    if product['pallet_id']:
        pallet = query_db(
            "SELECT * FROM pallets WHERE id = ?",
            (product['pallet_id'],), one=True
        )

    listings = query_db(
        "SELECT * FROM ebay_listings WHERE product_id = ? ORDER BY created_at DESC",
        (product_id,)
    )

    # Load existing draft for pre-filling the form
    draft = query_db(
        "SELECT * FROM ebay_listings WHERE product_id = ? AND status = 'draft' ORDER BY created_at DESC LIMIT 1",
        (product_id,), one=True
    )

    sales = query_db(
        "SELECT * FROM sales WHERE product_id = ? ORDER BY sold_at DESC",
        (product_id,)
    )

    custom_images = _load_custom_images(product)
    mismatch = detect_colour_mismatch(product)
    return render_page(
        TEMPLATE_PRODUCT_DETAIL_CONTENT,
        page_title=f'{product["name"]} - eBay Hub UK',
        active_page='pallets',
        product=product, pallet=pallet, listings=listings, sales=sales, draft=draft,
        shipping_groups=get_shipping_options_grouped(),
        custom_images=custom_images, colour_mismatch=mismatch,
    )


# -------------------------------------------------------------------
# Listings Template
# -------------------------------------------------------------------
TEMPLATE_LISTINGS_CONTENT = """
<div class="page-header">
    <h1><span>Listings</span></h1>
</div>

<div class="filter-tabs">
    <a href="/listings" class="filter-tab {{ 'active' if current_filter == 'all' }}">
        All <span class="filter-count">{{ counts.total }}</span>
    </a>
    <a href="/listings?status=active" class="filter-tab {{ 'active' if current_filter == 'active' }}">
        Active <span class="filter-count">{{ counts.active }}</span>
    </a>
    <a href="/listings?status=draft" class="filter-tab {{ 'active' if current_filter == 'draft' }}">
        Draft <span class="filter-count">{{ counts.draft }}</span>
    </a>
    <a href="/listings?status=sold" class="filter-tab {{ 'active' if current_filter == 'sold' }}">
        Sold <span class="filter-count">{{ counts.sold }}</span>
    </a>
    <a href="/listings?status=ended" class="filter-tab {{ 'active' if current_filter == 'ended' }}">
        Ended <span class="filter-count">{{ counts.ended }}</span>
    </a>
</div>

{% if listings %}
<div class="table-wrap">
    <table>
        <thead>
            <tr>
                <th></th>
                <th>Title</th>
                <th>Price</th>
                <th>Status</th>
                <th>eBay ID</th>
                <th>Views</th>
                <th>Watchers</th>
                <th>Created</th>
            </tr>
        </thead>
        <tbody>
            {% for l in listings %}
            <tr>
                <td style="width:40px;">
                    {% if l.image_url %}
                    <img src="{{ l.image_url }}" style="width:36px;height:36px;object-fit:contain;border-radius:4px;"
                         onerror="this.style.display='none'">
                    {% endif %}
                </td>
                <td>
                    {% if l.product_id %}
                    <a href="/product/{{ l.product_id }}" class="table-link">{{ l.title[:60] }}</a>
                    {% else %}
                    {{ l.title[:60] }}
                    {% endif %}
                </td>
                <td>{{ fmt_gbp(l.price_gbp) }}</td>
                <td><span class="badge {{ status_color(l.status) }}">{{ l.status }}</span></td>
                <td class="text-muted">{{ l.ebay_item_id or '-' }}</td>
                <td>{{ l.views }}</td>
                <td>{{ l.watchers }}</td>
                <td class="text-muted">{{ fmt_datetime(l.created_at) }}</td>
            </tr>
            {% endfor %}
        </tbody>
    </table>
</div>
{% else %}
<div class="empty-state">
    <span class="material-symbols-outlined">sell</span>
    <p>No listings yet. List products from your pallets to see them here.</p>
    <a href="/pallets" class="btn btn-cyan">
        <span class="material-symbols-outlined">inventory_2</span> Go to Pallets
    </a>
</div>
{% endif %}
"""

@app.route('/listings')
def listings_list():
    status_filter = request.args.get('status', 'all')

    if status_filter != 'all':
        listings = query_db("""
            SELECT l.*, p.name as product_name, p.image_url
            FROM ebay_listings l
            LEFT JOIN products p ON p.id = l.product_id
            WHERE l.status = ?
            ORDER BY l.created_at DESC
        """, (status_filter,))
    else:
        listings = query_db("""
            SELECT l.*, p.name as product_name, p.image_url
            FROM ebay_listings l
            LEFT JOIN products p ON p.id = l.product_id
            ORDER BY l.created_at DESC
        """)

    counts = query_db("""
        SELECT
            COUNT(*) as total,
            SUM(CASE WHEN status='active' THEN 1 ELSE 0 END) as active,
            SUM(CASE WHEN status='draft' THEN 1 ELSE 0 END) as draft,
            SUM(CASE WHEN status='ended' THEN 1 ELSE 0 END) as ended,
            SUM(CASE WHEN status='sold' THEN 1 ELSE 0 END) as sold
        FROM ebay_listings
    """, one=True)

    return render_page(
        TEMPLATE_LISTINGS_CONTENT,
        page_title='Listings - eBay Hub UK',
        active_page='listings',
        listings=listings, counts=counts, current_filter=status_filter
    )


# -------------------------------------------------------------------
# Orders Template
# -------------------------------------------------------------------
TEMPLATE_ORDERS_CONTENT = """
<div class="page-header">
    <h1><span>Orders</span></h1>
</div>

<div class="filter-tabs">
    <a href="/orders" class="filter-tab {{ 'active' if current_filter == 'all' }}">
        All <span class="filter-count">{{ counts.total }}</span>
    </a>
    <a href="/orders?status=new" class="filter-tab {{ 'active' if current_filter == 'new' }}">
        To Ship <span class="filter-count" style="background:rgba(255,107,155,0.2);color:var(--pink);">{{ counts.new_orders }}</span>
    </a>
    <a href="/orders?status=shipped" class="filter-tab {{ 'active' if current_filter == 'shipped' }}">
        Shipped <span class="filter-count">{{ counts.shipped }}</span>
    </a>
    <a href="/orders?status=delivered" class="filter-tab {{ 'active' if current_filter == 'delivered' }}">
        Delivered <span class="filter-count">{{ counts.delivered }}</span>
    </a>
</div>

{% if orders %}
<div class="table-wrap">
    <table>
        <thead>
            <tr>
                <th></th>
                <th>Product</th>
                <th>Buyer</th>
                <th>Price</th>
                <th>Source</th>
                <th>Status</th>
                <th>Sold</th>
                <th>Address</th>
                <th>Action</th>
            </tr>
        </thead>
        <tbody>
            {% for o in orders %}
            <tr style="{{ 'background:rgba(255,107,155,0.04);' if o.status == 'new' }}">
                <td style="width:40px;">
                    {% if o.image_url %}
                    <img src="{{ o.image_url }}" style="width:36px;height:36px;object-fit:contain;border-radius:4px;"
                         onerror="this.style.display='none'">
                    {% endif %}
                </td>
                <td>
                    {% if o.product_id %}
                    <a href="/product/{{ o.product_id }}" class="table-link">
                        {{ o.product_name or o.listing_title or 'Product #' ~ o.product_id }}
                    </a>
                    {% else %}
                    {{ o.listing_title or '-' }}
                    {% endif %}
                </td>
                <td>{{ o.buyer or '-' }}</td>
                <td class="text-lime">{{ fmt_gbp(o.price_gbp) }}</td>
                <td>
                    {% if (o.source or 'ebay') == 'private' %}
                    <span class="badge" style="background:rgba(190,238,0,0.15);color:var(--lime);border:1px solid rgba(190,238,0,0.3)">PRIVATE</span>
                    {% else %}
                    <span class="badge" style="background:rgba(143,245,255,0.12);color:var(--cyan);border:1px solid rgba(143,245,255,0.25)">eBay</span>
                    {% endif %}
                </td>
                <td><span class="badge {{ status_color(o.status) }}">{{ o.status }}</span></td>
                <td class="text-muted">{{ fmt_datetime(o.sold_at) }}</td>
                <td class="text-muted" style="max-width:200px;overflow:hidden;text-overflow:ellipsis;">
                    {{ o.shipping_address[:60] if o.shipping_address else (o.notes[:60] if o.notes else '-') }}
                </td>
                <td>
                    {% if o.status == 'new' %}
                    <form method="POST" action="/order/{{ o.id }}/ship" class="inline-form">
                        <button type="submit" class="btn btn-pink btn-sm">
                            <span class="material-symbols-outlined">local_shipping</span> Ship
                        </button>
                    </form>
                    {% elif o.status == 'shipped' %}
                    <span class="text-muted">In transit</span>
                    {% else %}
                    <span class="text-lime">Done</span>
                    {% endif %}
                </td>
            </tr>
            {% endfor %}
        </tbody>
    </table>
</div>
{% else %}
<div class="empty-state">
    <span class="material-symbols-outlined">local_shipping</span>
    <p>No orders yet. Sales will appear here when products are sold.</p>
</div>
{% endif %}
"""

@app.route('/orders')
def orders_list():
    status_filter = request.args.get('status', 'all')

    if status_filter != 'all':
        orders = query_db("""
            SELECT s.*, p.name as product_name, p.image_url,
                   l.title as listing_title
            FROM sales s
            LEFT JOIN products p ON p.id = s.product_id
            LEFT JOIN ebay_listings l ON l.id = s.listing_id
            WHERE s.status = ?
            ORDER BY s.sold_at DESC
        """, (status_filter,))
    else:
        orders = query_db("""
            SELECT s.*, p.name as product_name, p.image_url,
                   l.title as listing_title
            FROM sales s
            LEFT JOIN products p ON p.id = s.product_id
            LEFT JOIN ebay_listings l ON l.id = s.listing_id
            ORDER BY s.sold_at DESC
        """)

    counts = query_db("""
        SELECT
            COUNT(*) as total,
            SUM(CASE WHEN status='new' THEN 1 ELSE 0 END) as new_orders,
            SUM(CASE WHEN status='shipped' THEN 1 ELSE 0 END) as shipped,
            SUM(CASE WHEN status='delivered' THEN 1 ELSE 0 END) as delivered
        FROM sales
    """, one=True)

    return render_page(
        TEMPLATE_ORDERS_CONTENT,
        page_title='Orders - eBay Hub UK',
        active_page='orders',
        orders=orders, counts=counts, current_filter=status_filter
    )


# -------------------------------------------------------------------
# Settings Template
# -------------------------------------------------------------------
TEMPLATE_SETTINGS_CONTENT = """
<div class="page-header">
    <h1><span>Settings</span></h1>
</div>

<form method="POST" action="/settings">

    <!-- eBay API -->
    <div class="card mb-16">
        <div class="card-title" style="display:flex;align-items:center;gap:8px;">
            <span class="material-symbols-outlined text-cyan" style="font-size:20px;">api</span>
            eBay API Credentials
        </div>
        <p class="text-muted" style="font-size:0.85rem;margin:8px 0 16px 0;">
            Get your API keys from
            <span class="text-cyan">developer.ebay.com</span>.
            These are required to publish listings and sync orders.
        </p>
        <div class="form-row">
            <div class="form-group">
                <label class="form-label">App ID (Client ID)</label>
                <input type="text" name="ebay_app_id" class="form-control"
                       value="{{ config.ebay_app_id }}" placeholder="Your eBay App ID">
            </div>
            <div class="form-group">
                <label class="form-label">Cert ID (Client Secret)</label>
                <input type="password" name="ebay_cert_id" class="form-control"
                       value="{{ config.ebay_cert_id }}" placeholder="Your eBay Cert ID">
            </div>
        </div>
        <div class="form-row">
            <div class="form-group">
                <label class="form-label">Dev ID</label>
                <input type="text" name="ebay_dev_id" class="form-control"
                       value="{{ config.ebay_dev_id }}" placeholder="Your eBay Dev ID">
            </div>
            <div class="form-group">
                <label class="form-label">User Token</label>
                <input type="password" name="ebay_user_token" class="form-control"
                       value="{{ config.ebay_user_token }}" placeholder="OAuth user token">
            </div>
        </div>
    </div>

    <!-- Gemini AI -->
    <div class="card mb-16">
        <div class="card-title" style="display:flex;align-items:center;gap:8px;">
            <span class="material-symbols-outlined" style="font-size:20px;color:#f59e0b">auto_awesome</span>
            AI Title &amp; Description Generator
        </div>
        <p class="text-muted" style="font-size:0.85rem;margin:8px 0 16px 0;">
            Uses Google Gemini AI to generate eBay titles and descriptions. Get your key from
            <span class="text-cyan">aistudio.google.com</span>.
        </p>
        <div class="form-group">
            <label class="form-label">Gemini API Key</label>
            <input type="password" name="gemini_api_key" class="form-control"
                   value="{{ config.gemini_api_key }}" placeholder="AIzaSy...">
        </div>
    </div>

    <!-- Telegram -->
    <div class="card mb-16">
        <div class="card-title" style="display:flex;align-items:center;gap:8px;">
            <span class="material-symbols-outlined text-purple" style="font-size:20px;">send</span>
            Telegram Notifications
        </div>
        <p class="text-muted" style="font-size:0.85rem;margin:8px 0 16px 0;">
            Get notified about new sales and orders via Telegram bot.
        </p>
        <div class="form-row">
            <div class="form-group">
                <label class="form-label">Bot Token</label>
                <input type="text" name="telegram_bot_token" class="form-control"
                       value="{{ config.telegram_bot_token }}" placeholder="123456:ABC-DEF...">
                <div class="form-hint">Get from @BotFather on Telegram</div>
            </div>
            <div class="form-group">
                <label class="form-label">Chat ID</label>
                <input type="text" name="telegram_chat_id" class="form-control"
                       value="{{ config.telegram_chat_id }}" placeholder="-1001234567890">
            </div>
        </div>
    </div>

    <!-- Security -->
    <div class="card mb-16">
        <div class="card-title" style="display:flex;align-items:center;gap:8px;">
            <span class="material-symbols-outlined" style="font-size:20px;color:#ef4444">lock</span>
            Security
        </div>
        <div class="form-row">
            <div class="form-group">
                <label class="form-label">App PIN (required to access)</label>
                <input type="password" name="app_pin" class="form-control"
                       value="{{ config.app_pin }}" placeholder="Set a PIN (e.g. 1234)">
                <div class="form-hint">Leave empty to disable PIN protection</div>
            </div>
        </div>
    </div>

    <!-- Defaults -->
    <div class="card mb-16">
        <div class="card-title" style="display:flex;align-items:center;gap:8px;">
            <span class="material-symbols-outlined text-lime" style="font-size:20px;">tune</span>
            Default Settings
        </div>
        <div class="form-row">
            <div class="form-group">
                <label class="form-label">Default Shipping Method</label>
                <select name="default_shipping" class="form-control">
                    {% for group_label, items in shipping_groups %}
                    <optgroup label="{{ group_label }}">
                        {% for key, label in items %}
                        <option value="{{ key }}" {{ 'selected' if config.default_shipping == key }}>{{ label }}</option>
                        {% endfor %}
                    </optgroup>
                    {% endfor %}
                </select>
            </div>
            <div class="form-group">
                <label class="form-label">Returns Policy</label>
                <select name="returns_policy" class="form-control">
                    <option value="no" {{ 'selected' if (config.returns_policy or 'no') == 'no' }}>No returns accepted (private seller)</option>
                    <option value="14" {{ 'selected' if config.returns_policy == '14' }}>14 days</option>
                    <option value="30" {{ 'selected' if config.returns_policy == '30' }}>30 days</option>
                    <option value="60" {{ 'selected' if config.returns_policy == '60' }}>60 days</option>
                </select>
                <div class="form-hint">
                    <strong>No returns</strong> — OK dla <em>private seller</em> (wujek nie ma firmy). Kupujący nie zwróci bo mu się odwidziało.<br>
                    <strong>Business seller</strong> (zarejestrowana firma na eBay) musi dać ≥14 dni.<br>
                    <em>Uwaga:</em> eBay Money Back Guarantee i tak wymusi zwrot jeśli towar uszkodzony / nie zgadza się z opisem.
                </div>
            </div>
        </div>
        <div class="form-row">
            <div class="form-group">
                <label class="form-label">Shipping Pricing Mode</label>
                <select name="default_shipping_pricing" class="form-control">
                    <option value="flat" {{ 'selected' if (config.default_shipping_pricing or 'flat') == 'flat' }}>Flat rate (wpisujesz cenę ręcznie)</option>
                    <option value="calculated" {{ 'selected' if config.default_shipping_pricing == 'calculated' }}>Calculated (eBay liczy z wagi + kodu pocztowego)</option>
                </select>
                <div class="form-hint">
                    <strong>Flat</strong>: wpisujesz koszt wysyłki na każdym produkcie.<br>
                    <strong>Calculated</strong>: eBay liczy sam z wagi/wymiarów/kodu kupującego — działa tylko dla <em>Royal Mail</em> i <em>Parcelforce</em>. Dla innych kurierów (DPD, Evri, UPS...) apka sama cofa się do Flat.
                </div>
            </div>
            <div class="form-group">
                <label class="form-label">Seller Postcode (nadawcy)</label>
                <input type="text" name="seller_postcode" class="form-control"
                       value="{{ config.seller_postcode or '' }}" placeholder="np. SW1A 1AA" maxlength="10">
                <div class="form-hint">Kod pocztowy wujka (skąd wysyła). Wymagany dla Calculated Shipping.</div>
            </div>
        </div>
    </div>

    <button type="submit" class="btn btn-cyan">
        <span class="material-symbols-outlined">save</span> Save Settings
    </button>
</form>

<!-- Backups -->
<div class="card mb-16" style="margin-top:24px">
    <div class="card-title" style="display:flex;align-items:center;justify-content:space-between;">
        <div style="display:flex;align-items:center;gap:8px;">
            <span class="material-symbols-outlined text-lime" style="font-size:20px;">backup</span>
            Database Backups
        </div>
        <form method="POST" action="/settings/backup/create" class="inline-form">
            <button type="submit" class="btn btn-lime btn-sm">
                <span class="material-symbols-outlined">add_circle</span> Create Backup Now
            </button>
        </form>
    </div>
    <p class="text-muted" style="font-size:0.8rem;margin:8px 0 16px 0;">
        Auto-backup runs every hour. Max 48 backups kept (2 days).
    </p>

    {% if backups %}
    <div class="table-wrap">
        <table>
            <thead>
                <tr>
                    <th>Date</th>
                    <th>Size</th>
                    <th>File</th>
                    <th style="text-align:right">Actions</th>
                </tr>
            </thead>
            <tbody>
                {% for b in backups[:15] %}
                <tr>
                    <td>{{ b.date }}</td>
                    <td>{{ b.size_mb }} MB</td>
                    <td style="font-size:0.75rem;color:var(--text-muted);font-family:monospace">{{ b.name }}</td>
                    <td style="text-align:right">
                        <div class="d-flex gap-8" style="justify-content:flex-end">
                            <a href="/settings/backup/download/{{ b.name }}" class="btn btn-outline btn-sm" style="padding:4px 10px;font-size:0.7rem">
                                <span class="material-symbols-outlined" style="font-size:0.85rem">download</span> Download
                            </a>
                            <form method="POST" action="/settings/backup/restore/{{ b.name }}" class="inline-form" onsubmit="return confirm('Restore database from {{ b.name }}?\\n\\nCurrent data will be backed up first.')">
                                <button type="submit" class="btn btn-sm" style="padding:4px 10px;font-size:0.7rem;background:rgba(245,158,11,0.15);border:1px solid rgba(245,158,11,0.3);color:#f59e0b">
                                    <span class="material-symbols-outlined" style="font-size:0.85rem">restore</span> Restore
                                </button>
                            </form>
                        </div>
                    </td>
                </tr>
                {% endfor %}
            </tbody>
        </table>
    </div>
    {% if backups|length > 15 %}
    <div style="text-align:center;padding:8px;color:var(--text-muted);font-size:0.75rem">Showing 15 of {{ backups|length }} backups</div>
    {% endif %}
    {% else %}
    <div style="text-align:center;padding:20px;color:var(--text-muted)">No backups yet. Click "Create Backup Now" to make one.</div>
    {% endif %}
</div>

<!-- Upload Backup -->
<div class="card mb-16">
    <div class="card-title" style="display:flex;align-items:center;gap:8px;">
        <span class="material-symbols-outlined text-purple" style="font-size:20px;">upload_file</span>
        Upload Backup
    </div>
    <form method="POST" action="/settings/backup/upload" enctype="multipart/form-data">
        <div class="form-group">
            <label class="form-label">Upload a .db backup file to restore</label>
            <input type="file" name="backup_file" accept=".db" class="form-control" required style="padding:8px">
        </div>
        <button type="submit" class="btn btn-purple btn-sm" onclick="return confirm('Upload and restore this backup?\\n\\nCurrent data will be backed up first.')">
            <span class="material-symbols-outlined">upload</span> Upload & Restore
        </button>
    </form>
</div>
"""

@app.route('/settings', methods=['GET', 'POST'])
def settings():
    if request.method == 'POST':
        keys = [
            'app_pin',
            'ebay_app_id', 'ebay_cert_id', 'ebay_dev_id', 'ebay_user_token',
            'gemini_api_key',
            'telegram_bot_token', 'telegram_chat_id',
            'default_shipping', 'default_return_days',
            'default_shipping_pricing', 'seller_postcode',
            'returns_policy',
        ]
        for key in keys:
            val = request.form.get(key, '')
            set_config(key, val)
        flash('Settings saved.', 'success')
        return redirect(url_for('settings'))

    config = {}
    keys = [
        'app_pin',
        'ebay_app_id', 'ebay_cert_id', 'ebay_dev_id', 'ebay_user_token',
        'gemini_api_key',
        'telegram_bot_token', 'telegram_chat_id',
        'default_shipping', 'default_return_days',
        'default_shipping_pricing', 'seller_postcode',
        'returns_policy',
    ]
    for key in keys:
        config[key] = get_config(key, '')

    from modules.backup import get_backups
    backups = get_backups()

    return render_page(
        TEMPLATE_SETTINGS_CONTENT,
        page_title='Settings - eBay Hub UK',
        active_page='settings',
        config=config,
        backups=backups,
        shipping_groups=get_shipping_options_grouped(),
    )


@app.route('/settings/backup/create', methods=['POST'])
def backup_create():
    from modules.backup import create_backup
    result = create_backup()
    if result:
        flash(f'Backup created: {result.name}', 'success')
    else:
        flash('Backup failed!', 'error')
    return redirect(url_for('settings'))


@app.route('/settings/backup/restore/<backup_name>', methods=['POST'])
def backup_restore(backup_name):
    from modules.backup import restore_backup
    from modules.database import close_db
    ok, msg = restore_backup(backup_name)
    if ok:
        # Force close current DB connection so next request uses restored DB
        close_db()
        flash(f'Database restored from {backup_name}.', 'success')
    else:
        flash(f'Restore failed: {msg}', 'error')
    return redirect(url_for('settings'))


@app.route('/settings/backup/download/<backup_name>')
def backup_download(backup_name):
    from pathlib import Path
    backup_dir = Path(__file__).parent / 'backups'
    backup_path = backup_dir / backup_name
    if not backup_path.exists() or '..' in backup_name:
        flash('Backup not found.', 'error')
        return redirect(url_for('settings'))
    from flask import send_file
    return send_file(str(backup_path), as_attachment=True, download_name=backup_name)


@app.route('/settings/backup/upload', methods=['POST'])
def backup_upload():
    from modules.backup import create_backup
    from pathlib import Path
    import shutil

    file = request.files.get('backup_file')
    if not file or not file.filename.endswith('.db'):
        flash('Please upload a .db file.', 'error')
        return redirect(url_for('settings'))

    # Safety backup first
    create_backup()

    # Close current connection, replace DB, next request reconnects
    from modules.database import close_db
    close_db()
    db_path = Path(__file__).parent / 'ebay_hub.db'
    file.save(str(db_path))
    flash(f'Database restored from uploaded file: {file.filename}', 'success')
    return redirect(url_for('settings'))


# ===================================================================
# Help Page
# ===================================================================

TEMPLATE_HELP = """
<div class="page-header">
    <h1><span>Help & Guide</span></h1>
    <div style="display:flex;gap:8px;margin-left:auto">
        <button id="lang-pl" class="btn btn-lime btn-sm" onclick="setLang('pl')">Polski</button>
        <button id="lang-en" class="btn btn-outline btn-sm" onclick="setLang('en')">English</button>
    </div>
</div>

<style>
.help-lang { display: none; }
.help-lang.active { display: block; }
</style>

<script>
function setLang(lang) {
    document.querySelectorAll('.help-lang').forEach(function(el){
        el.classList.toggle('active', el.dataset.lang === lang);
    });
    var plBtn = document.getElementById('lang-pl');
    var enBtn = document.getElementById('lang-en');
    if (lang === 'pl') {
        plBtn.className = 'btn btn-lime btn-sm';
        enBtn.className = 'btn btn-outline btn-sm';
    } else {
        plBtn.className = 'btn btn-outline btn-sm';
        enBtn.className = 'btn btn-lime btn-sm';
    }
    try { localStorage.setItem('ebayhub_help_lang', lang); } catch(e){}
}
(function(){
    var saved = 'pl';
    try { saved = localStorage.getItem('ebayhub_help_lang') || 'pl'; } catch(e){}
    setLang(saved);
})();
</script>

<!-- =================================================================== -->
<!-- POLSKI                                                                -->
<!-- =================================================================== -->
<div class="help-lang" data-lang="pl">

<div class="card mb-16">
    <div class="card-title" style="color:#8ff5ff">Od czego zacząć — krok po kroku</div>
    <ol style="line-height:2.2;color:var(--text-muted);font-size:0.9rem">
        <li><strong>Dodaj paletę:</strong> Wejdź w <strong>Pallets</strong> → <strong>+ Add Pallet</strong> → wpisz nazwę, cenę zakupu (GBP) i dostawcę. Wgraj plik CSV/XLSX/ODS od dostawcy palety. <em>Amazon locale zostaw na „Auto-detect" — scraper sam znajdzie gdzie ASIN jest wystawiony.</em></li>
        <li><strong>Auto Pipeline:</strong> Na stronie palety kliknij <strong>Auto Pipeline</strong> — apka sama:
            <ul style="margin:4px 0 4px 20px;line-height:1.8">
                <li>Pobierze do 8 zdjęć każdego produktu z Amazon UK</li>
                <li>Wyciągnie specyfikację (marka, model, rozmiar itd.)</li>
                <li>Wygeneruje tytuły pod eBay przez AI (max 80 znaków)</li>
                <li>Wygeneruje profesjonalny opis HTML przez AI</li>
                <li>Dopasuje kategorię eBay</li>
                <li>Zrobi szkice aukcji gotowe do podglądu</li>
            </ul>
        </li>
        <li><strong>Ustaw ceny:</strong> W sekcji <strong>Set Prices</strong> wpisz ręcznie albo kliknij <strong>Apply Multiplier</strong> (np. koszt × 2.5 = twoja cena sprzedaży).</li>
        <li><strong>Sprawdź szkice:</strong> Wejdź w <strong>Listings</strong> → kliknij produkt żeby edytować tytuł/opis/cenę.</li>
        <li><strong>Wystaw:</strong> Wróć na paletę → kliknij <strong>Publish All</strong> → aukcje idą NA ŻYWO na eBay (zostaniesz zapytany o potwierdzenie).</li>
        <li><strong>Wyślij zamówienie:</strong> Gdy coś się sprzeda → <strong>Orders</strong> → <strong>Mark as Shipped</strong> + numer śledzenia.</li>
    </ol>
</div>

<div class="card mb-16">
    <div class="card-title" style="color:#8ff5ff">Amazon locale — działa automatycznie</div>
    <p style="color:var(--text-muted);font-size:0.9rem;line-height:1.8">
        Scraper <strong>sam wykrywa</strong>, na którym Amazonie jest wystawiony ASIN z twojego CSV-ka.
        Idzie po kolei przez Amazon.co.uk → .com → .de → .pl → .fr → .it → .es → .nl → .se i bierze pierwszą
        domenę, która ma ten ASIN jako prawdziwy produkt. Nic sam nie musisz ustawiać.
    </p>
    <p style="color:var(--text-muted);font-size:0.9rem;line-height:1.8">
        Po pierwszym udanym scrape'ie apka <strong>zapamiętuje wykrytą domenę</strong> na palecie — żeby
        następne scrape'y nie przechodziły znów przez wszystkie locale, tylko od razu szły do właściwego.
    </p>
    <p style="color:var(--text-muted);font-size:0.85rem;margin-top:8px">
        <strong>Po angielsku i tak jest:</strong> nawet jak ASIN okaże się na Amazon.de albo .fr, apka
        pobiera stronę w wersji angielskiej (URL <code>/-/en/</code>) — więc tytuły i bullet pointy
        zawsze są po angielsku pod eBay UK.
    </p>
    <p style="color:var(--text-muted);font-size:0.85rem;margin-top:8px">
        <strong>Ręczny override (opcjonalnie):</strong> jeżeli auto-detect trafi na zły wariant
        (np. na .co.uk pod tym samym ASIN jest biała choinka, a ty masz zieloną z .de) — wejdź w paletę,
        zmień <em>Amazon Locale</em> na konkretny (np. <code>amazon.de</code>), <strong>Save</strong>,
        i kliknij <strong>Scrape Images</strong> żeby pobrać na nowo. Żeby wrócić do auto, wybierz
        „Auto-detect" z dropdown.
    </p>
</div>

<div class="card mb-16">
    <div class="card-title" style="color:#beee00">Własne zdjęcia — ważne!</div>
    <p style="color:var(--text-muted);font-size:0.9rem;line-height:1.8">
        Amazon czasem pokazuje inny wariant produktu niż to co faktycznie jest w palecie
        (np. na Amazonie jest biała choinka, a w palecie zielona). Dlatego możesz dodać
        <strong>własne zdjęcia</strong> tego co naprawdę masz.
    </p>
    <ol style="line-height:2;color:var(--text-muted);font-size:0.9rem">
        <li>Wejdź w produkt (Pallets → paleta → produkt)</li>
        <li>Sekcja <strong>Your Photos</strong> — kliknij <strong>Choose Files</strong> i wybierz zdjęcia z telefonu/dysku</li>
        <li>Kliknij <strong>Upload</strong> — twoje zdjęcia pójdą jako <strong>pierwsze</strong> zarówno w karuzeli na stronie jak i w aukcji eBay</li>
        <li>Możesz usunąć każde zdjęcie osobno — ikona kosza pod zdjęciem</li>
    </ol>
    <p style="color:var(--warning);font-size:0.85rem;margin-top:8px">
        <strong>Uwaga:</strong> jeżeli apka wykryje że Amazon mówi co innego niż dostawca
        (np. inny kolor), zobaczysz żółte ostrzeżenie na górze strony produktu. Wtedy koniecznie
        zrób własne zdjęcia — AI też uwzględni opis od dostawcy jako prawdę.
    </p>
</div>

<div class="card mb-16">
    <div class="card-title" style="color:#f59e0b">Sprzedaż prywatna — jak dodać</div>
    <p style="color:var(--text-muted);font-size:0.9rem;line-height:1.8">
        Jeżeli sprzedałeś coś <strong>poza eBayem</strong> (znajomemu, na targu, za gotówkę w ręce) —
        wpisz to w apce, żeby przychód był kompletny na dashboardzie.
    </p>
    <ol style="line-height:2;color:var(--text-muted);font-size:0.9rem">
        <li>Wejdź w produkt (Pallets → paleta → produkt)</li>
        <li>Przewiń w dół — sekcja <strong>Sprzedaż prywatna</strong> (nad Edit Product)</li>
        <li>Wpisz <strong>cenę</strong> (GBP) — już wstępnie uzupełniona ceną z eBay, popraw jeśli sprzedałeś taniej</li>
        <li>Opcjonalnie: kto kupił (np. „Janek kolega") i notatka (np. „zapłacił gotówką")</li>
        <li>Zaznacz ✅ <strong>Already handed over</strong> jeśli już oddałeś towar kupującemu</li>
        <li>Kliknij <strong>Record private sale</strong> — potwierdź kwotę</li>
    </ol>
    <p style="color:var(--text-muted);font-size:0.85rem;margin-top:8px">
        <strong>Co się stanie:</strong> sprzedaż pojawi się w <strong>Orders</strong> z oznaczeniem
        <span class="badge" style="background:rgba(190,238,0,0.15);color:var(--lime);border:1px solid rgba(190,238,0,0.3)">PRIVATE</span>,
        produkt zostanie oznaczony jako sprzedany, a kwota wejdzie do Total Revenue na dashboardzie.
        Jeżeli już wcześniej ręcznie ustawiłeś status na „Sold" ale nie wpisałeś sprzedaży — sekcja
        nadal się pokaże z żółtym ostrzeżeniem, żeby można było uzupełnić wstecznie.
    </p>
</div>

<div class="card mb-16">
    <div class="card-title" style="color:#a855f7">Przyciski — co który robi</div>
    <div style="display:grid;grid-template-columns:170px 1fr;gap:10px 16px;font-size:0.85rem;color:var(--text-muted)">
        <div><span class="btn btn-purple btn-sm" style="font-size:0.7rem">Import CSV</span></div><div>Wgraj plik CSV/XLSX/ODS z produktami od dostawcy palety</div>
        <div><span class="btn btn-sm" style="font-size:0.7rem;background:linear-gradient(135deg,#8ff5ff,#a855f7);color:#000">Auto Pipeline</span></div><div><strong>Robi wszystko automatycznie!</strong> Zdjęcia, tytuły, opisy, kategorie, szkice — jednym kliknięciem</div>
        <div><span class="btn btn-outline btn-sm" style="font-size:0.7rem;border-color:#f59e0b;color:#f59e0b">Create Drafts</span></div><div>Zapisuje aukcje lokalnie — nic nie idzie jeszcze na eBay. Sprawdź przed publikacją</div>
        <div><span class="btn btn-lime btn-sm" style="font-size:0.7rem">Publish All</span></div><div>Wystawia wszystkie szkice na eBay NA ŻYWO! <strong>Upewnij się najpierw że ceny są ustawione!</strong></div>
        <div><span class="btn btn-outline btn-sm" style="font-size:0.7rem;border-color:#a855f7;color:#a855f7">Auto Categories</span></div><div>AI dopasowuje kategorie eBay do twoich produktów</div>
        <div><span class="btn btn-cyan btn-sm" style="font-size:0.7rem">Scrape Images</span></div><div>Pobiera ponownie zdjęcia produktów z Amazon UK</div>
        <div><span class="btn btn-outline btn-sm" style="font-size:0.7rem">Save Draft</span></div><div>Zapisuje tytuł, opis, cenę bez wystawiania na eBay</div>
        <div><span class="btn btn-lime btn-sm" style="font-size:0.7rem">Publish to eBay</span></div><div>Wystawia ten jeden produkt na eBay (od razu na żywo)</div>
        <div><span class="btn btn-outline btn-sm" style="font-size:0.7rem">Generate Title</span></div><div>AI generuje zoptymalizowany tytuł (max 80 znaków, po angielsku)</div>
        <div><span class="btn btn-outline btn-sm" style="font-size:0.7rem">Generate Description</span></div><div>AI generuje profesjonalny opis HTML z listą cech i bulletami</div>
        <div><span class="btn btn-lime btn-sm" style="font-size:0.7rem">Record private sale</span></div><div>Zapisuje sprzedaż prywatną (poza eBayem) — wchodzi do przychodu</div>
    </div>
</div>

<div class="card mb-16">
    <div class="card-title" style="color:#ff6b9b">Wysyłka — jak to działa</div>
    <p style="color:var(--text-muted);font-size:0.85rem;line-height:1.8">
        Gdy ktoś kupi twój produkt na eBay:<br><br>
        <strong>1.</strong> Zobaczysz zamówienie w zakładce <strong>Orders</strong> ze statusem <strong>„TO SHIP"</strong><br>
        <strong>2.</strong> Zapakuj produkt starannie<br>
        <strong>3.</strong> Zanieś na pocztę (Royal Mail) albo do punktu (Evri, DPD)<br>
        <strong>4.</strong> Weź numer śledzenia od kuriera<br>
        <strong>5.</strong> W apce kliknij <strong>Mark as Shipped</strong> i wpisz numer śledzenia<br>
        <strong>6.</strong> eBay przeleje ci pieniądze na konto (zwykle 2-3 dni po potwierdzeniu odbioru przez kupującego)<br><br>
        <strong style="color:#8ff5ff">Tip:</strong> Tańsze etykiety wysyłkowe kupisz bezpośrednio na eBay!<br>
        Wejdź: My eBay → Sold → Print Shipping Label — eBay ma zniżki z Royal Mail i Evri.<br><br>
        <strong style="color:#8ff5ff">Tip:</strong> Ustaw koszt wysyłki <strong>£0 (free postage)</strong> i wlicz go w cenę produktu — eBay wyżej pokazuje aukcje z darmową wysyłką!
    </p>
</div>

<div class="card mb-16">
    <div class="card-title" style="color:#f59e0b">Wskazówki jak sprzedawać więcej</div>
    <ul style="color:var(--text-muted);font-size:0.85rem;line-height:2.2">
        <li>Używaj <strong>darmowej wysyłki</strong> (£0) — eBay wyżej rankuje takie aukcje</li>
        <li>Używaj <strong>Auto Pipeline</strong> — AI generuje profesjonalne tytuły/opisy, oszczędzasz godziny pracy</li>
        <li>Dodaj <strong>dużo zdjęć</strong> — aukcje z 8+ fotkami mają 30% więcej sprzedaży na eBay UK</li>
        <li>Ustaw <strong>konkurencyjne ceny</strong> — sprawdź za ile podobne rzeczy poszły na eBay</li>
        <li>Wysyłaj <strong>w ciągu 1-2 dni</strong> — szybka wysyłka poprawia ocenę i ranking</li>
        <li>Zawsze podawaj <strong>numer śledzenia</strong> — chroni cię przed reklamacjami „nie doszło"</li>
        <li>Pisz tytuły ze <strong>słowami kluczowymi</strong> (marka, model, najważniejsze parametry)</li>
        <li>Nowe konta: eBay może <strong>trzymać pieniądze kilka dni</strong> — to normalne, po paru sprzedażach przestaje</li>
    </ul>
</div>

<div class="card mb-16">
    <div class="card-title" style="color:#ef4444">Typowe problemy</div>
    <div style="color:var(--text-muted);font-size:0.85rem;line-height:1.8">
        <p><strong>Błąd 120 — „Need additional information":</strong><br>Twoje konto sprzedawcy eBay nie jest do końca skonfigurowane. Wejdź na <a href="https://www.ebay.co.uk/sh/sell" target="_blank" style="color:#8ff5ff">ebay.co.uk/sh/sell</a> i dokończ dane bankowe + weryfikację tożsamości.</p>
        <p style="margin-top:12px"><strong>Błąd 107 — „Category not valid":</strong><br>Złe ID kategorii. Wejdź w produkt → zmień pole Category na prawidłowy numer kategorii eBay UK (np. 96915 dla ładowarek, 175673 dla części komputerowych).</p>
        <p style="margin-top:12px"><strong>Ostrzeżenie — „Funds on hold":</strong><br>To normalne dla nowych sprzedawców. eBay trzyma pieniądze dopóki kupujący nie potwierdzi dostawy. Po kilku udanych sprzedażach przestaje.</p>
        <p style="margin-top:12px"><strong>Nie pokazują się zdjęcia:</strong><br>Kliknij <strong>Scrape Images</strong> na stronie palety albo odpal <strong>Auto Pipeline</strong> jeszcze raz.</p>
        <p style="margin-top:12px"><strong>Żółte ostrzeżenie o kolorze na stronie produktu:</strong><br>Amazon pokazuje inny wariant niż dostawca. Wgraj własne zdjęcia w sekcji <strong>Your Photos</strong> — AI użyje opisu od dostawcy jako prawdy.</p>
    </div>
</div>

<div class="card mb-16">
    <div class="card-title">Potrzebujesz pomocy?</div>
    <p style="color:var(--text-muted);font-size:0.85rem">Zadzwoń do Adriana. Ma zdalny dostęp do apki i może naprawić każdy problem.</p>
</div>

</div><!-- /PL -->

<!-- =================================================================== -->
<!-- ENGLISH                                                               -->
<!-- =================================================================== -->
<div class="help-lang" data-lang="en">

<div class="card mb-16">
    <div class="card-title" style="color:#8ff5ff">Quick Start — The Easy Way</div>
    <ol style="line-height:2.2;color:var(--text-muted);font-size:0.9rem">
        <li><strong>Add a pallet:</strong> Go to <strong>Pallets</strong> → <strong>+ Add Pallet</strong> → fill in name, price (GBP), supplier. Upload your CSV/XLSX/ODS file from the joblot supplier. <em>Leave Amazon locale on "Auto-detect" — the scraper finds where each ASIN is listed.</em></li>
        <li><strong>Auto Pipeline:</strong> Click <strong>Auto Pipeline</strong> on the pallet page → the app automatically:
            <ul style="margin:4px 0 4px 20px;line-height:1.8">
                <li>Downloads up to 8 product images from Amazon UK</li>
                <li>Extracts product specifications (brand, model, etc.)</li>
                <li>Generates eBay-optimized titles using AI</li>
                <li>Generates professional HTML descriptions using AI</li>
                <li>Matches the correct eBay category</li>
                <li>Creates draft listings ready for review</li>
            </ul>
        </li>
        <li><strong>Set prices:</strong> Use the <strong>Set Prices</strong> section → type prices manually or click <strong>Apply Multiplier</strong> (e.g. cost × 2.5 = your selling price).</li>
        <li><strong>Review drafts:</strong> Go to <strong>Listings</strong> → check your drafts. Click on a product to edit title, description, or price.</li>
        <li><strong>Publish:</strong> Go back to the pallet → click <strong>Publish All</strong> → listings go LIVE on eBay (you'll be asked to confirm).</li>
        <li><strong>Ship orders:</strong> When something sells → go to <strong>Orders</strong> → click <strong>Mark as Shipped</strong> with the tracking number.</li>
    </ol>
</div>

<div class="card mb-16">
    <div class="card-title" style="color:#8ff5ff">Amazon locale — fully automatic</div>
    <p style="color:var(--text-muted);font-size:0.9rem;line-height:1.8">
        The scraper <strong>auto-detects</strong> which Amazon locale each ASIN from your CSV is listed on.
        It walks through Amazon.co.uk → .com → .de → .pl → .fr → .it → .es → .nl → .se in order and picks the first
        domain where the ASIN resolves to a real product page. You don't need to set anything manually.
    </p>
    <p style="color:var(--text-muted);font-size:0.9rem;line-height:1.8">
        After the first successful scrape, the app <strong>caches the detected locale</strong> on the pallet —
        so future scrapes jump straight to it instead of probing every domain again.
    </p>
    <p style="color:var(--text-muted);font-size:0.85rem;margin-top:8px">
        <strong>It's always in English:</strong> even if the ASIN turns out to be on Amazon.de or .fr, the app
        fetches the page in English (via Amazon's <code>/-/en/</code> URL prefix) — so titles and bullet points
        are always in English, ready for your eBay UK listing.
    </p>
    <p style="color:var(--text-muted);font-size:0.85rem;margin-top:8px">
        <strong>Manual override (optional):</strong> if auto-detect picks the wrong variant (e.g. .co.uk has
        a white Christmas tree under this ASIN but you have the green one from .de) — open the pallet, change
        <em>Amazon Locale</em> to a specific domain (e.g. <code>amazon.de</code>), hit <strong>Save</strong>,
        then click <strong>Scrape Images</strong> to re-fetch. Switch back to "Auto-detect" from the dropdown
        to resume automatic mode.
    </p>
</div>

<div class="card mb-16">
    <div class="card-title" style="color:#beee00">Your own photos — important!</div>
    <p style="color:var(--text-muted);font-size:0.9rem;line-height:1.8">
        Amazon sometimes shows a different variant than what's actually in your pallet
        (e.g. Amazon lists a white Christmas tree, but the pallet contains the green one).
        That's why you can upload <strong>your own photos</strong> of what you really have.
    </p>
    <ol style="line-height:2;color:var(--text-muted);font-size:0.9rem">
        <li>Open the product page (Pallets → pallet → product)</li>
        <li>Section <strong>Your Photos</strong> — click <strong>Choose Files</strong> and pick photos from your phone/disk</li>
        <li>Click <strong>Upload</strong> — your photos will appear <strong>first</strong> in both the carousel and in the eBay listing</li>
        <li>You can delete individual photos — trash icon below each one</li>
    </ol>
    <p style="color:var(--warning);font-size:0.85rem;margin-top:8px">
        <strong>Note:</strong> if the app detects a disagreement between Amazon and the supplier
        (e.g. different colour), you'll see a yellow warning at the top of the product page.
        Upload your own photos — the AI will also trust the supplier description over Amazon.
    </p>
</div>

<div class="card mb-16">
    <div class="card-title" style="color:#f59e0b">Private sale — how to record it</div>
    <p style="color:var(--text-muted);font-size:0.9rem;line-height:1.8">
        If you sold something <strong>outside eBay</strong> (to a friend, at a local market,
        cash in hand) — record it in the app so the revenue on the dashboard is complete.
    </p>
    <ol style="line-height:2;color:var(--text-muted);font-size:0.9rem">
        <li>Open the product page (Pallets → pallet → product)</li>
        <li>Scroll down — section <strong>Sprzedaż prywatna</strong> (above Edit Product)</li>
        <li>Enter the <strong>price</strong> (GBP) — pre-filled with the eBay price, adjust if you sold cheaper</li>
        <li>Optional: who bought it (e.g. "John — friend") and notes (e.g. "paid cash")</li>
        <li>Tick ✅ <strong>Already handed over</strong> if you already gave the item to the buyer</li>
        <li>Click <strong>Record private sale</strong> — confirm the amount</li>
    </ol>
    <p style="color:var(--text-muted);font-size:0.85rem;margin-top:8px">
        <strong>What happens:</strong> the sale appears in <strong>Orders</strong> tagged
        <span class="badge" style="background:rgba(190,238,0,0.15);color:var(--lime);border:1px solid rgba(190,238,0,0.3)">PRIVATE</span>,
        the product is marked as sold, and the amount flows into Total Revenue on the dashboard.
        If you previously set the status to "Sold" manually but never recorded a sale — the
        section still shows with a yellow warning so you can backfill the revenue.
    </p>
</div>

<div class="card mb-16">
    <div class="card-title" style="color:#a855f7">Buttons Explained</div>
    <div style="display:grid;grid-template-columns:170px 1fr;gap:10px 16px;font-size:0.85rem;color:var(--text-muted)">
        <div><span class="btn btn-purple btn-sm" style="font-size:0.7rem">Import CSV</span></div><div>Upload CSV/XLSX/ODS file with products from your joblot supplier</div>
        <div><span class="btn btn-sm" style="font-size:0.7rem;background:linear-gradient(135deg,#8ff5ff,#a855f7);color:#000">Auto Pipeline</span></div><div><strong>Does everything automatically!</strong> Images, titles, descriptions, categories, drafts — one click</div>
        <div><span class="btn btn-outline btn-sm" style="font-size:0.7rem;border-color:#f59e0b;color:#f59e0b">Create Drafts</span></div><div>Save listings locally — nothing goes to eBay yet. Review before publishing</div>
        <div><span class="btn btn-lime btn-sm" style="font-size:0.7rem">Publish All</span></div><div>Send all drafts to eBay — listings go LIVE immediately! <strong>Make sure prices are set first!</strong></div>
        <div><span class="btn btn-outline btn-sm" style="font-size:0.7rem;border-color:#a855f7;color:#a855f7">Auto Categories</span></div><div>Match eBay categories for your products using AI</div>
        <div><span class="btn btn-cyan btn-sm" style="font-size:0.7rem">Scrape Images</span></div><div>Re-download product images from Amazon UK</div>
        <div><span class="btn btn-outline btn-sm" style="font-size:0.7rem">Save Draft</span></div><div>Save title, description, and price without publishing to eBay</div>
        <div><span class="btn btn-lime btn-sm" style="font-size:0.7rem">Publish to eBay</span></div><div>Send this single product to eBay (goes live immediately)</div>
        <div><span class="btn btn-outline btn-sm" style="font-size:0.7rem">Generate Title</span></div><div>AI creates an optimized eBay title (max 80 characters, English)</div>
        <div><span class="btn btn-outline btn-sm" style="font-size:0.7rem">Generate Description</span></div><div>AI creates a professional HTML description with features and bullet points</div>
        <div><span class="btn btn-lime btn-sm" style="font-size:0.7rem">Record private sale</span></div><div>Record a sale made outside eBay — still counts toward revenue</div>
    </div>
</div>

<div class="card mb-16">
    <div class="card-title" style="color:#ff6b9b">Shipping — How It Works</div>
    <p style="color:var(--text-muted);font-size:0.85rem;line-height:1.8">
        When a customer buys your product on eBay:<br><br>
        <strong>1.</strong> You'll see the order in the <strong>Orders</strong> tab with status <strong>"TO SHIP"</strong><br>
        <strong>2.</strong> Pack the item securely<br>
        <strong>3.</strong> Take it to the post office (Royal Mail) or a drop-off point (Evri, DPD)<br>
        <strong>4.</strong> Get a tracking number from the courier<br>
        <strong>5.</strong> In the app, click <strong>Mark as Shipped</strong> and enter the tracking number<br>
        <strong>6.</strong> eBay transfers the money to your bank account (usually 2-3 days after the buyer confirms delivery)<br><br>
        <strong style="color:#8ff5ff">Tip:</strong> You can buy cheaper shipping labels directly from eBay!<br>
        Go to: My eBay → Sold → Print Shipping Label — eBay has discounted rates with Royal Mail and Evri.<br><br>
        <strong style="color:#8ff5ff">Tip:</strong> Set shipping cost to <strong>£0 (free postage)</strong> and include it in the product price — eBay shows free postage listings higher in search results!
    </p>
</div>

<div class="card mb-16">
    <div class="card-title" style="color:#f59e0b">Tips for Better Sales</div>
    <ul style="color:var(--text-muted);font-size:0.85rem;line-height:2.2">
        <li>Use <strong>free postage</strong> (set shipping cost to £0) — eBay ranks free postage listings higher in search</li>
        <li>Use <strong>Auto Pipeline</strong> to generate professional titles and descriptions — saves hours of work</li>
        <li>Add <strong>multiple photos</strong> — listings with 8+ photos get 30% more sales on eBay UK</li>
        <li>Set <strong>competitive prices</strong> — check what similar items sold for on eBay before pricing</li>
        <li>Ship <strong>within 1-2 days</strong> — fast dispatch improves your seller rating and search ranking</li>
        <li>Always provide a <strong>tracking number</strong> — it protects you against "item not received" claims</li>
        <li>Write titles with <strong>keywords buyers search for</strong> (brand, model, key specs)</li>
        <li>For new accounts: eBay may <strong>hold funds for a few days</strong> — this is normal and stops after you build reputation</li>
    </ul>
</div>

<div class="card mb-16">
    <div class="card-title" style="color:#ef4444">Common Problems</div>
    <div style="color:var(--text-muted);font-size:0.85rem;line-height:1.8">
        <p><strong>Error 120 — "Need additional information":</strong><br>Your eBay seller account is not fully set up. Go to <a href="https://www.ebay.co.uk/sh/sell" target="_blank" style="color:#8ff5ff">ebay.co.uk/sh/sell</a> and complete your bank details and identity verification.</p>
        <p style="margin-top:12px"><strong>Error 107 — "Category not valid":</strong><br>The matched category ID is wrong. Click on the product → change the Category field to a valid eBay UK category number (e.g. 96915 for Chargers, 175673 for Computer Parts).</p>
        <p style="margin-top:12px"><strong>Warning — "Funds on hold":</strong><br>This is normal for new sellers. eBay holds money until the buyer confirms delivery. After a few successful sales, this stops.</p>
        <p style="margin-top:12px"><strong>No images showing:</strong><br>Click <strong>Scrape Images</strong> on the pallet page, or run <strong>Auto Pipeline</strong> again.</p>
        <p style="margin-top:12px"><strong>Yellow colour-warning banner on product page:</strong><br>Amazon shows a different variant than the supplier. Upload your own photos in the <strong>Your Photos</strong> section — the AI will use the supplier description as the source of truth.</p>
    </div>
</div>

<div class="card mb-16">
    <div class="card-title">Need Help?</div>
    <p style="color:var(--text-muted);font-size:0.85rem">Contact Adrian for technical support. He can access the app remotely and fix any issues.</p>
</div>

</div><!-- /EN -->
"""

@app.route('/help')
def help_page():
    return render_page(TEMPLATE_HELP, page_title='Help - eBay Hub UK', active_page='help')


# ===================================================================
# Category Matching
# ===================================================================

@app.route('/api/suggest-category', methods=['POST'])
def api_suggest_category():
    """Get eBay category suggestions for a product name."""
    data = request.get_json() or {}
    query = data.get('query', '')
    if not query:
        return jsonify({'ok': False, 'error': 'No query'})

    ebay = get_ebay_client(get_config)
    if not ebay.is_configured() or not ebay.user_token:
        return jsonify({'ok': False, 'error': 'eBay API not configured'})

    categories = ebay.get_suggested_categories(query)
    return jsonify({'ok': True, 'categories': categories})


# ===================================================================
# Auto-Pipeline: Full product processing after CSV import
# ===================================================================

def _gemini_call(api_key, prompt, timeout=15):
    """Call Gemini API and return the text response, or None on failure."""
    try:
        resp = http_requests.post(
            f'https://generativelanguage.googleapis.com/v1beta/models/gemini-2.5-flash-lite:generateContent?key={api_key}',
            json={'contents': [{'parts': [{'text': prompt}]}]},
            timeout=timeout
        )
        result = resp.json()
        if 'candidates' in result and result['candidates']:
            return result['candidates'][0]['content']['parts'][0]['text'].strip()
    except Exception as e:
        print(f"[AutoPipeline] Gemini error: {e}")
    return None


def auto_process_products(pallet_id):
    """
    Auto-pipeline: After CSV import, process each product:
    1. Scrape Amazon for all images + item specifics (auto-detects locale
       unless the pallet has an explicit amazon_domain override)
    2. Generate eBay-optimized title via Gemini
    3. Generate HTML description via Gemini
    4. Match eBay category via Gemini
    5. Create draft listing with all data
    """
    pallet = query_db("SELECT * FROM pallets WHERE id = ?", (pallet_id,), one=True) or {}
    pallet_domain = (pallet.get('amazon_domain') or '').strip() or None
    products = query_db("SELECT * FROM products WHERE pallet_id = ?", (pallet_id,))
    gemini_key = get_config('gemini_api_key', '')

    processed = 0
    drafts_created = 0

    for product in products:
        pid = product['id']
        product_name = product['name']
        all_images = []
        item_specifics = {}
        bullet_points = []

        # --- Step 1: Scrape Amazon (auto-detect or pallet override) ---
        if product['asin']:
            try:
                data = scrape_amazon_product(product['asin'], pallet_domain)
                if data:
                    all_images = data.get('all_images', [])
                    item_specifics = data.get('item_specifics', {})
                    bullet_points = data.get('bullet_points', [])

                    # Cache the first auto-detected locale on the pallet so
                    # subsequent products in the same batch skip the cascade.
                    if not pallet_domain and data.get('source_domain'):
                        pallet_domain = data['source_domain']
                        execute_db(
                            "UPDATE pallets SET amazon_domain = ? WHERE id = ?",
                            (pallet_domain, pallet_id)
                        )

                    images_json = json.dumps(all_images)
                    specs_json = json.dumps(item_specifics)

                    # Update product with images and specs
                    execute_db(
                        "UPDATE products SET images=?, item_specifics=? WHERE id=?",
                        (images_json, specs_json, pid)
                    )

                    # Update main image if we got a better one
                    if data.get('image_url') and not product['image_url']:
                        execute_db("UPDATE products SET image_url=? WHERE id=?",
                                   (data['image_url'], pid))

                    # Update name if Amazon title is better
                    if data.get('title') and len(data['title']) > len(product_name):
                        product_name = data['title']
                        execute_db("UPDATE products SET name=? WHERE id=?",
                                   (product_name, pid))

                    # Update price if we have none
                    if data.get('price') and (product['ebay_price_gbp'] or 0) == 0:
                        execute_db("UPDATE products SET ebay_price_gbp=? WHERE id=?",
                                   (data['price'], pid))

                    processed += 1
            except Exception as e:
                print(f"[AutoPipeline] Scrape error for {product['asin']}: {e}")

        # --- Steps 2-4: Gemini AI (title, description, category) ---
        ebay_title = product_name[:80]
        ebay_description = ''
        ebay_category = product.get('category', '')

        if gemini_key:
            # Build context for AI. Order matters — the supplier description is
            # the ground truth (physically in the pallet) and must outrank Amazon
            # metadata when they disagree on colour/variant.
            supplier_desc = (product.get('supplier_description') or '').strip()
            context_parts = []
            if supplier_desc:
                context_parts.append("Supplier description (ground truth): " + supplier_desc)
            if bullet_points:
                context_parts.append("Features: " + "; ".join(bullet_points[:5]))
            if item_specifics:
                specs_text = ", ".join(f"{k}: {v}" for k, v in list(item_specifics.items())[:10])
                context_parts.append("Specs: " + specs_text)
            context = "\n".join(context_parts) if context_parts else ""

            # --- Step 2: Generate eBay-optimized title ---
            # Hierarchy of truth, strongest to weakest:
            #   1. Supplier description (physical contents of the pallet)
            #   2. Amazon Specs (brand, model, dimensions — usually right)
            #   3. Amazon Product title (often variant-stuffed — last resort only)
            title_prompt = (
                f'Generate a concise eBay UK listing title (max 80 characters) for this product. '
                f'Include brand, model, product type, size, and key functional specs.\n\n'
                f'Data source rules (apply in this order of trust):\n'
                f'1. "Supplier description" is the GROUND TRUTH — it describes what is '
                f'physically in the pallet. If it names a colour/variant/theme, use THAT.\n'
                f'2. "Specs" is from our Amazon scraper — use it for brand, model, size, '
                f'material. If Specs disagree with Supplier description on '
                f'colour/variant/theme, TRUST THE SUPPLIER. Amazon often lists a different '
                f'variant than what the supplier actually shipped.\n'
                f'3. "Product" name is the Amazon title and often bundles multiple variants '
                f'(colour, theme, pattern, seasonal motif) into one keyword-stuffed string. '
                f'Do NOT copy colour/theme/pattern words from it unless they also appear '
                f'in the Supplier description or Specs.\n\n'
                f'No quotes, no special characters. English only.\n\n'
                f'Product: {product_name}\n'
                f'{context}\n\n'
                f'Return ONLY the title, nothing else.'
            )
            ai_title = _gemini_call(gemini_key, title_prompt)
            if ai_title:
                ebay_title = ai_title.strip('"\'')[:80]
            time.sleep(1)

            # --- Step 3: Generate HTML description ---
            # Same hierarchy as the title prompt.
            desc_prompt = (
                f'Generate a professional eBay UK product description in HTML. '
                f'Include: product highlights as bullet points, a key specifications table, '
                f'and a professional closing.\n\n'
                f'Data source rules (apply in this order of trust):\n'
                f'1. "Supplier description" is the GROUND TRUTH — it describes what is '
                f'physically in the pallet. Use its wording for colour/variant/theme '
                f'everywhere they appear in the description.\n'
                f'2. "Specs" is from our Amazon scraper — build the specifications table '
                f'from it. If Specs disagree with Supplier description on colour or '
                f'variant, TRUST THE SUPPLIER and omit or override the conflicting Spec '
                f'rows (Amazon often lists a different variant than what actually shipped).\n'
                f'3. "Product" name is the Amazon title and often bundles multiple variants '
                f'into one keyword-stuffed string. Do NOT copy colour/theme/pattern words '
                f'from it unless they also appear in the Supplier description.\n'
                f'4. Features may come from the "Features:" block if present.\n\n'
                f'Do NOT mention the product condition anywhere — eBay displays it separately. '
                f'Use clean HTML (div, ul, li, p, strong, table tags). '
                f'Do NOT include <html>, <head>, or <body> tags. '
                f'Keep it concise but informative. English only.\n\n'
                f'Product: {product_name}\n'
                f'{context}\n\n'
                f'Return ONLY the HTML description, no markdown, no code blocks.'
            )
            ai_desc = _gemini_call(gemini_key, desc_prompt, timeout=20)
            if ai_desc:
                # Remove markdown code blocks if present
                text = ai_desc
                if text.startswith('```'):
                    text = text.split('\n', 1)[1] if '\n' in text else text
                    if text.endswith('```'):
                        text = text[:-3]
                ebay_description = text
            time.sleep(1)

            # --- Step 4: Match eBay category via Gemini ---
            if not ebay_category:
                cat_prompt = (
                    f'What is the best eBay UK category ID for this product? '
                    f'Use REAL eBay.co.uk category IDs (5 or 6 digit numbers). '
                    f'Return ONLY the category ID and name, format: "ID:Name"\n'
                    f'Common eBay UK categories: 171485:Laptop Accessories, 175673:Computer Components, '
                    f'44985:Cell Phone Accessories, 96915:Laptop Chargers, 20710:Consumer Electronics, '
                    f'11071:Home Furniture, 3270:Action Figures, 11700:Cameras, 64313:Car Parts, '
                    f'15032:Pet Supplies, 11450:Clothing, 11116:Garden Tools, 58058:Smart Home\n\n'
                    f'Product: {product_name[:120]}\n'
                    f'{context}'
                )
                ai_cat = _gemini_call(gemini_key, cat_prompt)
                if ai_cat and ':' in ai_cat:
                    ebay_category = ai_cat.strip('"\'')
                    execute_db("UPDATE products SET category=? WHERE id=?",
                               (ebay_category, pid))
                time.sleep(1)

        # --- Step 5: Create draft listing ---
        # Skip if draft already exists
        existing = query_db(
            "SELECT id FROM ebay_listings WHERE product_id = ? AND status = 'draft'",
            (pid,), one=True
        )
        if not existing:
            # Fallback description if Gemini didn't produce one
            if not ebay_description:
                ebay_description = (
                    '<div style="font-family:Arial,sans-serif">'
                    f'<h2>{ebay_title}</h2>'
                    f'<p>{product_name}</p>'
                    f'<p>Condition: {product.get("condition", "new").replace("_", " ").title()}</p>'
                    '<p>Fast dispatch from UK warehouse.</p>'
                    '</div>'
                )

            price = product['ebay_price_gbp'] or 0.0
            cat_id = ebay_category.split(':')[0] if ebay_category else ''

            execute_db(
                "INSERT INTO ebay_listings (product_id, title, description, price_gbp, "
                "status, category_id, item_specifics) VALUES (?, ?, ?, ?, 'draft', ?, ?)",
                (pid, ebay_title, ebay_description, price,
                 cat_id, json.dumps(item_specifics))
            )
            drafts_created += 1

    return processed, drafts_created


@app.route('/pallet/<int:pallet_id>/auto-pipeline', methods=['POST'])
def pallet_auto_pipeline(pallet_id):
    """Run the full auto-pipeline on all products in a pallet."""
    pallet = query_db("SELECT * FROM pallets WHERE id = ?", (pallet_id,), one=True)
    if not pallet:
        flash('Pallet not found.', 'error')
        return redirect(url_for('pallets_list'))

    processed, drafts = auto_process_products(pallet_id)
    gemini_key = get_config('gemini_api_key', '')

    msg = f'Auto-pipeline complete: {processed} products scraped, {drafts} drafts created.'
    if not gemini_key:
        msg += ' (AI features skipped - set Gemini API key in Settings)'
    flash(msg, 'success')
    return redirect(url_for('pallet_detail', pallet_id=pallet_id))


@app.route('/pallet/<int:pallet_id>/auto-categories', methods=['POST'])
def pallet_auto_categories(pallet_id):
    """Auto-match eBay categories for all products in pallet."""
    pallet = query_db("SELECT * FROM pallets WHERE id = ?", (pallet_id,), one=True)
    if not pallet:
        flash('Pallet not found.', 'error')
        return redirect(url_for('pallets_list'))

    ebay = get_ebay_client(get_config)
    if not ebay.is_configured() or not ebay.user_token:
        flash('eBay API not configured. Go to Settings.', 'error')
        return redirect(url_for('pallet_detail', pallet_id=pallet_id))

    products = query_db(
        "SELECT * FROM products WHERE pallet_id = ? AND name != ''",
        (pallet_id,)
    )

    matched = 0
    gemini_fallback = 0
    for product in products:
        try:
            cats = ebay.get_suggested_categories(product['name'][:80])
            if cats:
                best = cats[0]
                execute_db(
                    "UPDATE products SET category = ? WHERE id = ?",
                    (f"{best['category_id']}:{best['category_name']}", product['id'])
                )
                matched += 1
                continue
        except Exception as e:
            print(f"[Category] eBay API error for product {product['id']}: {e}")

        # Gemini fallback — ask AI to suggest eBay category ID
        try:
            api_key = get_config('gemini_api_key', '')
            if api_key:
                import requests as _req
                resp = _req.post(
                    f'https://generativelanguage.googleapis.com/v1beta/models/gemini-2.5-flash-lite:generateContent?key={api_key}',
                    json={'contents': [{'parts': [{'text':
                        f'What is the best eBay UK category ID for this product? '
                        f'Use REAL eBay.co.uk category IDs (5-6 digit numbers). '
                        f'Return ONLY format: "ID:Name"\n'
                        f'Common: 171485:Laptop Accessories, 175673:Computer Components, '
                        f'44985:Cell Phone Accessories, 96915:Laptop Chargers, 20710:Consumer Electronics\n\n'
                        f'Product: {product["name"][:100]}'
                    }]}]},
                    timeout=10
                )
                result = resp.json()
                if 'candidates' in result and result['candidates']:
                    cat_text = result['candidates'][0]['content']['parts'][0]['text'].strip()
                    if ':' in cat_text:
                        execute_db("UPDATE products SET category = ? WHERE id = ?", (cat_text, product['id']))
                        matched += 1
                        gemini_fallback += 1
        except Exception as e2:
            print(f"[Category] Gemini fallback error: {e2}")
        import time
        time.sleep(0.5)  # Rate limit

    msg = f'Matched categories for {matched}/{len(products)} products.'
    if gemini_fallback:
        msg += f' ({gemini_fallback} via AI fallback)'
    flash(msg, 'success')
    return redirect(url_for('pallet_detail', pallet_id=pallet_id))


# ===================================================================
# AI Generation (Gemini)
# ===================================================================

@app.route('/api/generate-title', methods=['POST'])
def api_generate_title():
    """Generate eBay-optimized title using Gemini AI."""
    data = request.get_json() or {}
    product_name = data.get('product_name', '')
    product_id = data.get('product_id')

    # Pull supplier description (ground truth about pallet contents) and
    # scraped specs from DB. Supplier description wins on colour/variant
    # disputes — see auto_process_products for the full rationale.
    specs_text = ''
    supplier_block = ''
    if product_id:
        row = query_db(
            "SELECT item_specifics, supplier_description FROM products WHERE id = ?",
            (product_id,), one=True)
        if row:
            supplier_desc = (row.get('supplier_description') or '').strip()
            if supplier_desc:
                supplier_block = 'Supplier description (ground truth): ' + supplier_desc + '\n'
            if row.get('item_specifics'):
                try:
                    sp = json.loads(row['item_specifics'])
                    if isinstance(sp, dict) and sp:
                        specs_text = 'Specs: ' + ', '.join(f'{k}: {v}' for k, v in list(sp.items())[:10])
                except (json.JSONDecodeError, TypeError):
                    pass

    if not product_name:
        return jsonify({'ok': False, 'error': 'No product name'})

    api_key = get_config('gemini_api_key', '')
    if not api_key:
        return jsonify({'ok': False, 'error': 'Gemini API key not set. Go to Settings.'})

    try:
        import requests as _req
        resp = _req.post(
            f'https://generativelanguage.googleapis.com/v1beta/models/gemini-2.5-flash-lite:generateContent?key={api_key}',
            json={
                'contents': [{'parts': [{'text':
                    f'Generate a concise eBay UK listing title (max 80 characters) for this product. '
                    f'Include brand, model, product type, size, and key functional specs.\n\n'
                    f'Data source rules (apply in this order of trust):\n'
                    f'1. "Supplier description" is GROUND TRUTH — it describes what is '
                    f'physically in the pallet. Use its colour/variant/theme wording.\n'
                    f'2. "Specs" is from our Amazon scraper. Use for brand/model/size. If Specs '
                    f'disagree with Supplier description on colour or variant, TRUST THE '
                    f'SUPPLIER — Amazon often lists a different variant than what shipped.\n'
                    f'3. "Product" name is the Amazon title and may bundle variant keywords '
                    f'(colour/theme/pattern). Do NOT copy those words unless they appear in '
                    f'the Supplier description or Specs.\n\n'
                    f'No quotes, no special characters. English only.\n\n'
                    f'Product: {product_name}\n'
                    f'{supplier_block}{specs_text}\n\n'
                    f'Return ONLY the title, nothing else.'
                }]}]
            },
            timeout=15
        )
        result = resp.json()
        if 'error' in result:
            return jsonify({'ok': False, 'error': result['error'].get('message', 'Gemini API error')[:100]})
        if 'candidates' not in result or not result['candidates']:
            return jsonify({'ok': False, 'error': 'Gemini returned no results. Check API key.'})
        text = result['candidates'][0]['content']['parts'][0]['text'].strip().strip('"\'')
        return jsonify({'ok': True, 'text': text[:80]})
    except Exception as e:
        print(f"[AI] Generate title error: {e}")
        return jsonify({'ok': False, 'error': str(e)[:100]})


@app.route('/api/generate-description', methods=['POST'])
def api_generate_description():
    """Generate eBay product description using Gemini AI."""
    data = request.get_json() or {}
    product_name = data.get('product_name', '')
    product_id = data.get('product_id')

    # Pull supplier description + scraped specs. Supplier wins on
    # colour/variant — see auto_process_products for the rationale.
    specs_text = ''
    supplier_block = ''
    if product_id:
        row = query_db(
            "SELECT item_specifics, supplier_description FROM products WHERE id = ?",
            (product_id,), one=True)
        if row:
            supplier_desc = (row.get('supplier_description') or '').strip()
            if supplier_desc:
                supplier_block = 'Supplier description (ground truth): ' + supplier_desc + '\n'
            if row.get('item_specifics'):
                try:
                    sp = json.loads(row['item_specifics'])
                    if isinstance(sp, dict) and sp:
                        specs_text = 'Specs: ' + ', '.join(f'{k}: {v}' for k, v in list(sp.items())[:10])
                except (json.JSONDecodeError, TypeError):
                    pass

    if not product_name:
        return jsonify({'ok': False, 'error': 'No product name'})

    api_key = get_config('gemini_api_key', '')
    if not api_key:
        return jsonify({'ok': False, 'error': 'Gemini API key not set. Go to Settings.'})

    try:
        import requests as _req
        resp = _req.post(
            f'https://generativelanguage.googleapis.com/v1beta/models/gemini-2.5-flash-lite:generateContent?key={api_key}',
            json={
                'contents': [{'parts': [{'text':
                    f'Generate a professional eBay UK product description in HTML for this product. '
                    f'Include: key features as bullet points, a specifications table (built from '
                    f'Specs), and a professional closing.\n\n'
                    f'Data source rules (apply in this order of trust):\n'
                    f'1. "Supplier description" is GROUND TRUTH — it describes what is physically '
                    f'in the pallet. Use its colour/variant/theme wording throughout.\n'
                    f'2. "Specs" is from our Amazon scraper — build the specifications table '
                    f'from it. If Specs disagree with Supplier description on colour or variant, '
                    f'TRUST THE SUPPLIER and omit or override conflicting Spec rows. Amazon '
                    f'often lists a different variant than what actually shipped.\n'
                    f'3. "Product" name is the Amazon title and may bundle variant keywords '
                    f'(colour/theme/pattern). Do NOT copy those words unless they appear in the '
                    f'Supplier description.\n\n'
                    f'Do NOT mention the product condition anywhere — eBay displays it separately. '
                    f'Use clean HTML (div, ul, li, p, strong, table tags). Keep it concise but '
                    f'informative. English only. Do NOT include the title.\n\n'
                    f'Product: {product_name}\n'
                    f'{supplier_block}{specs_text}\n\n'
                    f'Return ONLY the HTML description, no markdown, no code blocks.'
                }]}]
            },
            timeout=20
        )
        result = resp.json()
        if 'error' in result:
            return jsonify({'ok': False, 'error': result['error'].get('message', 'Gemini API error')[:100]})
        if 'candidates' not in result or not result['candidates']:
            return jsonify({'ok': False, 'error': 'Gemini returned no results. Check API key.'})
        text = result['candidates'][0]['content']['parts'][0]['text'].strip()
        # Remove markdown code blocks if present
        if text.startswith('```'):
            text = text.split('\n', 1)[1] if '\n' in text else text
            if text.endswith('```'):
                text = text[:-3]
        return jsonify({'ok': True, 'text': text})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)[:100]})


# ===================================================================
# Run
# ===================================================================

if __name__ == '__main__':
    # Start backup scheduler
    try:
        from modules.backup import start_backup_scheduler, create_backup
        start_backup_scheduler()
        create_backup()  # Initial backup on start
    except Exception as e:
        print(f"[WARN] Backup init error: {e}")

    print("=" * 50)
    print("  eBay Hub UK v1.0.0")
    print("  http://127.0.0.1:5002")
    print("=" * 50)
    app.run(debug=False, host='0.0.0.0', port=5002)
